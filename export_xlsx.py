"""
PDD EZ — Excel 导出模块
统一 GUI 和 CLI 两条路径的导出逻辑。
"""
import os
from datetime import datetime
from utils import get_base_dir


_STYLES_CACHE = None


def _create_styles():
    """返回统一的 Excel 样式对象（模块级缓存，避免高频导出重复创建）"""
    global _STYLES_CACHE
    if _STYLES_CACHE is not None:
        return _STYLES_CACHE
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    _STYLES_CACHE = {
        'fills': {
            'red': PatternFill('solid', fgColor='FFC7CE'),
            'yellow': PatternFill('solid', fgColor='FFEB9C'),
            'green': PatternFill('solid', fgColor='C6EFCE'),
        },
        'header_fill': PatternFill('solid', fgColor='4472C4'),
        'header_font': Font(name='微软雅黑', size=9, bold=True, color='FFFFFF'),
        'cell_font': Font(name='微软雅黑', size=9),
        'thin': Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin')),
        'center': Alignment(horizontal='center', vertical='center'),
    }
    return _STYLES_CACHE


def _get_default_export_dir() -> str:
    """默认导出目录：settings → 桌面"""
    import json
    try:
        sf = os.path.join(get_base_dir(), 'settings.json')
        if os.path.exists(sf):
            with open(sf, 'r', encoding='utf-8') as f:
                s = json.load(f)
                return s.get('export_path', os.path.join(os.path.expanduser('~'), 'Desktop'))
    except Exception:
        pass
    return os.path.join(os.path.expanduser('~'), 'Desktop')


def _unique_sheet_name(wb, base: str) -> str:
    """Sheet 重名保护 + Excel 命名规则 sanitize：
    非法字符（\\ / ? * [ ] :）替换为 _，超 31 字符截断，避免 create_sheet 抛异常。
    base 已被占用时追加 _2、_3…"""
    _ILLEGAL = set('\\/?*[]:')
    base = ''.join('_' if ch in _ILLEGAL else ch for ch in str(base)).strip() or 'Sheet'
    base = base[:31]
    existing = {ws.title for ws in wb.worksheets}
    if base not in existing:
        return base
    i = 2
    while True:
        cand = f"{base[:28]}_{i}"  # 保留重名后缀空间：_2 起最多到 _99，仍 ≤31 字符
        if cand not in existing:
            return cand
        i += 1


def _sanitize_cell(v):
    """Excel/CSV 公式注入防护：以 = + - @ 开头的字符串加 ' 前缀，
    避免用户可控文本（商品名）被 Excel 当作公式执行（=WEBSERVICE 外泄等）。
    已是 ' 前缀的不重复加，防止 ''=FORMULA 双层引号。"""
    if isinstance(v, str):
        s = v.strip()
        if s.startswith(('=', '+', '-', '@')) and not s.startswith("'"):
            return "'" + s
    return v




def export_cache_to_xlsx(cache: dict, export_dir: str = None) -> str:
    """
    GUI 路径：按地区分组的 cache → 追加 Sheet 到 PDD补货记录.xlsx
    返回文件路径。
    """
    import openpyxl
    from openpyxl.utils import get_column_letter

    if not export_dir:
        export_dir = _get_default_export_dir()
    os.makedirs(export_dir, exist_ok=True)  # 目录不存在时创建，避免保存崩溃
    path = os.path.join(export_dir, 'PDD补货记录.xlsx')

    ts_date = datetime.now().strftime('%m.%d_%H%M%S')
    styles = _create_styles()

    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        ts_date = _unique_sheet_name(wb, ts_date)
        ws = wb.create_sheet(ts_date)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = ts_date

    # 动态列（v1.3）：取第一个地区 plans 的勾选列；未配置回退默认商品字段
    sel_cols = []
    for _reg, _data in cache.items():
        _pl = (_data or {}).get('plans') or []
        if _pl and _pl[0].get('_sel_cols'):
            sel_cols = list(_pl[0]['_sel_cols'])
            break
    if not sel_cols:
        sel_cols = ['商品信息', '仓库总库存', '仓库预估总销售数']
    headers = ['地区', '仓库'] + list(sel_cols) + ['可售卖天数', '状态', '补货量']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = styles['header_font']
        c.fill = styles['header_fill']
        c.alignment = styles['center']
        c.border = styles['thin']

    row = 2
    for region, data in sorted(cache.items()):
        plans = data.get('plans', [])
        if not plans:
            continue
        for p in plans:
            raw = p.get('_raw') or {}
            vals = [_sanitize_cell(region), _sanitize_cell(p.get('warehouse', ''))]
            from ocr import strip_tail_noise  # 去「查看地址/查看」词条噪音（OCR 识别不稳定，导出层统一清）
            # 名称列用解析后的干净 name（与 GUI 一致，不含 ID:xxx）；其余勾选列用原文
            _name_col = None
            try:
                from utils import get_ocr_columns
                _nm = (get_ocr_columns().get('mapping') or {}).get('name')
                _name_col = _nm if _nm in sel_cols else None
            except Exception:
                _name_col = None
            for col in sel_cols:
                if _name_col and col == _name_col:
                    v = p.get('name', '')
                else:
                    v = raw.get(col)
                    if v is None or v == '':
                        v = p.get(col, '')
                if isinstance(v, str):
                    v = strip_tail_noise(v)
                vals.append(_sanitize_cell(v))
            vals += [p.get('ratio', p.get('days_left', '')), _sanitize_cell(p['status']), p['qty']]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.font = styles['cell_font']
                c.border = styles['thin']
                c.alignment = styles['center']
                if p.get('color') in styles['fills']:
                    c.fill = styles['fills'][p['color']]
            row += 1

    widths = [10, 12] + [20 if '名称' in c or '商品' in c else 12 for c in sel_cols] + [10, 12, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    try:
        wb.save(path)
    except OSError as e:
        # 覆盖磁盘满、路径非法、文件被锁定等常见 IO 错误
        raise OSError(f"无法保存 Excel 文件（{type(e).__name__}）：{e}\n"
                      f"请检查磁盘空间或关闭已打开的 PDD补货记录.xlsx 后重试。\n文件路径: {path}")
    return path




