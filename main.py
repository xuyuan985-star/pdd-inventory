"""
PDD 库存补货排期系统 — 纯标准库
公式：补货时间 = 库存 ÷ 当天销量 - 运输时间
"""

import os, csv, sys
from datetime import datetime, timedelta

def _base_dir():
    if getattr(sys, 'frozen', False):
        data_dir = os.path.join(os.environ.get('APPDATA', os.path.expanduser('~')), 'PDD补货助手')
        os.makedirs(data_dir, exist_ok=True)
        return data_dir
    return os.path.dirname(os.path.abspath(__file__))


def calculate_replenishment(inventory: list, sales: dict,
                            shipping_days: int = 3,
                            min_order: int = 100) -> list:
    today = datetime.now()
    lead_time = shipping_days + 1  # 补货时间 = 运输天数 + 1
    plans = []
    for item in inventory:
        sku = item['sku']
        name = item['name']
        stock = item['stock']
        sd = sales.get(sku, {})
        daily = sd.get('sales', 0) if sd else 0
        daily = max(daily, 1)
        ratio = stock / daily
        reorder_days = ratio - lead_time
        if reorder_days <= 0:
            status = '现在下单'; color = 'red'
            qty = max(daily * 8, min_order)
            qty = ((qty + min_order - 1) // min_order) * min_order
        elif reorder_days <= 2:
            status = f'{reorder_days:.0f}天后下单'; color = 'yellow'
            qty = max(daily * 8, min_order)
            qty = ((qty + min_order - 1) // min_order) * min_order
        else:
            status = f'{reorder_days:.0f}天后下单'; color = 'green'
            qty = 0
        order_date = today if qty > 0 else None
        arrive_date = order_date + timedelta(days=shipping_days) if order_date else None
        plans.append(dict(sku=sku, name=name, stock=stock, daily=daily,
            ratio=round(ratio,1), days_left=round(ratio,1), status=status,
            color=color, qty=qty,
            order_date=order_date.strftime('%Y-%m-%d') if order_date else '-',
            arrive_date=arrive_date.strftime('%Y-%m-%d') if arrive_date else '-'))
    plans.sort(key=lambda p: {'red':0,'yellow':1,'green':2}.get(p['color'],99))
    return plans


def generate_schedule(plans: list) -> list:
    schedule = []
    for p in plans:
        if p['qty'] > 0:
            schedule.append(dict(date=p['order_date'], type='下单', sku=p['sku'], name=p['name'], qty=p['qty']))
            schedule.append(dict(date=p['arrive_date'], type='到货', sku=p['sku'], name=p['name'], qty=p['qty']))
    schedule.sort(key=lambda x: x['date'])
    return schedule


def export_results(plans: list, schedule: list, output_dir: str = None) -> str:
    """导出到桌面 PDD补货记录.xlsx，每次追加新sheet（时间戳命名）"""
    if not output_dir:
        output_dir = os.path.join(os.path.expanduser('~'), 'Desktop')
    path = os.path.join(output_dir, 'PDD补货记录.xlsx')
    try:
        import openpyxl
        return _append_xlsx(plans, path)
    except ImportError:
        ts = datetime.now().strftime('%Y%m%d_%H%M')
        return _export_csv(plans, schedule, output_dir, ts)


def _append_xlsx(plans, path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ts = datetime.now().strftime('%m-%d %H.%M')  # 冒号对Excel非法，用点替代
    fills = {'red': PatternFill('solid', fgColor='FFC7CE'),
             'yellow': PatternFill('solid', fgColor='FFEB9C'),
             'green': PatternFill('solid', fgColor='C6EFCE')}
    cell_font = Font(name='微软雅黑', size=9)
    thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    center = Alignment(horizontal='center', vertical='center')
    header_fill = PatternFill('solid', fgColor='4472C4')
    header_font = Font(name='微软雅黑', size=9, bold=True, color='FFFFFF')

    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        ws = wb.create_sheet(ts)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = ts

    ts_date = ts.split()[0].replace('-', '.')  # 06.15
    headers = ['商品名称', f'库存({ts_date})', '当日销量', '可售卖天数', '补货状态', '建议补货量']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = center; c.border = thin

    for ri, p in enumerate(plans, 2):
        vals = [p['name'], p['stock'], p['daily'],
                p.get('ratio', p.get('days_left', '')), p['status'], p['qty']]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = cell_font; c.border = thin; c.alignment = center
            if p.get('color') in fills:
                c.fill = fills[p['color']]

    widths = [22, 12, 10, 10, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(path)
    return path


def _export_csv(plans, schedule, output_dir, ts):
    path = os.path.join(output_dir, f'补货计划_{ts}.csv')
    with open(path, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.writer(f)
        w.writerow(['商品', '规格', '库存', '销量', '库存÷销量', '状态', '补货量', '下单日', '到货日'])
        for p in plans:
            w.writerow([p['name'], p.get('sku',p['name']), p['stock'], p['daily'],
                        p.get('ratio',p.get('days_left','')), p['status'], p['qty'],
                        p['order_date'], p['arrive_date']])
    return path


def run_pipeline(order_csv: str, inv_csv: str, output_dir: str = None):
    from pdd_import import import_orders, import_inventory
    sales = import_orders(order_csv)
    inventory = import_inventory(inv_csv)
    plans = calculate_replenishment(inventory, sales)
    schedule = generate_schedule(plans)
    if not output_dir:
        output_dir = os.path.join(_base_dir(), 'output')
    path = export_results(plans, schedule, output_dir)
    print(f"\n[DONE] {path}")
    return path


if __name__ == '__main__':
    import sys
    if len(sys.argv) >= 3:
        run_pipeline(sys.argv[1], sys.argv[2])
    else:
        print("PDD库存补货排期系统")
        print("用法: python main.py 订单.csv 库存.csv")
