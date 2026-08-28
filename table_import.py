"""
PDD EZ — 表格导入通道（v1.4.7 商业升级 · WS-B · 任务卡 T-B1 + T-B4 模块侧）

把 CSV / XLSX 读成 ``(headers, rows)``，再喂 ``ocr.parse_items_generic``（强制复用点①，ocr.py:1024）
产出与 OCR 路径同构的 ``items``，并把行级问题汇总成 ``issues`` 供 GUI 显示"导入报告"。

设计铁律（来自 docs/PLAN_commercial_upgrade.md §3.2 T-B1 与 R7）：
1. 只复用、不重写 —— ``ocr.parse_items_generic`` / ``export_xlsx._sanitize_cell`` 全部 import 复用。
2. 失败显式 —— 编码探测失败、行超限、缺商品名列一律 ``ValueError``（§4 失败哲学）。
3. 列映射只做精确匹配 —— ``name`` 永不模糊（与宪法 §1 同纪律）；缺 ``name/stock/sales`` 不静默 fallback。
4. 公式注入防御 —— ``_sanitize_cell`` 由 GUI 在收口侧调用（强制复用点②），本模块 ``write_template``
   写出的示例单元格已脱敏，``import_items`` 输出原值供 GUI 走 sanitize 管线。
5. 程序端零 Tk 依赖 —— 不在导入层碰 GUI；worker 线程可放心调用。

不在本模块范围（由 GUI 集成任务做）：
- 按钮、文件对话框、预览对话框、worker 线程调度、Tk 消息框。
- 导入后调用 ``_sanitize_cell`` 清洗 items 字段名/region/warehouse（GUI 收口侧）。
- 写历史库（WS-A）。

常量：
- MAX_IMPORT_ROWS: 单文件最大行数 1 万；超限直接抛 ``ValueError`` 防冻 UI。
- SUPPORTED_EXT: 支持的扩展名（白名单）。
- LEGACY_EXT: 老格式 .xls 显式拒绝（必须另存为 .xlsx）。
"""
from __future__ import annotations

import csv
import os
import sys

# ocr 内部常量复用 —— 见 R7：必须复用 parse_items_generic，不得重写解析
from ocr import parse_items_generic, normalize_col_name, _parse_num_text
# export_xlsx 样式复用 —— 任务卡 T-B1 写模板要求
from export_xlsx import _create_styles, _sanitize_cell
# utils 列配置复用 —— 导入侧用与 OCR 同一份 mapping，避免"识别/导入两套列名"
from utils import get_ocr_columns


# ============================================================
# 常量与异常
# ============================================================

MAX_IMPORT_ROWS = 10000
"""单次导入最大行数（数据行，不含表头）；超限直接抛 ``ValueError``。"""

SUPPORTED_EXT = ('.csv', '.xlsx')
"""支持的扩展名白名单。"""

LEGACY_EXT = ('.xls',)
"""老格式 .xls —— 显式报错提示另存为 .xlsx，不静默猜测编码。"""

# 与 ocr.map_columns_to_fields 内 _FIELD_ALIASES 保持同步（PLAN §3.2 任务卡明示复用同款别名）。
# 字段别名：用户主名命中时，把字段的全部已知别名也加入匹配集（PDD 后台列名随版本变化）。
_FIELD_ALIASES = {
    'sales': ('仓库预估总销售数', '仓库销售库存', '仓库预估总销量'),
    'stock': ('仓库总库存', '仓库库存'),
}


# ============================================================
# 1. read_table_rows
# ============================================================

def _sniff_delimiter(sample: str) -> str:
    """嗅探 CSV 分隔符；失败回退 ','。Sniffer 在中文+混合分隔符时偶发抛
    ``csv.Error`` 或返回非字符串，统一兜底。"""
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
        delim = dialect.delimiter
        if isinstance(delim, str) and delim:
            return delim
    except Exception:
        pass
    return ','


def _read_csv(path: str) -> tuple:
    """CSV 编码探测：utf-8-sig → gbk → utf-8，全部 strict 失败抛 ``ValueError``。"""
    # 先用二进制读出用于编码探测 + Sniffer 嗅探的样本
    with open(path, 'rb') as f:
        raw = f.read()
    if not raw:
        raise ValueError('CSV 文件为空')

    # 编码探测顺序：utf-8-sig（去 BOM） → gbk（PDD 后台/Excel 中文版默认） → utf-8
    last_err = None
    for enc in ('utf-8-sig', 'gbk', 'utf-8'):
        try:
            text = raw.decode(enc, errors='strict')
        except UnicodeDecodeError as e:
            last_err = e
            continue
        # 探测成功：嗅探分隔符 → 解析为行
        sample = text[:4096]
        delim = _sniff_delimiter(sample)
        # 用 csv 模块的 reader 处理引号/转义
        reader = csv.reader(text.splitlines(), delimiter=delim)
        all_rows = [r for r in reader]
        return all_rows, enc, delim

    # 全部失败 —— 显式提示（§4 失败哲学），绝不静默 fallback
    raise ValueError(
        '无法识别文件编码，请用记事本另存为 UTF-8 后重试'
    ) from last_err


def _read_xlsx(path: str) -> tuple:
    """XLSX：openpyxl read_only + data_only（取计算值非公式串）；非 str 单元格统一 str() 规范化。"""
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError('缺少 openpyxl 依赖，请先 pip install openpyxl') from e

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            raise ValueError('XLSX 文件无有效 Sheet')
        all_rows = []
        for row in ws.iter_rows(values_only=True):
            # 非 str 单元格统一 str() 规范化（datetime/数字 → 字符串；None → ''）
            normalized = [
                ('' if c is None else (c if isinstance(c, str) else str(c)))
                for c in row
            ]
            all_rows.append(normalized)
        return all_rows, None, None
    finally:
        wb.close()


def read_table_rows(path: str) -> tuple:
    """文件 → ``(headers, rows)``。

    - ``headers``: 首个非空行的列表；空列名补 ``'列N'``（N 从 1 起）。
    - ``rows``: ``list[dict]``，每行 ``{列名: 单元格原文str}``，**与 OCR 识别结果同构**，
      可直接喂 ``ocr.parse_items_generic``。

    行为约束：
    - CSV：编码探测 ``utf-8-sig → gbk → utf-8``（全部 strict 失败抛 ``ValueError``），
      分隔符 ``csv.Sniffer`` 嗅探（`,;\t|`）嗅探失败回退 ``,``。
    - XLSX：``openpyxl.load_workbook(read_only=True, data_only=True)``，
      非 str 单元格（datetime/数字）统一 ``str()`` 规范化。
    - .xls：显式抛 ``ValueError`` 提示另存为 .xlsx（不静默猜测）。
    - 数据行超 ``MAX_IMPORT_ROWS`` 抛 ``ValueError``。

    Raises:
        FileNotFoundError: 文件不存在。
        ValueError: 扩展名不支持 / 编码不可识别 / 数据行超限 / XLSX 格式损坏。
    """
    if not os.path.isfile(path):
        raise FileNotFoundError(f'文件不存在: {path}')

    ext = os.path.splitext(path)[1].lower()
    if ext in LEGACY_EXT:
        raise ValueError(
            f'暂不支持 .{ext.lstrip(".")} 老格式，请用 Excel/WPS 打开后另存为 .xlsx 后重试'
        )
    if ext not in SUPPORTED_EXT:
        raise ValueError(
            f'不支持的文件类型: {ext}（支持 {", ".join(SUPPORTED_EXT)}）'
        )

    if ext == '.csv':
        all_rows, _enc, _delim = _read_csv(path)
    else:
        all_rows, _enc, _delim = _read_xlsx(path)

    if not all_rows:
        raise ValueError('文件为空，无任何行')

    # 找到第一个非空行作为表头
    header_idx = None
    for i, r in enumerate(all_rows):
        if r and any(str(c).strip() for c in r):
            header_idx = i
            break
    if header_idx is None:
        raise ValueError('文件无表头行（全部为空）')

    raw_headers = all_rows[header_idx]
    # 空列名补 '列N'（1 起），重复列名去重
    headers = []
    seen = set()
    for j, h in enumerate(raw_headers, 1):
        name = str(h).strip() if h is not None else ''
        if not name:
            name = f'列{j}'
        # 重复列名补 _2、_3…（避免下游 dict 覆盖丢字段）
        base = name
        k = 2
        while name in seen:
            name = f'{base}_{k}'
            k += 1
        seen.add(name)
        headers.append(name)

    # 数据行 = 表头之后的所有行；过滤纯空行
    data_rows_raw = all_rows[header_idx + 1:]
    data_rows = []
    for r in data_rows_raw:
        # 行内全部为空（仅含空字符串/None）→ 跳过
        if not any(str(c).strip() for c in r if c is not None):
            continue
        # 对齐列宽：列数不足补空字符串；超过截断（防御性，正常 XLSX 不会）
        if len(r) < len(headers):
            r = list(r) + [''] * (len(headers) - len(r))
        elif len(r) > len(headers):
            r = list(r)[:len(headers)]
        data_rows.append({headers[j]: ('' if r[j] is None else str(r[j])) for j in range(len(headers))})

    if len(data_rows) > MAX_IMPORT_ROWS:
        raise ValueError(
            f'数据行 {len(data_rows)} 超过上限 {MAX_IMPORT_ROWS}，请拆分文件或减少数据量后重试'
        )

    return headers, data_rows


# ============================================================
# 2. guess_mapping
# ============================================================

def _alias_targets(field: str, mapping: dict) -> set:
    """返回某字段在匹配时应被认作"自己的"所有列名（含主名 + 同款别名）。"""
    targets = set()
    main = (mapping or {}).get(field) or ''
    if main:
        targets.add(normalize_col_name(main))
    for alias in _FIELD_ALIASES.get(field, ()):
        targets.add(normalize_col_name(alias))
    return targets


def guess_mapping(headers: list) -> tuple:
    """把文件表头按 ``get_ocr_columns().mapping`` 值做精确匹配（normalize 后）。

    Args:
        headers: ``read_table_rows`` 返回的表头列表。

    Returns:
        ``({field: 实际列名}, 缺失字段清单)``。

    行为约束（PLAN §3.2 T-B1 强制）：
    - 仅精确匹配（normalize 后字符串相等），**不模糊**（宪法 §1：name 永不模糊）。
    - 当某字段主名未命中时，回退同款 ``_FIELD_ALIASES`` 别名（与 ``ocr.map_columns_to_fields`` 同语义）。
    - 缺 ``name/stock/sales`` 任一字段不静默 fallback —— 把缺失字段加入返回的 ``missing`` 列表，
      由 GUI 弹"映射预览"对话框让用户调整（v1.3 旧 mapping bug 教训）。
    - 缺 ``region/warehouse`` 是合法的（导入商品无区域/仓库时走 GUI 回退）。
    """
    mapping = get_ocr_columns().get('mapping') or {}
    norm_headers = {h: normalize_col_name(h) for h in headers}

    found = {}
    for field in ('name', 'stock', 'sales', 'region', 'warehouse'):
        targets = _alias_targets(field, mapping)
        hit = None
        for orig, norm in norm_headers.items():
            if norm in targets:
                hit = orig
                break
        if hit:
            found[field] = hit

    # 缺 name/stock/sales 视为必须解决的缺失
    missing = [f for f in ('name', 'stock', 'sales') if f not in found]
    return found, missing


# ============================================================
# 3. import_items
# ============================================================

def _truncate_name(s: str, limit: int = 40) -> str:
    """issues['name'] 用：截断到 limit 字符，方便 Treeview 展示。"""
    s = (s or '').strip()
    return s if len(s) <= limit else s[:limit - 1] + '…'


def import_items(path: str, mapping: dict = None) -> tuple:
    """主入口：``read_table_rows`` → ``guess_mapping`` → ``ocr.parse_items_generic`` → ``(items, issues)``。

    Args:
        path: CSV 或 XLSX 文件路径。
        mapping: 可选外部传入的 mapping 字典 ``{field: 列名}``；为 ``None`` 时
            用 ``guess_mapping`` 自动推断（但发现 ``name/stock/sales`` 缺失时仍抛
            ``ValueError``，不静默 fallback —— 由调用方在预览阶段补全后再传回）。

    Returns:
        ``(items, issues)``：
        - ``items``：同构于 OCR 识别结果，每条 ``{name, stock(int), sales(int),
          region, warehouse, sku_id, _raw, _missing_id?, _low_conf_col?}``，
          可直接喂既有 ``_calc_from_items`` 主干。
        - ``issues``：``list[dict]``，每条 ``{row, name, level, reason}``。
            - ``level='error'``：缺商品名行（OCR 同款语义：商品名为空整行丢弃）。
            - ``level='warning'``：stock/sales 单元格非空但 ``_parse_num_text`` 解析为 0
              （OCR 解析失败静默返 0 是 OCR 语义，导入通道必须显式暴露，§4 失败哲学）。

    Raises:
        FileNotFoundError: 文件不存在。
        ValueError: 扩展名不支持 / 编码不可识别 / 数据行超限 / mapping 缺 name|stock|sales。
    """
    headers, rows = read_table_rows(path)

    if mapping is None:
        mapping, missing = guess_mapping(headers)
        if missing:
            raise ValueError(
                f'导入文件列映射不完整，缺关键字段: {", ".join(missing)}。'
                f'请确认文件含"商品名称/库存/销量"列，或在映射预览对话框调整后重试。'
            )
    else:
        # 外部传入的 mapping 也要校验 name|stock|sales 三个核心字段
        for k in ('name', 'stock', 'sales'):
            if not mapping.get(k):
                raise ValueError(
                    f'外部 mapping 缺关键字段: {k}。name/stock/sales 必填，'
                    f'可由 guess_mapping 自动推断或人工指定。'
                )

    # === 强制复用：rows 直接喂 ocr.parse_items_generic（PLAN §3.2 R7 强制验收项） ===
    # parse_items_generic 内部已经完成：
    #   - map_columns_to_fields（按 mapping 归一化/别名/编辑距离匹配 + _split_name_id 拆 name+sku_id）
    #   - _parse_num_text 数字解析
    #   - strip_region_suffix / strip_warehouse_noise 噪音剥离
    #   - 无商品名行自动跳过
    items = parse_items_generic(rows, mapping)

    # === 跟踪行级 issues（必须自己记行号，parse_items_generic 不返回行号） ===
    # 直接基于"原文件行的 name 列单元格"判定 error —— 不依赖 parse 后的反向匹配（脆弱）。
    issues = []
    name_col = mapping.get('name') or ''
    stock_col = mapping.get('stock') or ''
    sales_col = mapping.get('sales') or ''

    for data_idx, row in enumerate(rows, start=1):
        name_v = str(row.get(name_col, '') or '').strip() if name_col else ''
        stock_raw = str(row.get(stock_col, '') or '').strip() if stock_col else ''
        sales_raw = str(row.get(sales_col, '') or '').strip() if sales_col else ''

        # 1) error：缺商品名行（name 列为空 / 全空白）
        if not name_v:
            issues.append({
                'row': data_idx,
                'name': '',
                'level': 'error',
                'reason': '缺商品名（name 列为空）',
            })
            continue

        # 2) warning：stock/sales 单元格非空但 _parse_num_text 解析为 0
        # parse_items_generic 内部 _parse_num_text 静默返 0，导入侧必须显式暴露（§4 失败哲学）
        if stock_raw and _parse_num_text(stock_raw) == 0:
            issues.append({
                'row': data_idx,
                'name': _truncate_name(name_v),
                'level': 'warning',
                'reason': f'库存单元格无法解析为数字（原文: "{stock_raw}"）',
            })
        if sales_raw and _parse_num_text(sales_raw) == 0:
            issues.append({
                'row': data_idx,
                'name': _truncate_name(name_v),
                'level': 'warning',
                'reason': f'销量单元格无法解析为数字（原文: "{sales_raw}"）',
            })

    return items, issues


# ============================================================
# 4. write_template
# ============================================================

def _format_field_examples() -> dict:
    """模板说明页：每个业务字段支持的原文格式示例（与 _parse_num_text 语义对齐）。"""
    return {
        'name': '示例商品A500g/袋 ID:12345678901（含商品 ID 时自动拆为商品名 + sku_id）',
        'stock': '500 / 1,234 / 1.2万 / 5k / 100份 查看',
        'sales': '300 / 2,500 / 1.5w / 8k / 共 200',
        'region': '广东省广州市 / 浙江（自动去"省/市"等尾缀）',
        'warehouse': '广州仓 / 华东中心仓（自动去"查看地址"等噪音）',
    }


def _build_template_rows(mapping: dict) -> tuple:
    """构造模板数据样例：表头 + 2 行示例。"""
    headers = [mapping.get(k) or k for k in ('name', 'stock', 'sales', 'region', 'warehouse')]
    row1 = [
        '示例商品A500g/袋 ID:12345678901',
        '1,200',
        '1.5w',
        '广东省广州市',
        '广州中心仓',
    ]
    row2 = [
        '示例商品B300g/盒 ID:98765432100',
        '500',
        '300',
        '浙江省杭州市',
        '华东中心仓',
    ]
    return headers, [row1, row2]


def _build_template_legend(mapping: dict) -> tuple:
    """构造 Sheet2 列名说明：每列 = 业务字段、含义、列名、支持格式示例。"""
    headers = ['业务字段', '含义', '列名（按 settings 配置）', '支持格式示例']
    examples = _format_field_examples()
    rows = []
    for field, meaning in (
        ('name', '商品名（必填）'),
        ('stock', '仓库库存（必填，数字）'),
        ('sales', '仓库预估销量（必填，数字）'),
        ('region', '销售区域（可空，缺省归入当前地区）'),
        ('warehouse', '仓库信息（可空）'),
    ):
        rows.append([
            field,
            meaning,
            mapping.get(field, ''),
            examples.get(field, ''),
        ])
    return headers, rows


def write_template(path: str) -> str:
    """按当前 ``get_ocr_columns().mapping`` 运行时生成模板 xlsx。

    Args:
        path: 输出 .xlsx 文件路径（**不**随包分发 —— 由调用方放到
            ``get_base_dir()/templates/`` 或用户指定目录）。

    Returns:
        写入的文件路径（与入参相同，便于链式调用）。

    Sheet 结构：
    - Sheet1 "数据样例"：表头 = mapping 字段对应列名 + 2 行示例。
      样式复用 ``export_xlsx._create_styles``（表头蓝底白字 + 边框 + 居中）。
    - Sheet2 "列名说明"：业务字段 / 含义 / 列名 / 支持格式示例。

    Raises:
        RuntimeError: openpyxl 缺失。
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise RuntimeError('缺少 openpyxl 依赖，请先 pip install openpyxl') from e

    mapping = get_ocr_columns().get('mapping') or {}
    styles = _create_styles()

    wb = openpyxl.Workbook()

    # --- Sheet1: 数据样例 ---
    ws1 = wb.active
    ws1.title = '数据样例'
    headers1, example_rows = _build_template_rows(mapping)
    for ci, h in enumerate(headers1, 1):
        c = ws1.cell(row=1, column=ci, value=h)
        c.font = styles['header_font']
        c.fill = styles['header_fill']
        c.alignment = styles['center']
        c.border = styles['thin']
    for ri, row in enumerate(example_rows, 2):
        for ci, v in enumerate(row, 1):
            c = ws1.cell(row=ri, column=ci, value=_sanitize_cell(v))
            c.font = styles['cell_font']
            c.border = styles['thin']
            c.alignment = styles['center']
    # 列宽
    for ci, _ in enumerate(headers1, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = 22

    # --- Sheet2: 列名说明 ---
    ws2 = wb.create_sheet('列名说明')
    headers2, legend_rows = _build_template_legend(mapping)
    for ci, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = styles['header_font']
        c.fill = styles['header_fill']
        c.alignment = styles['center']
        c.border = styles['thin']
    for ri, row in enumerate(legend_rows, 2):
        for ci, v in enumerate(row, 1):
            c = ws2.cell(row=ri, column=ci, value=_sanitize_cell(v))
            c.font = styles['cell_font']
            c.border = styles['thin']
            c.alignment = styles['center']
    # 列宽
    ws2.column_dimensions['A'].width = 10
    ws2.column_dimensions['B'].width = 28
    ws2.column_dimensions['C'].width = 24
    ws2.column_dimensions['D'].width = 50

    # 确保目录存在
    out_dir = os.path.dirname(os.path.abspath(path))
    if out_dir and not os.path.isdir(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    wb.save(path)
    return path


# ============================================================
# 模块自检（python -m table_import 可直接跑）
# ============================================================

def _selftest():
    """最小自检：import 自检用，避免在测试文件外加 helper。"""
    import tempfile
    import shutil
    tmp = tempfile.mkdtemp()
    try:
        # 写 GBK CSV（含 5 列齐全 + 一行缺商品名）
        csv_gbk = os.path.join(tmp, 'in_gbk.csv')
        with open(csv_gbk, 'wb') as f:
            f.write('商品信息,仓库总库存,仓库预估总销售数,销售区域,仓库信息\n'.encode('gbk'))
            f.write('示例商品A,1,200,广东省广州市,广州中心仓\n'.encode('gbk'))
            f.write('示例商品B ID:12345678,500,300,浙江省,华东中心仓\n'.encode('gbk'))
            f.write(',100,50,江苏,南京仓\n'.encode('gbk'))  # 缺商品名行

        headers, rows = read_table_rows(csv_gbk)
        items, issues = import_items(csv_gbk)
        assert len(items) == 2, f'GBK 导入应 2 条，实际 {len(items)}'
        assert items[0]['name'].startswith('示例商品A'), f'商品名解析错: {items[0]!r}'
        assert items[0]['stock'] == 1, f'stock 解析错: {items[0]!r}'
        assert items[0]['sales'] == 200, f'sales 解析错: {items[0]!r}'
        assert items[1].get('sku_id') == '12345678', f'sku_id 拆分错: {items[1]!r}'

        # 应有 1 条 error issues（行 3 缺商品名）
        err_issues = [i for i in issues if i['level'] == 'error']
        assert len(err_issues) == 1, f'应恰好 1 条 error issues，实际 {len(err_issues)}: {issues!r}'
        assert err_issues[0]['row'] == 3, f'error 行号应=3，实际 {err_issues[0]["row"]}'

        # 写 XLSX 模板
        tpl_path = os.path.join(tmp, 'tpl.xlsx')
        write_template(tpl_path)
        assert os.path.isfile(tpl_path)
        from openpyxl import load_workbook
        wb = load_workbook(tpl_path, read_only=True, data_only=True)
        assert '数据样例' in wb.sheetnames
        assert '列名说明' in wb.sheetnames
        wb.close()

        print('OK: table_import selftest passed')
        return 0
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


if __name__ == '__main__':
    sys.exit(_selftest())
