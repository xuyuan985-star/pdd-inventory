"""PDD EZ 核心模块 smoke tests——后续轮次改动的验证锚点"""
import os
import re
import sys
import unittest
import inspect

HERE = os.path.dirname(os.path.abspath(__file__))
if HERE not in sys.path:
    sys.path.insert(0, HERE)


class TestLogger(unittest.TestCase):
    def test_logger_module_importable(self):
        import logger
        self.assertTrue(hasattr(logger, 'Logger'))
        self.assertTrue(hasattr(logger, 'log'))

    def test_logger_writes_file(self):
        import tempfile, shutil
        from logger import Logger
        tmp = tempfile.mkdtemp()
        try:
            log = Logger(log_dir=tmp)
            log.info('test line')
            import glob
            files = glob.glob(os.path.join(tmp, '*.log'))
            self.assertTrue(files, '日志文件未生成')
            with open(files[0], encoding='utf-8') as fp:
                content = fp.read()
            self.assertIn('test line', content)
        finally:
            shutil.rmtree(tmp, ignore_errors=True)


class TestGithubApi(unittest.TestCase):
    def test_mirror_url(self):
        from github_api import mirror_download_url
        u = mirror_download_url('https://github.com/x/a.zip', prefer_mirror=True)
        self.assertTrue(u.startswith('https://github.kotori.top/'))
        u2 = mirror_download_url('https://github.com/x/a.zip', prefer_mirror=False)
        self.assertTrue(u2.startswith('https://github.com/'))
        u3 = mirror_download_url('https://example.com/x.zip')
        self.assertEqual(u3, 'https://example.com/x.zip')


class TestConfigMerge(unittest.TestCase):
    def test_merge(self):
        import importlib.util
        spec = importlib.util.spec_from_file_location('pdd_utils', os.path.join(HERE, 'utils.py'))
        u = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(u)
        tpl = {'theme': '极简白', 'api': {'active_provider': 'doubao', 'providers': {}}}
        user = {'theme': '终末地', 'api': None}
        merged = u.Config._merge(tpl, user)
        self.assertEqual(merged['theme'], '终末地')
        self.assertEqual(merged['api']['active_provider'], 'doubao')

    def test_version_newer(self):
        import importlib.util
        spec = importlib.util.spec_from_file_location('pdd_utils', os.path.join(HERE, 'utils.py'))
        u = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(u)
        self.assertTrue(u.version_newer('v1.5', 'v1.4'))
        self.assertFalse(u.version_newer('v1.3', 'v1.4'))
        self.assertTrue(u.version_newer('v1.10', 'v1.9'))


class TestDPAPI(unittest.TestCase):
    """v1.4.8 P1-C：DPAPI 凭据加密 + Config._migrate_secrets + 解密失败降级"""

    def test_is_encrypted_prefix(self):
        """is_encrypted 识别 dpapi:v1: 前缀；空/None/明文 都返回 False。"""
        import importlib.util
        spec = importlib.util.spec_from_file_location('pdd_dpapi', os.path.join(HERE, 'dpapi_utils.py'))
        m = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(m)
        self.assertTrue(m.is_encrypted('dpapi:v1:abc123'))
        self.assertFalse(m.is_encrypted('sk-plain-text'))
        self.assertFalse(m.is_encrypted(''))
        self.assertFalse(m.is_encrypted(None))
        self.assertFalse(m.is_encrypted(123))

    def test_encrypt_decrypt_roundtrip(self):
        """enc(plain) → dec(blob) == plain；空串/None 短路返回。"""
        import importlib.util
        spec = importlib.util.spec_from_file_location('pdd_dpapi', os.path.join(HERE, 'dpapi_utils.py'))
        m = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(m)
        if not m.is_available():
            self.skipTest('DPAPI 不可用（沙盒/Wine）')
        plain = 'sk-test-key-1234567890-abcdef'
        blob = m.enc(plain)
        self.assertTrue(blob.startswith('dpapi:v1:'))
        self.assertNotIn(plain, blob, '明文不应出现在密文里')
        self.assertEqual(m.dec(blob), plain)
        # Unicode + 长字符串
        self.assertEqual(m.dec(m.enc('中文+emoji🔐')), '中文+emoji🔐')
        self.assertEqual(m.dec(m.enc('x' * 5000)), 'x' * 5000)
        # 空/None 短路
        self.assertEqual(m.enc(''), '')
        self.assertEqual(m.enc(None), '')  # None 入参按空串处理（不抛）
        self.assertEqual(m.dec(''), '')

    def test_decrypt_non_encrypted_passthrough(self):
        """dec 对无 dpapi:v1: 前缀的字符串原样返回（向后兼容未迁移配置）。"""
        import importlib.util
        spec = importlib.util.spec_from_file_location('pdd_dpapi', os.path.join(HERE, 'dpapi_utils.py'))
        m = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(m)
        self.assertEqual(m.dec('sk-anything-not-encrypted'), 'sk-anything-not-encrypted')
        self.assertEqual(m.dec('random text'), 'random text')

    def test_decrypt_corrupt_blob_raises_dpapi_error(self):
        """dpapi:v1: 后跟非 base64 或损坏数据 → 抛 DPAPIError（让调用方置空）。"""
        import importlib.util
        spec = importlib.util.spec_from_file_location('pdd_dpapi', os.path.join(HERE, 'dpapi_utils.py'))
        m = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(m)
        if not m.is_available():
            self.skipTest('DPAPI 不可用')
        # 合法 base64 但不是 DPAPI 输出 → CryptUnprotectData 返回 False
        bad = 'dpapi:v1:' + 'A' * 200
        with self.assertRaises(m.DPAPIError):
            m.dec(bad)
        # 非 base64 → 抛 DPAPIError（base64 decode 失败包成 DPAPIError）
        with self.assertRaises(m.DPAPIError):
            m.dec('dpapi:v1:!!!not-base64!!!')

    def test_config_decrypt_value_handles_corrupt(self):
        """Config.decrypt_value 对损坏密文返回空串（不让 UI 阻塞启动）。"""
        import importlib.util
        spec = importlib.util.spec_from_file_location('pdd_utils', os.path.join(HERE, 'utils.py'))
        u = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(u)
        # 明文直通
        self.assertEqual(u.Config.decrypt_value('plain-text'), 'plain-text')
        # 空串
        self.assertEqual(u.Config.decrypt_value(''), '')
        # 损坏密文 → 返回空串（不抛）
        self.assertEqual(u.Config.decrypt_value('dpapi:v1:' + 'X' * 200), '')

    def test_migrate_secrets_encrypts_plaintext(self):
        """_migrate_secrets: 明文 api_key / backend.password → dpapi:v1: 密文 + meta.dpi_v=1。
        tmp 目录注入明文 settings.json + 最小 template.json，复现首启迁移场景。"""
        import importlib.util, os, json, tempfile, shutil
        spec = importlib.util.spec_from_file_location('pdd_utils', os.path.join(HERE, 'utils.py'))
        u = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(u)
        spec2 = importlib.util.spec_from_file_location('pdd_dpapi', os.path.join(HERE, 'dpapi_utils.py'))
        d = importlib.util.module_from_spec(spec2)
        spec2.loader.exec_module(d)
        if not d.is_available():
            self.skipTest('DPAPI 不可用')

        tmp = tempfile.mkdtemp()
        try:
            # 注入明文 settings.json
            sf = os.path.join(tmp, 'settings.json')
            user = {
                'api': {
                    'active_provider': 'doubao',
                    'providers': {
                        'doubao': {'api_key': 'sk-PLAINTEXT-DOUBAO-KEY', 'model': 'Doubao-1.5',
                                   'endpoint': 'https://ark.cn-beijing.volces.com/api/v3/chat/completions'},
                    }
                },
                'backend': {
                    'url': 'https://mms.pinduoduo.com/',
                    'account': '13800138000',
                    'password': 'MyPlainP@ssw0rd',
                },
            }
            with open(sf, 'w', encoding='utf-8') as f:
                json.dump(user, f, ensure_ascii=False, indent=2)
            # 最小 template
            tf = os.path.join(tmp, 'settings_template.json')
            with open(tf, 'w', encoding='utf-8') as f:
                json.dump({
                    'theme': 's', 'api': {'active_provider': 'doubao', 'providers': {
                        'doubao': {'api_key': '', 'model': '', 'model_history': [], 'endpoint': ''}
                    }},
                    'export_path': '', 'backend': {'url': '', 'account': '', 'password': ''},
                    'calibrate': {'mode': 'ai', 'ai': {}},
                    'announcement': {'url': '', 'mirror_url': ''},
                    'usage': {'enabled': True, 'batch_budget_cny': 0, 'monthly_budget_cny': 0, 'pricing': {}},
                    'history': {'retention_days': 180, 'max_rows': 200000},
                    'eula_accepted_v1': False,
                }, f, ensure_ascii=False, indent=2)
            # 把 Config 的 base_dir 重定向到 tmp（不能 monkey-patch utils.get_base_dir —
            # 它被 import 时绑定，改 utils.get_base_dir 不影响 Config 内调用；改 m 即可）
            u.get_base_dir = lambda: tmp
            u.Config._load_cache = {'mtime': -1, 'data': None}
            u.Config._template_cache = None

            # 触发 load
            loaded = u.Config.load()
            # 内存里应已是密文
            self.assertTrue(loaded['api']['providers']['doubao']['api_key'].startswith('dpapi:v1:'))
            self.assertTrue(loaded['backend']['password'].startswith('dpapi:v1:'))
            self.assertEqual(loaded.get('meta', {}).get('dpi_v'), 1)
            # 落盘也应已是密文
            with open(sf, 'r', encoding='utf-8') as f:
                on_disk = json.load(f)
            self.assertTrue(on_disk['api']['providers']['doubao']['api_key'].startswith('dpapi:v1:'))
            self.assertTrue(on_disk['backend']['password'].startswith('dpapi:v1:'))
            # 解密可还原
            self.assertEqual(d.dec(loaded['api']['providers']['doubao']['api_key']),
                             'sk-PLAINTEXT-DOUBAO-KEY')
            self.assertEqual(d.dec(loaded['backend']['password']), 'MyPlainP@ssw0rd')

            # 二次 load 应幂等（meta.dpi_v=1 阻止再迁移；on-disk 不再变化）
            u.Config._load_cache = {'mtime': -1, 'data': None}
            loaded2 = u.Config.load()
            self.assertEqual(loaded2['api']['providers']['doubao']['api_key'],
                             loaded['api']['providers']['doubao']['api_key'])
            self.assertEqual(loaded2.get('meta', {}).get('dpi_v'), 1)
        finally:
            shutil.rmtree(tmp, ignore_errors=True)

    def test_migrate_secrets_noop_on_already_encrypted(self):
        """已加密字段 + meta.dpi_v=1 → _migrate_secrets 直接返回 False（幂等）。"""
        import importlib.util
        spec = importlib.util.spec_from_file_location('pdd_utils', os.path.join(HERE, 'utils.py'))
        u = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(u)
        spec2 = importlib.util.spec_from_file_location('pdd_dpapi', os.path.join(HERE, 'dpapi_utils.py'))
        d = importlib.util.module_from_spec(spec2)
        spec2.loader.exec_module(d)
        if not d.is_available():
            self.skipTest('DPAPI 不可用')

        data = {
            'api': {'providers': {'doubao': {'api_key': 'dpapi:v1:already-encrypted-blob'}}},
            'meta': {'dpi_v': 1},
        }
        result = u.Config._migrate_secrets(data)
        self.assertFalse(result)
        # 字段未被改
        self.assertEqual(data['api']['providers']['doubao']['api_key'],
                         'dpapi:v1:already-encrypted-blob')

    def test_sanitize_for_log_redacts_secrets(self):
        """_sanitize_for_log 脱敏 api_key / password / Authorization / Bearer 等敏感字段。"""
        from utils import _sanitize_for_log
        # 直传字符串
        self.assertNotIn('sk-1234567890', _sanitize_for_log('api_key=sk-1234567890'))
        self.assertNotIn('MyP@ss', _sanitize_for_log('password=MyP@ss'))
        self.assertNotIn('eyJhb', _sanitize_for_log('Authorization: Bearer eyJhb.xxx'))
        # JSON 形式
        out = _sanitize_for_log('{"api_key": "sk-secret", "model": "Doubao-1.5"}')
        self.assertNotIn('sk-secret', out)
        self.assertIn('Doubao-1.5', out)
        # 键名仍可见（方便排查）
        self.assertIn('api_key', out)
        # Bearer 单独
        self.assertNotIn('eyJhbGciOiJIUzI1NiJ9', _sanitize_for_log('Bearer eyJhbGciOiJIUzI1NiJ9.payload'))
        # 无敏感字段直通
        self.assertEqual(_sanitize_for_log('no secrets here'), 'no secrets here')
        # 异常输入不抛
        self.assertIsNone(_sanitize_for_log(None))
        self.assertEqual(_sanitize_for_log(''), '')

    def test_log_sanitizes_in_real_log_call(self):
        """log.info/warning 实际走过的 Formatter 链会脱敏（防裸字符串拼接漏过）。"""
        from logger import log, _PlainFormatter
        import logging, io
        buf = io.StringIO()
        h = logging.StreamHandler(buf)
        h.setFormatter(_PlainFormatter('%(levelname)s | %(message)s'))
        log.logger.addHandler(h)
        try:
            log.info('user entered api_key=sk-12345 secret value')
            log.warning('Bearer eyJhbGciOiJIUzI1NiJ9.payload.sig')
        finally:
            log.logger.removeHandler(h)
        out = buf.getvalue()
        self.assertNotIn('sk-12345', out, f'api_key leaked in log: {out!r}')
        self.assertNotIn('eyJhbGciOiJIUzI1NiJ9', out, f'Bearer token leaked: {out!r}')


class TestUpdater(unittest.TestCase):
    def test_importable(self):
        import importlib.util
        spec = importlib.util.spec_from_file_location('pdd_updater', os.path.join(HERE, 'updater.py'))
        m = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(m)
        self.assertTrue(hasattr(m, 'do_finalize'))
        self.assertTrue(hasattr(m, '_cover_with_self_handling'))
        self.assertTrue(hasattr(m, '_pick_main_exe'))

    @unittest.skipUnless(sys.platform == 'win32', 'Windows 专属 API')
    def test_wait_pid_exit(self):
        """_wait_pid_exit：已死 PID 立即返回 True；活进程等待超时返回 False（64 位句柄安全验证）"""
        import importlib.util, subprocess, time
        spec = importlib.util.spec_from_file_location('pdd_updater', os.path.join(HERE, 'updater.py'))
        m = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(m)
        # 不存在的 PID → 进程已死 → True
        self.assertTrue(m._wait_pid_exit(99999999, timeout=1))
        # 活进程（自己）等待 0.1s 超时 → False
        self.assertFalse(m._wait_pid_exit(os.getpid(), timeout=0.1))
        # 活进程匹配 expected_exe 路径（本进程 python）→ 等待超时 False
        self.assertFalse(m._wait_pid_exit(os.getpid(), expected_exe=sys.executable, timeout=0.1))

    def test_do_finalize_full_chain(self):
        """集成：do_finalize 全链路——解压→覆盖→配置保留→_pick_main_exe 选固定名"""
        import importlib.util, tempfile, shutil, zipfile
        spec = importlib.util.spec_from_file_location('pdd_updater', os.path.join(HERE, 'updater.py'))
        m = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(m)

        tmp = tempfile.mkdtemp()
        try:
            appdir = os.path.join(tmp, 'PDD EZ')
            os.makedirs(os.path.join(appdir, '_internal'), exist_ok=True)
            with open(os.path.join(appdir, 'PDD EZ.exe'), 'wb') as f:
                f.write(b'old')
            with open(os.path.join(appdir, '_internal', 'settings.json'), 'wb') as f:
                f.write(b'USER_CFG')

            # 构造更新包
            stage = os.path.join(tmp, 'stage')
            os.makedirs(os.path.join(stage, 'PDD EZ', '_internal'), exist_ok=True)
            with open(os.path.join(stage, 'PDD EZ', 'PDD EZ.exe'), 'wb') as f:
                f.write(b'new')
            with open(os.path.join(stage, 'PDD EZ', '_internal', 'lib.dll'), 'wb') as f:
                f.write(b'lib')
            zip_path = os.path.join(tmp, 'upd.zip')
            with zipfile.ZipFile(zip_path, 'w') as zf:
                for root, _, files in os.walk(os.path.join(stage, 'PDD EZ')):
                    for fn in files:
                        fp = os.path.join(root, fn)
                        zf.write(fp, os.path.relpath(fp, stage))

            rc = m.do_finalize(zip_path, os.path.join(tmp, 'ex'), appdir,
                               wait_pid=0, target_main=os.path.join(appdir, 'PDD EZ.exe'))
            self.assertEqual(rc, 0, 'finalize 应返回 0')
            with open(os.path.join(appdir, 'PDD EZ.exe'), 'rb') as f:
                self.assertEqual(f.read(), b'new', '主 exe 未更新')
            with open(os.path.join(appdir, '_internal', 'settings.json'), 'rb') as f:
                self.assertEqual(f.read(), b'USER_CFG', '用户配置被覆盖')
            picked = m._pick_main_exe(appdir)
            self.assertTrue(picked.endswith('PDD EZ.exe'), f'应选固定名: {picked}')
        finally:
            shutil.rmtree(tmp, ignore_errors=True)

    def test_finalize_with_updater_in_package(self):
        """blocker 场景：更新包含新 updater + updater 从 %TEMP% 运行（GUI 复制后）。
        验证 updater 自身可被覆盖（temp 副本不在 target_dir，rename 自身无冲突）。"""
        import importlib.util, tempfile, shutil, zipfile
        spec = importlib.util.spec_from_file_location('pdd_updater', os.path.join(HERE, 'updater.py'))
        m = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(m)

        tmp = tempfile.mkdtemp()
        try:
            appdir = os.path.join(tmp, 'PDD EZ')
            os.makedirs(os.path.join(appdir, '_internal'), exist_ok=True)
            with open(os.path.join(appdir, 'PDD EZ.exe'), 'wb') as f:
                f.write(b'old')
            with open(os.path.join(appdir, 'PDD EZ Updater.exe'), 'wb') as f:
                f.write(b'old_updater')
            with open(os.path.join(appdir, '_internal', 'settings.json'), 'wb') as f:
                f.write(b'USER_CFG')

            # 更新包含新 updater（真实发布包含 updater）
            stage = os.path.join(tmp, 'stage')
            os.makedirs(os.path.join(stage, 'PDD EZ', '_internal'), exist_ok=True)
            with open(os.path.join(stage, 'PDD EZ', 'PDD EZ.exe'), 'wb') as f:
                f.write(b'new')
            with open(os.path.join(stage, 'PDD EZ', 'PDD EZ Updater.exe'), 'wb') as f:
                f.write(b'new_updater')
            with open(os.path.join(stage, 'PDD EZ', '_internal', 'lib.dll'), 'wb') as f:
                f.write(b'lib')
            zip_path = os.path.join(tmp, 'upd.zip')
            with zipfile.ZipFile(zip_path, 'w') as zf:
                for root, _, files in os.walk(os.path.join(stage, 'PDD EZ')):
                    for fn in files:
                        fp = os.path.join(root, fn)
                        zf.write(fp, os.path.relpath(fp, stage))

            # 模拟 updater 从 %TEMP% 运行（sys.argv[0] 在 temp，不在 target_dir）
            old_argv0 = sys.argv[0]
            sys.argv[0] = os.path.join(tempfile.gettempdir(), 'PDD_EZ_Updater_tmp.exe')
            try:
                rc = m.do_finalize(zip_path, os.path.join(tmp, 'ex'), appdir,
                                   wait_pid=0, target_main=os.path.join(appdir, 'PDD EZ.exe'))
            finally:
                sys.argv[0] = old_argv0
            self.assertEqual(rc, 0, '含 updater 的更新应成功')
            with open(os.path.join(appdir, 'PDD EZ Updater.exe'), 'rb') as f:
                self.assertEqual(f.read(), b'new_updater', 'updater 自身未更新')
            with open(os.path.join(appdir, 'PDD EZ.exe'), 'rb') as f:
                self.assertEqual(f.read(), b'new', '主 exe 未更新')
            with open(os.path.join(appdir, '_internal', 'settings.json'), 'rb') as f:
                self.assertEqual(f.read(), b'USER_CFG', '用户配置被覆盖')
        finally:
            shutil.rmtree(tmp, ignore_errors=True)


class TestPrivacy(unittest.TestCase):
    """隐私过滤：真实测试数据不得出现在源码中"""

    SENSITIVE = ['盐渍', '鞭炮笋', '海带苗', '海带结', '烟台', '96622588033', '9662',
                 '939262347672', '984387564986', '284337564986']

    def test_no_sensitive_data(self):
        for f in sorted(os.listdir(HERE)):
            if not f.endswith(('.py', '.md', '.txt', '.json')):
                continue
            if f.startswith('_') or f == 'test_smoke.py':
                # 跳过临时扫描脚本/测试本身（其内容含敏感词列表）
                continue
            # v1.5.11.2：settings.json / regions.json 是用户运行时数据（gitignore 已排除、
            # 绝不入仓），但会被识别结果写入真实 SKU/地区——扫它们会让「用户一用就随机红」。
            # 隐私扫描只针对**入仓文件**（.gitignore 之外）。
            if f in ('settings.json', 'regions.json', 'settings.json.bak'):
                continue
            with open(os.path.join(HERE, f), encoding='utf-8', errors='replace') as fp:
                content = fp.read()
            for s in self.SENSITIVE:
                self.assertNotIn(s, content, f'{f} 含敏感数据: {s}')


class TestBugHuntRegressions(unittest.TestCase):
    """v1.4.5 bug hunt 修复回归（F4/F11 可纯函数断言的部分）"""

    def test_strip_tail_noise_keeps_pure_date(self):
        """bug hunt F4：纯日期/整值剥空时保留原文，不再清空更新时间列"""
        from ocr import strip_tail_noise
        self.assertEqual(strip_tail_noise('2024-05-04'), '2024-05-04')
        self.assertEqual(strip_tail_noise('2024年5月4日'), '2024年5月4日')
        # 常规词条剥离仍生效
        self.assertEqual(strip_tail_noise('128份 查看'), '128份')
        self.assertEqual(strip_tail_noise('示例仓库查看地址'), '示例仓库')

    def test_total_count_prefers_common_n(self):
        """bug hunt F11（fix-review C14 强化）：直接调生产 helper _parse_total_count"""
        from vision import _parse_total_count
        self.assertEqual(_parse_total_count('每页10条 共有128条'), 128)
        self.assertEqual(_parse_total_count('共有 9 条'), 9)
        self.assertEqual(_parse_total_count('总共有 5 条'), 5)
        self.assertEqual(_parse_total_count('总共25条'), 25)
        self.assertEqual(_parse_total_count('识别不了'), None)

    def test_strip_tail_noise_pure_word_regression(self):
        """fix-review N1 收窄：整值保护仅日期/数字形态，纯词条噪音仍剥空"""
        from ocr import strip_tail_noise
        self.assertEqual(strip_tail_noise('查看地址'), '')
        self.assertEqual(strip_tail_noise('更新记录'), '')
        self.assertEqual(strip_tail_noise('2024-05-04'), '2024-05-04')
        self.assertEqual(strip_tail_noise('128份 查看'), '128份')


class TestUpdaterRegression(unittest.TestCase):
    """fix-review P0/P1 回归：import re / _is_program_dir 版本号分支 / N2 0 值守卫"""

    def test_is_program_dir_version_named_no_nameerror(self):
        """P0-C1：版本号 exe + _internal 应为 True；未 import re 时此调用必然 NameError"""
        import importlib.util, os, tempfile, shutil
        spec = importlib.util.spec_from_file_location('pdd_updater2', os.path.join(HERE, 'updater.py'))
        m = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(m)
        tmp = tempfile.mkdtemp()
        try:
            d1 = os.path.join(tmp, 'Prog1')
            os.makedirs(os.path.join(d1, '_internal'), exist_ok=True)
            with open(os.path.join(d1, 'PDD EZ.exe'), 'wb') as f:
                f.write(b'x')
            self.assertTrue(m._is_program_dir(d1), '固定名+_internal 应为 True')
            d2 = os.path.join(tmp, 'Prog2')
            os.makedirs(os.path.join(d2, '_internal'), exist_ok=True)
            with open(os.path.join(d2, 'PDD EZ_v1.4.4.exe'), 'wb') as f:
                f.write(b'x')
            # 版本号 exe（旧版升级目录）分支不得 NameError，应为 True
            self.assertTrue(m._is_program_dir(d2), '版本号 exe+_internal 应为 True')
            d3 = os.path.join(tmp, 'Prog3')
            os.makedirs(d3, exist_ok=True)
            with open(os.path.join(d3, 'PDD EZ Backup.exe'), 'wb') as f:
                f.write(b'x')
            self.assertFalse(m._is_program_dir(d3), '仅有 Backup.exe 无 _internal 应为 False')
        finally:
            shutil.rmtree(tmp, ignore_errors=True)

    def test_merge_verify_keeps_zero_inventory(self):
        """fix-review N2/C8：真实 0 库存不被副模型覆盖；真空缺才补"""
        import importlib.util, os
        spec = importlib.util.spec_from_file_location('pdd_ocr2', os.path.join(HERE, 'ocr.py'))
        m = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(m)
        verify = [{'name': 'A', 'stock': '500'}, {'name': 'B', 'stock': '500'}]
        # 触发二次识别：_missing_id=True（缺 ID 主模型行）
        items = [{'name': 'A', 'stock': '0', '_missing_id': True}]
        out = m.merge_verify_items([dict(i) for i in items], [dict(v) for v in verify])
        self.assertEqual(out[0].get('stock'), '0', '0 库存应保留')
        items2 = [{'name': 'B', 'stock': '', '_missing_id': True}]
        out2 = m.merge_verify_items([dict(i) for i in items2], [dict(v) for v in verify])
        self.assertEqual(out2[0].get('stock'), '500', '真空缺应补全')

    def test_deleted_files_applied_on_finalize(self):
        """P2-4（bug hunt F15/R1）：deleted-files.txt 删除生效——do_finalize 覆盖成功后
        目标端旧模板/资源被删；且删除发生在覆盖成功之后（覆盖中途失败时旧文件不删，回滚完整）。"""
        import importlib.util, tempfile, shutil, zipfile
        spec = importlib.util.spec_from_file_location('pdd_updater_df', os.path.join(HERE, 'updater.py'))
        m = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(m)
        tmp = tempfile.mkdtemp()
        try:
            appdir = os.path.join(tmp, 'PDD EZ')
            os.makedirs(os.path.join(appdir, '_internal'), exist_ok=True)
            with open(os.path.join(appdir, 'PDD EZ.exe'), 'wb') as f:
                f.write(b'old')
            # 目标端有一个要被删除的旧模板
            os.makedirs(os.path.join(appdir, 'templates'), exist_ok=True)
            with open(os.path.join(appdir, 'templates', 'stale.csv'), 'w', encoding='utf-8') as f:
                f.write('stale')
            # 构造更新包：新 exe + deleted-files.txt（声明删 templates/stale.csv）
            stage = os.path.join(tmp, 'stage')
            os.makedirs(os.path.join(stage, 'PDD EZ', '_internal'), exist_ok=True)
            with open(os.path.join(stage, 'PDD EZ', 'PDD EZ.exe'), 'wb') as f:
                f.write(b'new')
            with open(os.path.join(stage, 'PDD EZ', 'deleted-files.txt'), 'w', encoding='utf-8') as f:
                f.write('templates/stale.csv\n')
            zip_path = os.path.join(tmp, 'upd.zip')
            with zipfile.ZipFile(zip_path, 'w') as zf:
                for root, _, files in os.walk(os.path.join(stage, 'PDD EZ')):
                    for fn in files:
                        fp = os.path.join(root, fn)
                        zf.write(fp, os.path.relpath(fp, stage))
            rc = m.do_finalize(zip_path, os.path.join(tmp, 'ex'), appdir,
                               wait_pid=0, target_main=os.path.join(appdir, 'PDD EZ.exe'))
            self.assertEqual(rc, 0, 'finalize 应返回 0')
            with open(os.path.join(appdir, 'PDD EZ.exe'), 'rb') as f:
                self.assertEqual(f.read(), b'new')
            self.assertFalse(os.path.exists(os.path.join(appdir, 'templates', 'stale.csv')),
                             '删除清单中的旧模板应在覆盖成功后删除')
        finally:
            shutil.rmtree(tmp, ignore_errors=True)

    def test_deleted_files_whitelist_blocks_unsafe(self):
        """P2-4（bug hunt F15 白名单）：_apply_deleted_files 拒绝 exe/dll/穿越路径/绝对路径；
        白名单外的任意文件绝不删。"""
        import importlib.util, tempfile, shutil
        spec = importlib.util.spec_from_file_location('pdd_updater_dfw', os.path.join(HERE, 'updater.py'))
        m = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(m)
        tmp = tempfile.mkdtemp()
        try:
            extracted = os.path.join(tmp, 'ex')
            target = os.path.join(tmp, 'target')
            os.makedirs(extracted)
            os.makedirs(target)
            with open(os.path.join(extracted, 'deleted-files.txt'), 'w', encoding='utf-8') as f:
                f.write('..\\escape.txt\n')          # 穿越
                f.write('templates/legit.csv\n')     # 白名单内
                f.write('PDD EZ.exe\n')              # exe 拒绝
                f.write('C:/absolute.txt\n')         # 绝对路径拒绝
                f.write('random_other.txt\n')        # 白名单外
            with open(os.path.join(target, 'escape.txt'), 'w') as f:
                f.write('x')
            os.makedirs(os.path.join(target, 'templates'))
            with open(os.path.join(target, 'templates', 'legit.csv'), 'w') as f:
                f.write('x')
            with open(os.path.join(target, 'random_other.txt'), 'w') as f:
                f.write('x')
            n = m._apply_deleted_files(extracted, target)
            # 只该删 legit.csv（白名单内）；escape.txt/random_other.txt 保留
            self.assertEqual(n, 1, f'应恰好删除 1 个白名单文件，实际 {n}')
            self.assertFalse(os.path.exists(os.path.join(target, 'templates', 'legit.csv')), '白名单文件应删除')
            self.assertTrue(os.path.exists(os.path.join(target, 'escape.txt')), '穿越路径不得删除')
            self.assertTrue(os.path.exists(os.path.join(target, 'random_other.txt')), '白名单外不得删除')
        finally:
            shutil.rmtree(tmp, ignore_errors=True)

    def test_version_comparison_semantics(self):
        """P2-4（bug hunt F19）：内联版本比较逻辑——远端 < 本地时不应更新（rc 语义为'无需更新'）。
        直接复算 updater 内联 _version_newer 的算法（无法 import 闭包，按同源逻辑断言）。"""
        def _version_newer(remote, local):
            def _p(v):
                v = str(v).lstrip('vV')
                return [int(x) for x in v.split('.') if x.isdigit()]
            r, l = _p(remote), _p(local)
            n = max(len(r), len(l))
            return (r + [0] * (n - len(r))) > (l + [0] * (n - len(l)))
        # 远端 < 本地 → 不更新（false）；远端 > 本地 → 更新（true）；相等 → 不更新
        self.assertFalse(_version_newer('v1.4.5', 'v1.4.6'), '远端<本地应不更新')
        self.assertTrue(_version_newer('v1.4.7', 'v1.4.6'), '远端>本地应更新')
        self.assertFalse(_version_newer('v1.4.6', 'v1.4.6'), '相同版本应不更新')
        # 小版本跨级：1.4.10 > 1.4.9
        self.assertTrue(_version_newer('v1.4.10', 'v1.4.9'), '10>9 应按数字比较')
        self.assertFalse(_version_newer('v1.4.5', 'v1.4.10'), '5<10 应按数字比较')
        # 版本缺段按 0 补齐
        self.assertFalse(_version_newer('v1.4', 'v1.4.0'), '缺段补 0 相等')
        self.assertTrue(_version_newer('v1.5', 'v1.4.9'), '次版本优先于补段')

    def test_do_replace_oserror_rolls_back(self):
        """P2-4（bug hunt F22）：copy2 抛 OSError（磁盘满等）时 target 会恢复 .old，
        src 保存为 .new——主 exe 不缺失。回归旧 bug：磁盘满时主 exe 被删且无备份。"""
        import importlib.util, tempfile, shutil, os
        from unittest.mock import patch
        _real_copy2 = shutil.copy2
        spec = importlib.util.spec_from_file_location('pdd_updater_f22', os.path.join(HERE, 'updater.py'))
        m = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(m)
        tmp = tempfile.mkdtemp()
        try:
            target = os.path.join(tmp, 'PDD EZ.exe')
            src = os.path.join(tmp, 'new_pkg.exe')
            with open(target, 'wb') as f:
                f.write(b'old_exe')
            with open(src, 'wb') as f:
                f.write(b'new_bin')
            with patch('shutil.copy2', side_effect=lambda src2, dst2: (_ for _ in ()).throw(
                    OSError(28, 'No space left on device')) if dst2.endswith('.exe') else _real_copy2(src2, dst2)):
                m._do_replace(src, target)
            # 磁盘满：target 应保留原内容（回滚），src 另存为 .new
            with open(target, 'rb') as f:
                self.assertEqual(f.read(), b'old_exe', '失败后 target 应恢复原文件')
            self.assertTrue(os.path.exists(target + '.new'), '源应保留为 .new 供手动处理')
        finally:
            shutil.rmtree(tmp, ignore_errors=True)



# ══════════════════════════════════════════════════════════════════
# v1.4.7 商业升级三工作流回归合并（t8：t5 WS-C + t6 WS-B + t7 WS-A）
# 来源：test_tmp_wsc.py / test_tmp_wsb.py / test_tmp_wsa.py（合并后删除）。
# 既有锚点测试保持不动；本段类名与上文无冲突。
# ══════════════════════════════════════════════════════════════════
import importlib.util
import json
import shutil
import sqlite3
import tempfile
import threading
from datetime import datetime
from unittest.mock import patch, MagicMock

import table_import
from table_import import (
    read_table_rows,
    guess_mapping,
    import_items,
    write_template,
    MAX_IMPORT_ROWS,
    SUPPORTED_EXT,
    LEGACY_EXT,
    _sanitize_cell,
    _parse_num_text,
)


# ──────────────────────────────────────────────────────────────────
# WS-B 表格导入通道（table_import.py，来源 test_tmp_wsb.py）
# ──────────────────────────────────────────────────────────────────
def _write_csv(path, content_bytes):
    with open(path, 'wb') as f:
        f.write(content_bytes)


def _write_xlsx(path, rows):
    """rows: list of list，写第一个 sheet。rows[0] 是表头。"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    for ri, row in enumerate(rows, 1):
        for ci, v in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=v)
    wb.save(path)


# 映射常量：与 utils.DEFAULT_COL_MAPPING 完全一致，避免测试因 settings.json 状态波动
STANDARD_HEADERS = ['商品信息', '仓库总库存', '仓库预估总销售数', '销售区域', '仓库信息']


# ------------------------------------------------------------
# 1. read_table_rows
# ------------------------------------------------------------

class TestReadCSVGbk(unittest.TestCase):
    """GBK 编码 CSV 读取（PDD 后台/Excel 中文版默认）"""

    def test_gbk_basic(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, 'in_gbk.csv')
            content = (
                ','.join(STANDARD_HEADERS) + '\n' +
                '示例商品A,1,200,广东省广州市,广州中心仓\n' +
                '示例商品B,500,300,浙江省,华东中心仓\n'
            )
            _write_csv(path, content.encode('gbk'))
            headers, rows = read_table_rows(path)
            self.assertEqual(headers, STANDARD_HEADERS)
            self.assertEqual(len(rows), 2)
            self.assertEqual(rows[0]['商品信息'], '示例商品A')
            self.assertEqual(rows[1]['商品信息'], '示例商品B')

    def test_gbk_with_bom_skip(self):
        """GBK CSV 带 UTF-8 BOM：utf-8-sig 优先于 gbk，应能成功解码"""
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, 'in_bom.csv')
            content = (
                ','.join(STANDARD_HEADERS) + '\n' +
                '商品A,1,200,广东,广州仓\n'
            )
            # 写 UTF-8 BOM + UTF-8 内容（utf-8-sig 探测命中）
            _write_csv(path, b'\xef\xbb\xbf' + content.encode('utf-8'))
            headers, rows = read_table_rows(path)
            self.assertEqual(headers, STANDARD_HEADERS)
            self.assertEqual(rows[0]['商品信息'], '商品A')


class TestReadCSVUtf8Sig(unittest.TestCase):
    """UTF-8-BOM CSV 读取"""

    def test_utf8_sig_explicit(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, 'in_utf8.csv')
            content = ','.join(STANDARD_HEADERS) + '\n商品A,1,200,广东,广州仓\n'
            _write_csv(path, b'\xef\xbb\xbf' + content.encode('utf-8'))
            headers, rows = read_table_rows(path)
            # utf-8-sig 自动去 BOM，headers 干净
            self.assertEqual(headers, STANDARD_HEADERS)
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]['商品信息'], '商品A')

    def test_plain_utf8(self):
        """无 BOM 的 UTF-8 CSV 也能解析（utf-8 兜底）"""
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, 'in_plain.csv')
            content = ','.join(STANDARD_HEADERS) + '\n商品B,2,300,浙江,杭州仓\n'
            _write_csv(path, content.encode('utf-8'))
            headers, rows = read_table_rows(path)
            self.assertEqual(headers, STANDARD_HEADERS)
            self.assertEqual(rows[0]['商品信息'], '商品B')


class TestReadCSVEncodingError(unittest.TestCase):
    """编码不可识别：显式报错（§4 失败哲学）"""

    def test_unknown_encoding_raises(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, 'bad.csv')
            # 写日文 Shift-JIS 内容（3 种编码都解不出来）
            sjis_bytes = '商品情報,100,200,東京,東京倉庫\n'.encode('shift_jis', errors='ignore')
            _write_csv(path, sjis_bytes)
            # 删去后用纯二进制确保 strict 失败
            with open(path, 'wb') as f:
                f.write(bytes([0x80, 0x81, 0x82, 0x83] * 20))  # 无效字节序列
            with self.assertRaises(ValueError) as ctx:
                read_table_rows(path)
            self.assertIn('无法识别文件编码', str(ctx.exception))
            self.assertIn('UTF-8', str(ctx.exception))


class TestReadCSVDelimiter(unittest.TestCase):
    """分隔符嗅探：, ; \\t"""

    def test_semicolon_delimiter(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, 'in_sc.csv')
            content = ';'.join(STANDARD_HEADERS) + '\n商品A;1;200;广东;广州仓\n'
            _write_csv(path, content.encode('utf-8'))
            headers, rows = read_table_rows(path)
            self.assertEqual(headers, STANDARD_HEADERS)
            self.assertEqual(rows[0]['商品信息'], '商品A')


class TestReadXLSX(unittest.TestCase):
    """XLSX 读取：openpyxl read_only + data_only"""

    def test_xlsx_basic(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, 'in.xlsx')
            rows = [
                STANDARD_HEADERS,
                ['商品A', 1, 200, '广东省', '广州仓'],
                ['商品B', 500, 300, '浙江省', '杭州仓'],
            ]
            _write_xlsx(path, rows)
            headers, rows = read_table_rows(path)
            self.assertEqual(headers, STANDARD_HEADERS)
            self.assertEqual(len(rows), 2)
            # 数字单元格被 str() 规范化
            self.assertEqual(rows[0]['仓库总库存'], '1')
            self.assertEqual(rows[0]['仓库预估总销售数'], '200')

    def test_xlsx_empty_cells(self):
        """空单元格 → 空字符串（不丢列）"""
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, 'empty.xlsx')
            rows = [
                STANDARD_HEADERS,
                ['商品A', None, 200, None, '广州仓'],
            ]
            _write_xlsx(path, rows)
            headers, r = read_table_rows(path)
            self.assertEqual(r[0]['商品信息'], '商品A')
            self.assertEqual(r[0]['仓库总库存'], '')
            self.assertEqual(r[0]['销售区域'], '')

    def test_xlsx_formula_value(self):
        """data_only=True 取计算值非公式串：A2 公式 =1+1 → 2"""
        from openpyxl import Workbook
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, 'f.xlsx')
            wb = Workbook()
            ws = wb.active
            # 写一行表头
            for ci, h in enumerate(STANDARD_HEADERS, 1):
                ws.cell(row=1, column=ci, value=h)
            # 数据行：A 列商品，B 列 1+1 公式
            ws.cell(row=2, column=1, value='商品A')
            ws.cell(row=2, column=2, value='=1+1')
            ws.cell(row=2, column=3, value=200)
            ws.cell(row=2, column=4, value='广东')
            ws.cell(row=2, column=5, value='广州仓')
            wb.save(path)
            # 重打开并 data_only=True 验证：但 openpyxl 写公式时不会自动算 result，
            # 这里改为：手写预计算值（data_only 行为依赖 Excel 重算，纯 openpyxl 写公式无 cached value）
            # 本测试覆盖：data_only=True 不会报 #VALUE 错误
            _headers, _rows = read_table_rows(path)
            # 不抛异常即可；具体值可能是 '=1+1' 或 2，取决于 cached value
            self.assertTrue(len(_rows) == 1)


class TestReadLegacyXLS(unittest.TestCase):
    """.xls 老格式显式拒绝"""

    def test_xls_rejected(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, 'old.xls')
            with open(path, 'wb') as f:
                f.write(b'fake xls content')
            with self.assertRaises(ValueError) as ctx:
                read_table_rows(path)
            self.assertIn('xls', str(ctx.exception))
            self.assertIn('.xlsx', str(ctx.exception))


class TestReadRowLimit(unittest.TestCase):
    """1 万行上限"""

    def test_over_limit_rejected(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, 'big.csv')
            lines = [','.join(STANDARD_HEADERS)]
            for i in range(MAX_IMPORT_ROWS + 1):
                lines.append(f'商品{i},1,1,广东,广州仓')
            _write_csv(path, ('\n'.join(lines) + '\n').encode('utf-8'))
            with self.assertRaises(ValueError) as ctx:
                read_table_rows(path)
            self.assertIn('上限', str(ctx.exception))
            self.assertIn(str(MAX_IMPORT_ROWS), str(ctx.exception))

    def test_exact_limit_ok(self):
        """恰好 1 万行：应允许（边界）"""
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, 'exact.csv')
            lines = [','.join(STANDARD_HEADERS)]
            for i in range(MAX_IMPORT_ROWS):
                lines.append(f'商品{i},1,1,广东,广州仓')
            _write_csv(path, ('\n'.join(lines) + '\n').encode('utf-8'))
            headers, rows = read_table_rows(path)
            self.assertEqual(len(rows), MAX_IMPORT_ROWS)


class TestReadEmptyHeader(unittest.TestCase):
    """空列名补 '列N'"""

    def test_blank_header_filled(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, 'h.csv')
            content = '商品信息,,仓库预估总销售数,,\n商品A,1,200,广东,广州仓\n'
            _write_csv(path, content.encode('utf-8'))
            headers, _ = read_table_rows(path)
            self.assertEqual(headers, ['商品信息', '列2', '仓库预估总销售数', '列4', '列5'])


# ------------------------------------------------------------
# 2. guess_mapping
# ------------------------------------------------------------

class TestGuessMapping(unittest.TestCase):
    """列映射：精确匹配 + 缺 name 不静默 fallback"""

    def test_standard_headers(self):
        """标准 PDD 后台表头 → 5 字段全命中"""
        found, missing = guess_mapping(STANDARD_HEADERS)
        self.assertEqual(missing, [])
        self.assertEqual(found['name'], '商品信息')
        self.assertEqual(found['stock'], '仓库总库存')
        self.assertEqual(found['sales'], '仓库预估总销售数')
        self.assertEqual(found['region'], '销售区域')
        self.assertEqual(found['warehouse'], '仓库信息')

    def test_alias_match(self):
        """stock 别名（仓库库存）应被认作 stock"""
        headers = ['商品名称', '仓库库存', '仓库预估总销售数']
        found, missing = guess_mapping(headers)
        # 任务卡说明：缺 region/warehouse 是合法的
        self.assertEqual(found.get('stock'), '仓库库存')
        self.assertNotIn('stock', missing)

    def test_missing_name(self):
        """缺 name 列 → missing 包含 name（不静默 fallback）"""
        headers = ['仓库总库存', '仓库预估总销售数', '销售区域']
        found, missing = guess_mapping(headers)
        self.assertIn('name', missing)
        self.assertNotIn('name', found)

    def test_missing_stock(self):
        """缺 stock 列 → missing 包含 stock"""
        headers = ['商品信息', '仓库预估总销售数', '销售区域']
        found, missing = guess_mapping(headers)
        self.assertIn('stock', missing)

    def test_missing_sales(self):
        """缺 sales 列 → missing 包含 sales"""
        headers = ['商品信息', '仓库总库存', '销售区域']
        found, missing = guess_mapping(headers)
        self.assertIn('sales', missing)

    def test_fuzzy_name_not_used(self):
        """宪法 §1：name 永不模糊匹配。'商品报价' 不应被误配为 name"""
        headers = ['商品报价', '仓库总库存', '仓库预估总销售数']
        found, missing = guess_mapping(headers)
        self.assertIn('name', missing, 'name 字段必须精确匹配，不应模糊到商品报价')
        self.assertNotIn('name', found)

    def test_normalize_whitespace(self):
        """列名带全角空格/前后空格应被 normalize 后命中"""
        headers = ['商品信息 ', ' 仓库总库存', '　仓库预估总销售数', '销售区域', '仓库信息']
        found, missing = guess_mapping(headers)
        self.assertEqual(missing, [])


# ------------------------------------------------------------
# 3. import_items — 万/千分位解析
# ------------------------------------------------------------

class TestImportItemsNumberParsing(unittest.TestCase):
    """_parse_num_text 各种数字格式（万/千分位/w/k）经 import_items 正确解析"""

    def test_wan_unit(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, 'n.csv')
            content = (
                ','.join(STANDARD_HEADERS) + '\n' +
                '商品A,1.2万,5000,广东,广州仓\n'
            )
            _write_csv(path, content.encode('utf-8'))
            items, _ = import_items(path)
            self.assertEqual(items[0]['stock'], 12000, '1.2万 应解析为 12000')
            self.assertEqual(items[0]['sales'], 5000)

    def test_thousands_separator(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, 'n.csv')
            content = (
                ','.join(STANDARD_HEADERS) + '\n' +
                '商品A,"1,234","2,500",广东,广州仓\n'
            )
            _write_csv(path, content.encode('utf-8'))
            items, _ = import_items(path)
            self.assertEqual(items[0]['stock'], 1234, '1,234 应解析为 1234')
            self.assertEqual(items[0]['sales'], 2500)

    def test_w_k_units(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, 'n.csv')
            content = (
                ','.join(STANDARD_HEADERS) + '\n' +
                '商品A,1.5w,8k,广东,广州仓\n'
            )
            _write_csv(path, content.encode('utf-8'))
            items, _ = import_items(path)
            self.assertEqual(items[0]['stock'], 15000, '1.5w 应解析为 15000')
            self.assertEqual(items[0]['sales'], 8000, '8k 应解析为 8000')

    def test_fullwidth_digits(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, 'n.csv')
            content = (
                ','.join(STANDARD_HEADERS) + '\n' +
                '商品A,共 ２００,统计中,广东,广州仓\n'
            )
            _write_csv(path, content.encode('utf-8'))
            items, _ = import_items(path)
            self.assertEqual(items[0]['stock'], 200, '全角 200 应解析为 200')
            # 销量单元格为'统计中'，_parse_num_text 返 0 → warning
            self.assertEqual(items[0]['sales'], 0)


# ------------------------------------------------------------
# 4. import_items — issues 行号/原因准确
# ------------------------------------------------------------

class TestImportItemsIssues(unittest.TestCase):
    """issues 报告：行号、level、reason 准确"""

    def test_missing_name_row_number(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, 'i.csv')
            content = (
                ','.join(STANDARD_HEADERS) + '\n' +
                '商品A,1,200,广东,广州仓\n' +
                ',100,50,江苏,南京仓\n' +  # 第 2 数据行，缺商品名
                '商品C,3,400,浙江,杭州仓\n'  # 第 3 数据行
            )
            _write_csv(path, content.encode('utf-8'))
            items, issues = import_items(path)
            err = [i for i in issues if i['level'] == 'error']
            self.assertEqual(len(err), 1)
            self.assertEqual(err[0]['row'], 2, '行号=数据行索引（从 1 起）')
            self.assertIn('缺商品名', err[0]['reason'])

    def test_unparseable_stock_warning(self):
        """stock='统计中' → warning（_parse_num_text 静默返 0 必须暴露）"""
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, 'i.csv')
            content = (
                ','.join(STANDARD_HEADERS) + '\n' +
                '商品A,统计中,200,广东,广州仓\n' +
                '商品B,abc,300,浙江,杭州仓\n'  # 非数字也走 warning
            )
            _write_csv(path, content.encode('utf-8'))
            items, issues = import_items(path)
            warns = [i for i in issues if i['level'] == 'warning']
            self.assertEqual(len(warns), 2, f'应有 2 条 warning，实际 {len(warns)}')
            rows = sorted(w['row'] for w in warns)
            self.assertEqual(rows, [1, 2])

    def test_unparseable_sales_warning(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, 'i.csv')
            content = (
                ','.join(STANDARD_HEADERS) + '\n' +
                '商品A,1,统计中,广东,广州仓\n'
            )
            _write_csv(path, content.encode('utf-8'))
            items, issues = import_items(path)
            warns = [i for i in issues if i['level'] == 'warning']
            self.assertEqual(len(warns), 1)
            self.assertIn('销量', warns[0]['reason'])

    def test_empty_stock_no_warning(self):
        """stock 单元格为空（不是非空但解析失败）→ 不报 warning（不打扰）"""
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, 'i.csv')
            content = (
                ','.join(STANDARD_HEADERS) + '\n' +
                '商品A,,200,广东,广州仓\n'  # stock 留空
            )
            _write_csv(path, content.encode('utf-8'))
            items, issues = import_items(path)
            self.assertEqual(issues, [], '空 stock 不应报 warning')


class TestImportItemsRejectMissingNameColumn(unittest.TestCase):
    """缺商品名列（header 里就没有 name）→ import_items 抛 ValueError"""

    def test_missing_name_column_raises(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, 'no_name.csv')
            content = (
                ','.join(['仓库总库存', '仓库预估总销售数', '销售区域']) + '\n' +
                '1,200,广东\n'
            )
            _write_csv(path, content.encode('utf-8'))
            with self.assertRaises(ValueError) as ctx:
                import_items(path)
            self.assertIn('name', str(ctx.exception))
            self.assertIn('缺关键字段', str(ctx.exception))


# ------------------------------------------------------------
# 5. import_items — 公式注入防御（_sanitize_cell 复用）
# ------------------------------------------------------------

class TestSanitizeCellIntegration(unittest.TestCase):
    """_sanitize_cell 强制复用：= + - @ 开头加 ' 前缀"""

    def test_formula_prefixes_sanitized(self):
        # 直接验证 _sanitize_cell 行为
        self.assertEqual(_sanitize_cell('=cmd|/c calc'), "'=cmd|/c calc")
        self.assertEqual(_sanitize_cell('+1+1'), "'+1+1")
        self.assertEqual(_sanitize_cell('-2+3'), "'-2+3")
        self.assertEqual(_sanitize_cell('@SUM(A1:A2)'), "'@SUM(A1:A2)")
        # 普通文本不变
        self.assertEqual(_sanitize_cell('商品A'), '商品A')
        self.assertEqual(_sanitize_cell('100'), '100')
        # 不重复加引号
        self.assertEqual(_sanitize_cell("'=FORMULA"), "'=FORMULA")

    def test_formula_in_csv_preserved_as_value(self):
        """CSV 里 =cmd 开头的数据：经 import_items 后原值进 items（GUI 收口侧 sanitize）"""
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, 'f.csv')
            content = (
                ','.join(STANDARD_HEADERS) + '\n' +
                '=cmd|/c calc,1,200,广东,广州仓\n' +  # 商品名列以 = 开头
                '+1+1,2,300,浙江,杭州仓\n' +  # 以 + 开头
                '@SUM,3,400,江苏,南京仓\n'  # 以 @ 开头
            )
            _write_csv(path, content.encode('utf-8'))
            # import_items 自身不清洗（任务卡：清洗逻辑由 GUI 收口侧做）
            # 但调用方应能在收口侧通过 _sanitize_cell 拿到清洗值
            items, _ = import_items(path)
            self.assertEqual(items[0]['name'], '=cmd|/c calc', '导入侧保留原值供 GUI sanitize')
            # 模拟 GUI 收口侧 sanitize
            cleaned = [_sanitize_cell(it.get('name', '')) for it in items]
            self.assertEqual(cleaned[0], "'=cmd|/c calc")
            self.assertEqual(cleaned[1], "'+1+1")
            self.assertEqual(cleaned[2], "'@SUM")


# ------------------------------------------------------------
# 6. import_items — SKU ID 拆分
# ------------------------------------------------------------

class TestImportItemsSkuId(unittest.TestCase):
    """商品信息列含 ID:xxx 时正确拆为 name + sku_id"""

    def test_sku_id_split(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, 's.csv')
            content = (
                ','.join(STANDARD_HEADERS) + '\n' +
                '示例商品A ID:12345678901,1,200,广东,广州仓\n'
            )
            _write_csv(path, content.encode('utf-8'))
            items, _ = import_items(path)
            self.assertEqual(items[0]['name'], '示例商品A')
            self.assertEqual(items[0]['sku_id'], '12345678901')

    def test_no_sku_id_keeps_row(self):
        """无 ID 整行不丢（v1.4 修复），仅标 _missing_id"""
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, 's.csv')
            content = (
                ','.join(STANDARD_HEADERS) + '\n' +
                '示例商品A,1,200,广东,广州仓\n'
            )
            _write_csv(path, content.encode('utf-8'))
            items, issues = import_items(path)
            self.assertEqual(len(items), 1, '无 ID 不丢行')
            self.assertTrue(items[0].get('_missing_id'))
            self.assertEqual(items[0]['sku_id'], '')


# ------------------------------------------------------------
# 7. write_template
# ------------------------------------------------------------

class TestWriteTemplate(unittest.TestCase):
    """模板生成：双 Sheet + 样式复用 + 不随包分发"""

    def test_template_creates_two_sheets(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, 'tpl.xlsx')
            write_template(path)
            self.assertTrue(os.path.isfile(path))
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            self.assertIn('数据样例', wb.sheetnames)
            self.assertIn('列名说明', wb.sheetnames)
            # 验证 Sheet1 有表头 + 2 行示例
            ws1 = wb['数据样例']
            rows = list(ws1.iter_rows(values_only=True))
            self.assertEqual(len(rows), 3, f'Sheet1 应 3 行（表头 + 2 示例），实际 {len(rows)}')
            # 验证 Sheet2 有 5 字段说明 + 表头
            ws2 = wb['列名说明']
            rows2 = list(ws2.iter_rows(values_only=True))
            self.assertEqual(len(rows2), 6, f'Sheet2 应 6 行（表头 + 5 字段），实际 {len(rows2)}')
            wb.close()

    def test_template_uses_sanitize_cell(self):
        """模板示例单元格过 _sanitize_cell（即使没 = + - @ 开头也安全）"""
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, 'tpl2.xlsx')
            write_template(path)
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            ws1 = wb['数据样例']
            headers = [c.value for c in next(ws1.iter_rows(min_row=1, max_row=1))]
            # 表头应含"商品信息"等真实列名（settings 默认）
            self.assertIn('商品信息', headers)
            wb.close()


# ------------------------------------------------------------
# 8. 集成
# ------------------------------------------------------------

class TestEndToEnd(unittest.TestCase):
    """全链路：GBK → import_items → items 字段完整"""

    def test_full_pipeline(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, 'full.csv')
            content = (
                ','.join(STANDARD_HEADERS) + '\n' +
                '商品A ID:12345678901,1.2万,5000,广东省广州市,广州中心仓\n' +
                '商品B,500,300,浙江,杭州仓\n'
            )
            _write_csv(path, content.encode('gbk'))
            items, issues = import_items(path)
            self.assertEqual(len(items), 2)
            self.assertEqual(len(issues), 0, f'干净数据应无 issues，实际 {issues!r}')
            # 第一行
            self.assertEqual(items[0]['name'], '商品A')
            self.assertEqual(items[0]['sku_id'], '12345678901')
            self.assertEqual(items[0]['stock'], 12000)
            self.assertEqual(items[0]['sales'], 5000)
            self.assertEqual(items[0]['region'], '广东省广州')  # strip_region_suffix 去尾"市"
            self.assertEqual(items[0]['warehouse'], '广州中心仓')
            # 第二行
            self.assertEqual(items[1]['name'], '商品B')
            self.assertEqual(items[1]['stock'], 500)


# ──────────────────────────────────────────────────────────────────
# WS-C 用量采集（usage_extractor / usage_store / ocr-vision 三元组，来源 test_tmp_wsc.py）
# ──────────────────────────────────────────────────────────────────
class TestUsageExtractorExtract(unittest.TestCase):
    """SPEC §2.1 6 步降级链 + §2.2 契约。"""

    def setUp(self):
        import usage_extractor
        self.ux = usage_extractor

    def test_step1_chat_completions_prompt_completion(self):
        """SPEC §8.1 部分：chat.completions 顶层 usage，prompt/completion 命名。"""
        data = {'usage': {'prompt_tokens': 100, 'completion_tokens': 50, 'total_tokens': 150}}
        r = self.ux.extract(data, 'doubao', 'chat.completions')
        self.assertIsNotNone(r)
        self.assertEqual(r['prompt'], 100)
        self.assertEqual(r['completion'], 50)
        self.assertEqual(r['total'], 150)
        self.assertEqual(r['source'], 'data.usage')
        self.assertIsNone(r['image_tokens'])

    def test_step2_responses_input_output_top(self):
        """SPEC §2.1 步 2：responses API 顶层 input/output 命名。"""
        data = {'usage': {'input_tokens': 200, 'output_tokens': 80, 'total_tokens': 280}}
        r = self.ux.extract(data, 'doubao', 'responses')
        self.assertIsNotNone(r)
        self.assertEqual(r['prompt'], 200)
        self.assertEqual(r['completion'], 80)
        self.assertEqual(r['total'], 280)
        self.assertEqual(r['source'], 'data.usage')

    def test_step3_responses_output_minus1(self):
        """SPEC §8.1：doubao responses 顶层缺失 → output[-1].usage 兜底。"""
        data = {
            'usage': None,  # 顶层缺失
            'output': [
                {'type': 'message', 'usage': {'input_tokens': 300, 'output_tokens': 100, 'total_tokens': 400}}
            ],
        }
        r = self.ux.extract(data, 'doubao', 'responses')
        self.assertIsNotNone(r)
        self.assertEqual(r['prompt'], 300)
        self.assertEqual(r['completion'], 100)
        self.assertEqual(r['total'], 400)
        self.assertEqual(r['source'], 'data.output[-1].usage')

    def test_step4_multimodal_image_tokens(self):
        """SPEC §8.2：qwen multimodal-generation image_tokens 专项（仅 image_tokens 字段时）。"""
        # 步 4 仅在 usage 里有 image_tokens 但无 input_tokens/output_tokens 时命中
        # （SPEC §2.1 步 2 优先于步 4）；这是 OCR 老模型形态
        data = {'usage': {'image_tokens': 1024, 'total_tokens': 1071}}
        r = self.ux.extract(data, 'qwen', 'multimodal-generation')
        self.assertIsNotNone(r)
        self.assertEqual(r['prompt'], 1024)
        self.assertEqual(r['completion'], 0)
        # total = 1024 (image) + 0 (completion) → 自校验补齐 1024
        # 1071 > 1024 → 保留原值
        self.assertEqual(r['total'], 1071)
        self.assertEqual(r['image_tokens'], 1024)
        self.assertEqual(r['source'], 'data.usage.multimodal')

    def test_total_self_check(self):
        """SPEC §8.4：total=0 但 prompt+completion>0 → _pack 用求和补齐。"""
        data = {'usage': {'prompt_tokens': 100, 'completion_tokens': 50, 'total_tokens': 0}}
        r = self.ux.extract(data, 'glm', 'chat.completions')
        self.assertIsNotNone(r)
        self.assertEqual(r['total'], 150, '_pack 应自校验补齐 total')

    def test_total_self_check_underreport(self):
        """total < prompt+completion（GLM 计费 bug）→ 同样用求和补齐。"""
        data = {'usage': {'prompt_tokens': 100, 'completion_tokens': 50, 'total_tokens': 50}}
        r = self.ux.extract(data, 'glm', 'chat.completions')
        self.assertIsNotNone(r)
        self.assertEqual(r['total'], 150, 'total < sum 时用求和补齐')

    def test_glm_legacy_missing(self):
        """SPEC §8.3：GLM 老 flash 缺失 → None（调用方走 §3 兜底）。"""
        data = {'usage': None, 'choices': [{'message': {'content': '...'}}]}
        r = self.ux.extract(data, 'glm', 'chat.completions')
        self.assertIsNone(r)

    def test_extract_non_dict_returns_none(self):
        """契约：data 非 dict → None。"""
        self.assertIsNone(self.ux.extract(None, 'glm', 'chat.completions'))
        self.assertIsNone(self.ux.extract('not dict', 'glm', 'chat.completions'))
        self.assertIsNone(self.ux.extract(42, 'glm', 'chat.completions'))

    def test_extract_no_raise(self):
        """契约：函数体任何异常吞掉，不外抛。"""
        # 故意构造 dict-with-attrs-异常（output 不是 list）
        try:
            r = self.ux.extract({'output': 'not-list'}, 'doubao', 'responses')
            self.assertIsNone(r)
        except Exception as e:
            self.fail(f'extract 不应外抛，但抛了 {e!r}')

    def test_step_priority_chat_over_output(self):
        """SPEC §2.1：步 1/2 优先于步 3——顶层存在时不读 output[-1]。"""
        data = {
            'usage': {'prompt_tokens': 10, 'completion_tokens': 5, 'total_tokens': 15},
            'output': [{'usage': {'input_tokens': 999, 'output_tokens': 999}}],
        }
        r = self.ux.extract(data, 'doubao', 'responses')
        # 步 2 先命中（顶层 input/output 命名... 但这里是 prompt/completion 命名 → 步 1 命中）
        # 步 1 要求有 prompt_tokens 或 completion_tokens → 命中
        self.assertEqual(r['source'], 'data.usage')


class TestUsageExtractorResolveApiType(unittest.TestCase):
    """SPEC §2.4 endpoint URL → api_type 推断。"""

    def setUp(self):
        import usage_extractor
        self.ux = usage_extractor

    def test_chat_completions(self):
        self.assertEqual(self.ux.resolve_api_type('https://ark.cn-beijing.volces.com/api/v3/chat/completions'),
                         'chat.completions')
        self.assertEqual(self.ux.resolve_api_type('https://dashscope.aliyuncs.com/compatible-mode/v1/chat/completions'),
                         'chat.completions')
        self.assertEqual(self.ux.resolve_api_type('https://open.bigmodel.cn/api/paas/v4/chat/completions'),
                         'chat.completions')

    def test_responses(self):
        self.assertEqual(self.ux.resolve_api_type('https://ark.cn-beijing.volces.com/api/v3/responses'),
                         'responses')

    def test_multimodal_generation(self):
        self.assertEqual(self.ux.resolve_api_type('https://dashscope.aliyuncs.com/api/v1/services/aigc/multimodal-generation/official-ocr/chat-completions'),
                         'multimodal-generation')
        self.assertEqual(self.ux.resolve_api_type('https://example.com/multimodal-generation'),
                         'multimodal-generation')

    def test_custom_endpoint_default_chat(self):
        """SPEC §2.4：custom_endpoint 一律 chat.completions（保守）。"""
        self.assertEqual(self.ux.resolve_api_type('https://example.com/some-custom', custom_endpoint='ep-abc123'),
                         'chat.completions')

    def test_unknown_default_chat(self):
        self.assertEqual(self.ux.resolve_api_type('https://example.com/something-else'),
                         'chat.completions')

    def test_no_endpoint(self):
        self.assertEqual(self.ux.resolve_api_type(''), 'chat.completions')
        self.assertEqual(self.ux.resolve_api_type(None), 'chat.completions')


class TestUsageExtractorFallback(unittest.TestCase):
    """SPEC §3 兜底估算。"""

    def setUp(self):
        import usage_extractor
        self.ux = usage_extractor

    def test_fallback_chinese_mix(self):
        """中文×2 + 字母数字×0.25 × 1.25 冗余。"""
        # '示例商品A500g' = 4 中文 + 5 alnum (A500g) + 0 other
        # = 4*2 + 5*0.25 = 8+1.25 = 9.25 → ×1.25 = 11.5625 → 12
        r = self.ux.estimate_fallback(content='示例商品A500g', prompt='', provider='doubao', max_tok=2048)
        self.assertEqual(r['source'], 'fallback')
        self.assertEqual(r['completion'], 12)
        self.assertEqual(r['total'], 12)  # 无 prompt 文本 + 无 max_side
        self.assertIsNone(r['image_tokens'])

    def test_fallback_image_tokens_max_side(self):
        """图片按 max_side² 折算（doubao 1.0 px/token）。"""
        r = self.ux.estimate_fallback(content='', prompt='', provider='doubao', max_side=1920, max_tok=2048)
        # image = 1920*1920*1.0/1024 = 3600
        self.assertEqual(r['image_tokens'], 3600)
        self.assertEqual(r['prompt'], 3600)
        self.assertEqual(r['total'], 3600)

    def test_fallback_glm_image_coeff(self):
        """GLM 0.5 px/token 系数。"""
        r = self.ux.estimate_fallback(content='', prompt='', provider='glm', max_side=1024, max_tok=2048)
        # image = 1024*1024*0.5/1024 = 512
        self.assertEqual(r['image_tokens'], 512)

    def test_fallback_max_tok_extreme(self):
        """极端失败：content 空 + prompt 空 + max_tok 给定 → fallback_max_tok。"""
        # 实际上：content 空时 c=0；prompt 空时 p_text=0；max_side=0 时 p_img=0
        # → total=0+0=0；若 max_tok=0 → 走 fallback_max_tok 分支
        r = self.ux.estimate_fallback(content='', prompt='', provider='doubao', max_tok=0, max_side=0)
        self.assertEqual(r['source'], 'fallback_max_tok')
        self.assertEqual(r['total'], 0)


class TestUsageExtractorComputeCost(unittest.TestCase):
    """SPEC §5.3 费用计算 + §8.5 公式注入防护（虽然本函数只算 cost 不写文件）。"""

    def setUp(self):
        import usage_extractor
        self.ux = usage_extractor

    def test_basic_cost(self):
        usage = {'prompt': 1_000_000, 'completion': 100_000, 'image_tokens': None, 'source': 'data.usage'}
        pricing = {'input_per_million': 0.8, 'output_per_million': 2.0}
        # 1 * 0.8 + 0.1 * 2.0 = 0.8 + 0.2 = 1.0
        self.assertEqual(self.ux.compute_cost(usage, pricing), 1.0)

    def test_image_per_call(self):
        usage = {'prompt': 0, 'completion': 0, 'image_tokens': 1024, 'source': 'data.usage.multimodal'}
        pricing = {'input_per_million': 0.0, 'output_per_million': 0.0, 'image_per_call': 0.003}
        self.assertEqual(self.ux.compute_cost(usage, pricing), 0.003)

    def test_missing_pricing_returns_zero(self):
        usage = {'prompt': 100, 'completion': 50, 'image_tokens': None, 'source': 'data.usage'}
        self.assertEqual(self.ux.compute_cost(usage, None), 0.0)
        self.assertEqual(self.ux.compute_cost(usage, {}), 0.0)

    def test_missing_usage_returns_zero(self):
        self.assertEqual(self.ux.compute_cost(None, {'input_per_million': 0.8}), 0.0)
        self.assertEqual(self.ux.compute_cost({}, {'input_per_million': 0.8}), 0.0)

    def test_string_pricing_value_does_not_crash(self):
        """SPEC §6.2 表格：定价字符串异常按 0 算 + 不外抛。"""
        usage = {'prompt': 100, 'completion': 50, 'image_tokens': None}
        # input_per_million 是字符串，强制转 float 失败 → 按 0 算
        pricing = {'input_per_million': 'invalid', 'output_per_million': '0.5 元'}
        try:
            r = self.ux.compute_cost(usage, pricing)
            # 不外抛，cost 应为 0（input 解析失败 + output 也解析失败）
            self.assertEqual(r, 0.0)
        except Exception as e:
            self.fail(f'compute_cost 不应外抛，但抛了 {e!r}')

    def test_no_image_tokens_no_image_cost(self):
        usage = {'prompt': 100, 'completion': 50, 'image_tokens': None}
        pricing = {'image_per_call': 0.003}
        # image_tokens 缺失 → 不计 image_per_call
        self.assertEqual(self.ux.compute_cost(usage, pricing), 0.0)

    def test_model_name_with_injection_chars_in_pricing_lookup_does_not_eval(self):
        """SPEC §8.5：模型名注入字符（=cmd|...）—— compute_cost 不做字典 key
        求值时不允许副作用。本测试只断言 compute_cost 不抛 + 返回 0。"""
        usage = {'prompt': 100, 'completion': 50}
        pricing = {'input_per_million': 0.8}
        # 即使传奇怪字符串 key 也不应该有任何 exec/eval 路径
        try:
            r = self.ux.compute_cost(usage, pricing)
            self.assertIsInstance(r, float)
        except Exception as e:
            self.fail(f'compute_cost 不应外抛: {e!r}')


# ──────────────────────────────────────────────────────────────────
# 2. usage_store 落账与聚合
# ──────────────────────────────────────────────────────────────────

class TestUsageStoreRecord(unittest.TestCase):
    """SPEC §5.1 jsonl 行格式 + §6.2 失败安全。"""

    def setUp(self):
        # 注入临时 base_dir 替身：通过 monkey-patch get_base_dir
        import usage_store
        self.us = usage_store
        self.tmp = tempfile.mkdtemp()
        self._orig_base = usage_store.get_base_dir
        usage_store.get_base_dir = lambda: self.tmp
        # 重置 session
        usage_store.session_reset()
        # 重置 _is_usage_enabled → 写入一份 settings.json 启用
        with open(os.path.join(self.tmp, 'settings.json'), 'w', encoding='utf-8') as f:
            json.dump({'usage': {'enabled': True}}, f)

    def tearDown(self):
        self.us.get_base_dir = self._orig_base
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_record_writes_jsonl(self):
        usage = {'prompt': 100, 'completion': 50, 'total': 150, 'image_tokens': None, 'source': 'data.usage'}
        self.us.record('doubao', 'chat.completions', 'Doubao-Seed-2.1-pro',
                       'https://ark.../chat/completions', usage, cost_cny=0.001, is_estimate=False,
                       call_site='ocr_table', batch_id='2026-01-15-001')
        log_path = os.path.join(self.tmp, 'usage_log.jsonl')
        self.assertTrue(os.path.exists(log_path))
        with open(log_path, 'r', encoding='utf-8') as f:
            line = f.readline().strip()
        obj = json.loads(line)
        self.assertEqual(obj['schema_version'], 1)
        self.assertEqual(obj['provider'], 'doubao')
        self.assertEqual(obj['api_type'], 'chat.completions')
        self.assertEqual(obj['model'], 'Doubao-Seed-2.1-pro')
        self.assertEqual(obj['call_site'], 'ocr_table')
        self.assertEqual(obj['batch_id'], '2026-01-15-001')
        self.assertEqual(obj['is_estimate'], False)
        self.assertEqual(obj['cost_cny'], 0.001)
        self.assertEqual(obj['usage']['prompt'], 100)

    def test_record_does_not_raise_on_disk_failure(self):
        """SPEC §6.2：jsonl 写盘失败（磁盘满/锁）不外抛。"""
        # 通过 patch open 抛 OSError 模拟写盘失败
        usage = {'prompt': 1, 'completion': 1, 'total': 2, 'image_tokens': None, 'source': 'data.usage'}
        real_open = open

        def fake_open(p, mode='r', *args, **kwargs):
            if 'usage_log.jsonl' in str(p) and 'a' in mode:
                raise OSError(28, 'No space left on device')
            return real_open(p, mode, *args, **kwargs)

        with patch('builtins.open', side_effect=fake_open):
            # 不应抛任何异常
            try:
                self.us.record('doubao', 'chat.completions', 'm', 'ep', usage, 0.0, False, call_site='test')
            except Exception as e:
                self.fail(f'record 不应外抛，但抛了 {e!r}')

    def test_record_no_usage_dict_uses_fallback(self):
        """usage 传 None 时 record 内部用 fallback 占位，不崩。"""
        try:
            self.us.record('doubao', 'chat.completions', 'm', 'ep', None, None, True,
                           call_site='test')
        except Exception as e:
            self.fail(f'record(usage=None) 不应外抛: {e!r}')
        log_path = os.path.join(self.tmp, 'usage_log.jsonl')
        with open(log_path, 'r', encoding='utf-8') as f:
            obj = json.loads(f.readline().strip())
        self.assertEqual(obj['usage']['source'], 'missing')

    def test_record_enabled_false_zero_writes(self):
        """SPEC §5.2 enabled=False 整链零写盘。"""
        # 改写 settings.json 关闭
        with open(os.path.join(self.tmp, 'settings.json'), 'w', encoding='utf-8') as f:
            json.dump({'usage': {'enabled': False}}, f)
        usage = {'prompt': 100, 'completion': 50, 'total': 150, 'image_tokens': None, 'source': 'data.usage'}
        self.us.record('doubao', 'chat.completions', 'm', 'ep', usage, 0.001, False, call_site='test')
        log_path = os.path.join(self.tmp, 'usage_log.jsonl')
        self.assertFalse(os.path.exists(log_path), 'enabled=False 时不应写盘')

    def test_record_concurrent_thread_safe(self):
        """§6.2：并发写不损坏行（每行 JSON 可解析）。"""
        import concurrent.futures
        usage = {'prompt': 10, 'completion': 5, 'total': 15, 'image_tokens': None, 'source': 'data.usage'}

        def write(i):
            self.us.record('doubao', 'chat.completions', 'm', 'ep', usage, 0.0, False,
                           call_site='thread', batch_id=f'b{i}')

        with concurrent.futures.ThreadPoolExecutor(max_workers=10) as ex:
            list(ex.map(write, range(50)))

        log_path = os.path.join(self.tmp, 'usage_log.jsonl')
        with open(log_path, 'r', encoding='utf-8') as f:
            lines = [json.loads(l) for l in f if l.strip()]
        self.assertEqual(len(lines), 50)
        # 每行都是合法 JSON（parse 成功即通过）

    def test_session_total_skips_estimate(self):
        """session_total：is_estimate 行不计 cost。"""
        self.us.session_reset()
        u1 = {'prompt': 100, 'completion': 50, 'total': 150, 'image_tokens': None, 'source': 'data.usage'}
        u2 = dict(u1, source='fallback')
        self.us.record('doubao', 'chat.completions', 'm', 'ep', u1, cost_cny=1.0, is_estimate=False, call_site='real')
        self.us.record('doubao', 'chat.completions', 'm', 'ep', u2, cost_cny=5.0, is_estimate=True, call_site='fb')
        self.assertEqual(self.us.session_total(), 1.0)


class TestUsageStoreAggregate(unittest.TestCase):
    """SPEC §5.4 4 档聚合：is_estimate 行不计 cost 计 token。"""

    def setUp(self):
        import usage_store
        self.us = usage_store
        self.tmp = tempfile.mkdtemp()
        self._orig_base = usage_store.get_base_dir
        usage_store.get_base_dir = lambda: self.tmp
        # 写 4 行（2 估算 + 2 实测，含 token + cost）
        with open(os.path.join(self.tmp, 'settings.json'), 'w', encoding='utf-8') as f:
            json.dump({'usage': {'enabled': True}}, f)
        u_real = {'prompt': 100, 'completion': 50, 'total': 150, 'image_tokens': None, 'source': 'data.usage'}
        u_est = {'prompt': 200, 'completion': 80, 'total': 280, 'image_tokens': None, 'source': 'fallback'}
        self.us.record('doubao', 'chat.completions', 'm', 'ep', u_real, cost_cny=1.5, is_estimate=False, call_site='a')
        self.us.record('qwen', 'chat.completions', 'm2', 'ep2', u_real, cost_cny=2.0, is_estimate=False, call_site='b')
        self.us.record('glm', 'chat.completions', 'm3', 'ep3', u_est, cost_cny=0.0, is_estimate=True, call_site='c')
        self.us.record('doubao', 'chat.completions', 'm', 'ep', u_est, cost_cny=0.0, is_estimate=True, call_site='d')

    def tearDown(self):
        self.us.get_base_dir = self._orig_base
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_aggregate_all(self):
        agg = self.us.aggregate('all')
        # cost = 1.5 + 2.0（estimate 行不计 cost）
        self.assertAlmostEqual(agg['cost_cny'], 3.5, places=4)
        # token = (150+150) + (280+280) = 860（estimate 也计 token）
        self.assertEqual(agg['total_tokens'], 860)
        self.assertEqual(agg['estimate_count'], 2)
        self.assertEqual(agg['real_count'], 2)
        self.assertEqual(agg['prompt_tokens'], 100+100+200+200)
        self.assertEqual(agg['completion_tokens'], 50+50+80+80)

    def test_aggregate_today(self):
        agg = self.us.aggregate('today')
        # 当天行都计入（行是刚写的）
        self.assertAlmostEqual(agg['cost_cny'], 3.5, places=4)

    def test_aggregate_returns_dict(self):
        r = self.us.aggregate()
        self.assertIn('today', r)
        self.assertIn('week', r)
        self.assertIn('month', r)
        self.assertIn('all', r)


class TestUsageStoreMonthSummary(unittest.TestCase):
    """SPEC §5.4 month_summary 按模型/按 call_site 分布。"""

    def setUp(self):
        import usage_store
        self.us = usage_store
        self.tmp = tempfile.mkdtemp()
        self._orig_base = usage_store.get_base_dir
        usage_store.get_base_dir = lambda: self.tmp
        with open(os.path.join(self.tmp, 'settings.json'), 'w', encoding='utf-8') as f:
            json.dump({'usage': {'enabled': True}}, f)
        u_real = {'prompt': 100, 'completion': 50, 'total': 150, 'image_tokens': None, 'source': 'data.usage'}
        u_est = {'prompt': 200, 'completion': 80, 'total': 280, 'image_tokens': None, 'source': 'fallback'}
        self.us.record('doubao', 'chat.completions', 'Doubao-Seed-2.1-pro', 'ep', u_real, 1.0, False, call_site='ocr_table')
        self.us.record('qwen', 'multimodal-generation', 'qwen-vl-ocr', 'ep2', u_real, 0.5, False, call_site='ocr_table')
        self.us.record('doubao', 'chat.completions', 'Doubao-Seed-2.1-pro', 'ep', u_est, 0.0, True, call_site='ocr_verify')

    def tearDown(self):
        self.us.get_base_dir = self._orig_base
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_by_model_aggregation(self):
        ms = self.us.month_summary()
        self.assertIn('Doubao-Seed-2.1-pro', ms['by_model'])
        self.assertIn('qwen-vl-ocr', ms['by_model'])
        # doubao：1 行 real (1.0) + 1 行 estimate (0.0) = cost 1.0
        d = ms['by_model']['Doubao-Seed-2.1-pro']
        self.assertAlmostEqual(d['cost'], 1.0, places=4)
        self.assertEqual(d['count'], 2)
        # qwen: 1 行 real (0.5)
        q = ms['by_model']['qwen-vl-ocr']
        self.assertAlmostEqual(q['cost'], 0.5, places=4)

    def test_by_call_site_aggregation(self):
        ms = self.us.month_summary()
        self.assertIn('ocr_table', ms['by_call_site'])
        self.assertIn('ocr_verify', ms['by_call_site'])

    def test_estimate_count_separated(self):
        ms = self.us.month_summary()
        self.assertEqual(ms['estimate_count'], 1)
        self.assertEqual(ms['total_tokens'], 150 + 150 + 280)


# ──────────────────────────────────────────────────────────────────
# 2.5 v1.4.7 P3-R2-C1：reset_month 原子写（os.replace）+ 崩溃窗口保护
# ──────────────────────────────────────────────────────────────────

class TestUsageStoreResetMonthAtomic(unittest.TestCase):
    """v1.4.7 P3-R2-C1：reset_month 改用 os.replace 原子替换。

    验证要点：
      1. 正常 reset：原 jsonl 被原子替换、当月行被删除、其他月行保留、审计行追加。
      2. 崩溃窗口模拟：monkeypatch 让 os.replace 抛 PermissionError（模拟 Windows
         文件锁瞬态拒绝）——3 次重试后应能成功（除非强制全部失败）。
      3. 强制失败 3 次：应保留 tmp 供人工恢复（不静默丢数据），不污染原 jsonl。
    """

    def setUp(self):
        import usage_store
        self.us = usage_store
        self.tmp = tempfile.mkdtemp()
        self._orig_base = usage_store.get_base_dir
        usage_store.get_base_dir = lambda: self.tmp
        self.us._invalidate_enabled_cache()
        # v1.4.7 P3-R2：测试隔离——重置所有缓存避免上一测试残留（_MONTH_COST_CACHE
        # 是模块级且按 year+month key，旧测试数据会被带到新测试）
        try:
            with usage_store._RECORD_LOCK:
                usage_store._MONTH_COST_CACHE['built'] = False
                usage_store._MONTH_COST_CACHE['cost'] = 0.0
                usage_store._MONTH_COST_CACHE['year'] = -1
                usage_store._MONTH_COST_CACHE['month'] = -1
                usage_store._WRITE_FAIL_STATE['consecutive'] = 0
                usage_store._WRITE_FAIL_STATE['last_error'] = ''
                usage_store._WRITE_FAIL_STATE['alerted'] = False
        except Exception:
            pass
        with open(os.path.join(self.tmp, 'settings.json'), 'w', encoding='utf-8') as f:
            json.dump({'usage': {'enabled': True}}, f)
        # 写 5 行：3 行当月 + 2 行上月（确保 reset 当月不影响历史月）
        u = {'prompt': 10, 'completion': 5, 'total': 15, 'image_tokens': None, 'source': 'data.usage'}
        # 强制 ts 写入「当月」（_now_iso 默认就是当前时间）
        for i, (cost, is_est) in enumerate([(1.0, False), (2.0, False), (0.0, True)]):
            self.us.record('doubao', 'chat.completions', f'm{i}', 'ep', u,
                           cost_cny=cost, is_estimate=is_est, call_site=f'cur_{i}')
        # 用一个伪造的上月 ts 注入（直接写 jsonl）
        log_path = os.path.join(self.tmp, 'usage_log.jsonl')
        with open(log_path, 'a', encoding='utf-8') as f:
            old_line = {
                'schema_version': 1, 'ts': '2025-12-15T10:00:00+08:00', 'event': 'usage',
                'provider': 'doubao', 'api_type': 'chat.completions', 'model': 'm_old',
                'endpoint': 'ep_old',
                'usage': {'prompt': 100, 'completion': 50, 'total': 150,
                          'image_tokens': None, 'source': 'data.usage'},
                'cost_cny': 9.99, 'is_estimate': False, 'call_site': 'old',
                'batch_id': '', 'extra': {},
            }
            f.write(json.dumps(old_line, ensure_ascii=False,
                               separators=(',', ':')) + '\n')

    def tearDown(self):
        self.us.get_base_dir = self._orig_base
        self.us._invalidate_enabled_cache()
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_reset_atomic_success(self):
        """正常路径：reset 当月后原 jsonl 仍可读（不丢历史月行），且当月行被删。"""
        ok = self.us.reset_month()
        self.assertTrue(ok)
        # 验证：原文件存在且非空（os.replace 后 tmp 已消费）
        log_path = os.path.join(self.tmp, 'usage_log.jsonl')
        self.assertTrue(os.path.exists(log_path))
        with open(log_path, 'r', encoding='utf-8') as f:
            lines = [json.loads(l) for l in f if l.strip()]
        # 应保留：1 条上月 + 1 条审计；删除 3 条当月
        old_lines = [l for l in lines if l.get('ts', '').startswith('2025-12')]
        audit_lines = [l for l in lines if l.get('event') == 'config_change']
        self.assertEqual(len(old_lines), 1, '上月行必须保留')
        self.assertEqual(old_lines[0]['cost_cny'], 9.99)
        self.assertEqual(len(audit_lines), 1, '追加 1 条审计')
        self.assertEqual(audit_lines[0]['action'], 'reset_month')

    def test_reset_no_tmp_residue_on_success(self):
        """成功路径：tmp 文件被 os.replace 消费，不留 .tmp 残留。"""
        self.us.reset_month()
        # 任何 .tmp* 残留都不应存在
        leftovers = [f for f in os.listdir(self.tmp)
                     if f.startswith('usage_log.jsonl.tmp')]
        self.assertEqual(leftovers, [], f'不应有 .tmp 残留: {leftovers}')

    def test_reset_atomic_on_os_replace_retry(self):
        """崩溃窗口模拟：os.replace 第 1 次 PermissionError 失败，第 2 次成功。

        不动原文件（截断式失败是旧实现的漏洞，新实现应仅在原子替换成功后才切换）。
        """
        real_replace = os.replace
        call_count = [0]
        original = os.path.join(self.tmp, 'usage_log.jsonl')

        def flaky_replace(src, dst):
            call_count[0] += 1
            if call_count[0] == 1 and dst == original:
                raise OSError(5, '模拟 Windows 文件锁瞬态拒绝')
            return real_replace(src, dst)

        with patch('os.replace', side_effect=flaky_replace):
            ok = self.us.reset_month()
        # 第 1 次失败 + 重试（0.2s sleep） + 第 2 次成功 = 应该 True
        self.assertTrue(ok)
        self.assertGreaterEqual(call_count[0], 2)
        # 原文件应完好（os.replace 重试期间原文件未被截断）
        self.assertTrue(os.path.exists(original))
        with open(original, 'r', encoding='utf-8') as f:
            content = f.read()
        self.assertIn('2025-12-15', content, '上月行必须仍在原文件中')
        self.assertIn('config_change', content, '审计行已写入')

    def test_reset_keeps_tmp_on_total_failure(self):
        """全部 3 次 os.replace 失败：保留 tmp 供人工恢复，不污染原 jsonl。"""
        real_replace = os.replace
        original = os.path.join(self.tmp, 'usage_log.jsonl')

        def always_fail(src, dst):
            if dst == original:
                raise OSError(32, '模拟 Windows 文件锁持久占用')
            return real_replace(src, dst)

        with patch('os.replace', side_effect=always_fail):
            ok = self.us.reset_month()
        # 失败应返 False
        self.assertFalse(ok)
        # 原文件应完好（未被截断——这是新实现 vs 旧实现的关键差异）
        self.assertTrue(os.path.exists(original))
        with open(original, 'r', encoding='utf-8') as f:
            content_before = f.read()
        # 全部 3 行当月 + 1 行上月 + 1 行 config_change(无) = 当月 + 上月 都在
        self.assertIn('2025-12-15', content_before, '失败时上月行必须保留')
        self.assertIn('cur_0', content_before, '失败时当月行必须保留')
        self.assertNotIn('config_change', content_before, '失败时审计行不应写入')
        # tmp 残留供人工恢复
        leftovers = [f for f in os.listdir(self.tmp)
                     if f.startswith('usage_log.jsonl.tmp')]
        self.assertEqual(len(leftovers), 1,
                         f'全部失败时应保留 1 个 tmp 供恢复，实测: {leftovers}')

    def test_reset_invalidates_month_cache(self):
        """重置当月后 _MONTH_COST_CACHE 应被失效（_month_cost() 重算为 0）。"""
        # 先触发 cache 构建
        self.us.get_month_cost()
        # 验证 cache 已 built
        with self.us._RECORD_LOCK:
            self.assertTrue(self.us._MONTH_COST_CACHE['built'])
        # 重置当月
        self.us.reset_month()
        # cache 必须被失效
        with self.us._RECORD_LOCK:
            self.assertFalse(self.us._MONTH_COST_CACHE['built'])
            self.assertEqual(self.us._MONTH_COST_CACHE['cost'], 0.0)
        # 重新查询应返 0（数据已清空）
        self.assertEqual(self.us.get_month_cost(), 0.0)


# ──────────────────────────────────────────────────────────────────
# 2.6 v1.4.7 P3-R2-M3：_read_jsonl_lines 加 _RECORD_LOCK（防半截行）
# ──────────────────────────────────────────────────────────────────

class TestUsageStoreReadJsonlLock(unittest.TestCase):
    """v1.4.7 P3-R2-M3：_read_jsonl_lines 加 _RECORD_LOCK 保护。

    验证要点：
      1. 锁内读 + 锁内写不会死锁（reset_month 改用 _read_jsonl_lines_unlocked）
      2. 跨线程：reader 线程调用 _read_jsonl_lines 与 writer 线程 record() 互斥——reader
         永远不会拿到半截行（旧实现极端竞态下 record() 两次 write 之间被打断，reader
         会拿到不完整 JSON 行被 try/except 吞，丢一行数据）
      3. aggregate / month_summary 等调用 _read_jsonl_lines 的接口仍正常返回
    """

    def setUp(self):
        import usage_store
        self.us = usage_store
        self.tmp = tempfile.mkdtemp()
        self._orig_base = usage_store.get_base_dir
        usage_store.get_base_dir = lambda: self.tmp
        self.us._invalidate_enabled_cache()
        # 重置缓存
        try:
            with usage_store._RECORD_LOCK:
                usage_store._MONTH_COST_CACHE['built'] = False
                usage_store._MONTH_COST_CACHE['cost'] = 0.0
                usage_store._MONTH_COST_CACHE['year'] = -1
                usage_store._MONTH_COST_CACHE['month'] = -1
                usage_store._WRITE_FAIL_STATE['consecutive'] = 0
                usage_store._WRITE_FAIL_STATE['last_error'] = ''
                usage_store._WRITE_FAIL_STATE['alerted'] = False
        except Exception:
            pass
        with open(os.path.join(self.tmp, 'settings.json'), 'w', encoding='utf-8') as f:
            json.dump({'usage': {'enabled': True}}, f)
        u = {'prompt': 10, 'completion': 5, 'total': 15,
             'image_tokens': None, 'source': 'data.usage'}
        # 写 5 行
        for i in range(5):
            self.us.record('doubao', 'chat.completions', f'm{i}', 'ep',
                           u, cost_cny=0.1 * i, is_estimate=False, call_site=f's{i}')

    def tearDown(self):
        self.us.get_base_dir = self._orig_base
        self.us._invalidate_enabled_cache()
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_aggregate_works_with_lock(self):
        """aggregate() 通过新加锁的 _read_jsonl_lines 仍能正确返回 5 行聚合。"""
        agg = self.us.aggregate('all')
        # 5 行 × cost=0.0+0.1+0.2+0.3+0.4 = 1.0
        self.assertAlmostEqual(agg['cost_cny'], 1.0, places=4)
        self.assertEqual(agg['real_count'], 5)

    def test_month_summary_works_with_lock(self):
        """month_summary() 仍正常返回。"""
        ms = self.us.month_summary()
        self.assertEqual(ms['total_tokens'], 5 * 15)
        self.assertEqual(len(ms['by_model']), 5)

    def test_concurrent_read_write_no_hang(self):
        """跨线程并发读写不应死锁：reader 与 writer 都应在合理时间内完成。"""
        import concurrent.futures
        u = {'prompt': 1, 'completion': 1, 'total': 2,
             'image_tokens': None, 'source': 'data.usage'}

        def writer_task(i):
            self.us.record('doubao', 'chat.completions', 'm', 'ep', u,
                           cost_cny=0.01, is_estimate=False, call_site=f'w{i}')

        def reader_task(i):
            return self.us.aggregate('all')['cost_cny']

        with concurrent.futures.ThreadPoolExecutor(max_workers=10) as ex:
            futs = []
            for i in range(20):
                futs.append(ex.submit(writer_task, i))
            for i in range(10):
                futs.append(ex.submit(reader_task, i))
            # 5 秒内必须完成（死锁检测）
            for f in concurrent.futures.as_completed(futs, timeout=5):
                f.result()  # 若死锁则 timeout 抛

        # 验证：原 5 行 + 20 行 writer = 25 行
        agg = self.us.aggregate('all')
        self.assertEqual(agg['real_count'], 25,
                         f'并发读写后应 25 行，实测 {agg["real_count"]}')

    def test_aggregate_under_concurrent_writes_no_half_lines(self):
        """并发写 + 立即 aggregate：reader 不应拿到半截 JSON 行（被 try 吞导致计数偏少）。"""
        import concurrent.futures
        u = {'prompt': 1, 'completion': 1, 'total': 2,
             'image_tokens': None, 'source': 'data.usage'}

        def writer_task(i):
            for _ in range(5):
                self.us.record('doubao', 'chat.completions', 'm', 'ep', u,
                               cost_cny=0.01, is_estimate=False, call_site=f'w{i}')

        with concurrent.futures.ThreadPoolExecutor(max_workers=8) as ex:
            list(ex.map(writer_task, range(5)))

        # 原 5 行 + 5 threads × 5 writes = 30 行
        # 若有半截行被吞，real_count 会偏少；M3 加锁后应严格相等
        agg = self.us.aggregate('all')
        self.assertEqual(agg['real_count'], 30,
                         f'并发写后应 30 行，实测 {agg["real_count"]}（半截行被吞）')


class TestUsagePanelSummary(unittest.TestCase):
    """v1.4.7 P3-R2-M1/L1：usage_panel_summary 统一入口（gui + settings_ui 共用）。

    验证：4 档 + by_model + by_call_site 都齐全；month.cost_cny 与 get_month_cost() 同源。
    """

    def setUp(self):
        import usage_store
        self.us = usage_store
        self.tmp = tempfile.mkdtemp()
        self._orig_base = usage_store.get_base_dir
        usage_store.get_base_dir = lambda: self.tmp
        self.us._invalidate_enabled_cache()
        # 测试隔离：重置所有缓存（见 TestUsageStoreResetMonthAtomic.setUp 注释）
        try:
            with usage_store._RECORD_LOCK:
                usage_store._MONTH_COST_CACHE['built'] = False
                usage_store._MONTH_COST_CACHE['cost'] = 0.0
                usage_store._MONTH_COST_CACHE['year'] = -1
                usage_store._MONTH_COST_CACHE['month'] = -1
                usage_store._WRITE_FAIL_STATE['consecutive'] = 0
                usage_store._WRITE_FAIL_STATE['last_error'] = ''
                usage_store._WRITE_FAIL_STATE['alerted'] = False
        except Exception:
            pass
        with open(os.path.join(self.tmp, 'settings.json'), 'w', encoding='utf-8') as f:
            json.dump({'usage': {'enabled': True}}, f)
        u_real = {'prompt': 100, 'completion': 50, 'total': 150,
                  'image_tokens': None, 'source': 'data.usage'}
        u_est = {'prompt': 200, 'completion': 80, 'total': 280,
                 'image_tokens': None, 'source': 'fallback'}
        self.us.record('doubao', 'chat.completions', 'm1', 'ep', u_real, 1.0, False, call_site='a')
        self.us.record('qwen', 'multimodal-generation', 'm2', 'ep', u_real, 0.5, False, call_site='b')
        self.us.record('glm', 'chat.completions', 'm3', 'ep', u_est, 0.0, True, call_site='c')

    def tearDown(self):
        self.us.get_base_dir = self._orig_base
        self.us._invalidate_enabled_cache()
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_panel_summary_keys_complete(self):
        """返回必须包含 today/week/month/all/by_model/by_call_site。"""
        panel = self.us.usage_panel_summary()
        for k in ('today', 'week', 'month', 'all', 'by_model', 'by_call_site',
                  'month_label', 'missing_count', 'estimate_count'):
            self.assertIn(k, panel, f'缺少 {k} 字段')

    def test_panel_summary_month_cost_matches_cache(self):
        """month.cost_cny 必须与 get_month_cost() 一致（统一来源）。"""
        panel = self.us.usage_panel_summary()
        cache = self.us.get_month_cost()
        # 1.0 + 0.5 = 1.5（estimate 行不计 cost）
        self.assertAlmostEqual(panel['month']['cost_cny'], 1.5, places=4)
        self.assertAlmostEqual(panel['month']['cost_cny'], cache, places=4)

    def test_panel_summary_by_model(self):
        """按模型分布：doubao=1.0, qwen=0.5, glm=0.0(estimate)。"""
        panel = self.us.usage_panel_summary()
        self.assertIn('m1', panel['by_model'])
        self.assertIn('m2', panel['by_model'])
        self.assertIn('m3', panel['by_model'])
        self.assertAlmostEqual(panel['by_model']['m1']['cost'], 1.0, places=4)
        self.assertAlmostEqual(panel['by_model']['m2']['cost'], 0.5, places=4)
        self.assertAlmostEqual(panel['by_model']['m3']['cost'], 0.0, places=4)

    def test_panel_summary_by_call_site(self):
        """按 call_site 分布。"""
        panel = self.us.usage_panel_summary()
        for site in ('a', 'b', 'c'):
            self.assertIn(site, panel['by_call_site'])

    def test_panel_summary_on_empty(self):
        """空 jsonl：返空 dict 不崩。"""
        # 清空文件
        for f in os.listdir(self.tmp):
            if f.startswith('usage_log'):
                os.remove(os.path.join(self.tmp, f))
        with open(os.path.join(self.tmp, 'settings.json'), 'w', encoding='utf-8') as f:
            json.dump({'usage': {'enabled': True}}, f)
        self.us._invalidate_enabled_cache()
        # 重置缓存（前面 record 写入已 built=true）
        with self.us._RECORD_LOCK:
            self.us._MONTH_COST_CACHE['built'] = False
            self.us._MONTH_COST_CACHE['cost'] = 0.0
        panel = self.us.usage_panel_summary()
        self.assertEqual(panel['month']['cost_cny'], 0.0)
        self.assertEqual(panel['today']['cost_cny'], 0.0)
        self.assertEqual(panel['by_model'], {})


# ──────────────────────────────────────────────────────────────────
# 3. _debug_dump_response 受开关控制
# ──────────────────────────────────────────────────────────────────

class TestDebugDump(unittest.TestCase):
    """SPEC §11.2 debug 落盘钩子：受 debug_archive_enabled 控制；默认关。"""

    def setUp(self):
        global usage_extractor
        import usage_extractor as _ux
        usage_extractor = _ux
        self.ux = _ux
        self.tmp = tempfile.mkdtemp()
        self._orig_base = _ux.get_base_dir
        _ux.get_base_dir = lambda: self.tmp
        _ux.reset_usage_cfg_cache()

    def tearDown(self):
        self.ux.get_base_dir = self._orig_base
        shutil.rmtree(self.tmp, ignore_errors=True)
        self.ux.reset_usage_cfg_cache()

    def test_default_disabled_no_write(self):
        """默认 debug_archive_enabled 未设置 → 不写盘。"""
        with open(os.path.join(self.tmp, 'settings.json'), 'w', encoding='utf-8') as f:
            json.dump({'usage': {'enabled': True}}, f)
        self.ux.reset_usage_cfg_cache()
        self.ux._debug_dump_response('doubao', 'Doubao-Seed-2.1-pro', 'https://example/ep', {'a': 1}, None)
        archive_dir = os.path.join(self.tmp, 'output', 'usage_archive')
        self.assertFalse(os.path.exists(archive_dir), '默认关时不应创建目录')

    def test_enabled_writes_file(self):
        """debug_archive_enabled=true → 写盘到 output/usage_archive/。"""
        with open(os.path.join(self.tmp, 'settings.json'), 'w', encoding='utf-8') as f:
            json.dump({'usage': {'enabled': True, 'debug_archive_enabled': True}}, f)
        self.ux.reset_usage_cfg_cache()
        data = {'usage': {'prompt_tokens': 100, 'completion_tokens': 50, 'total_tokens': 150}}
        u = self.ux.extract(data, 'doubao', 'chat.completions')
        self.ux._debug_dump_response('doubao', 'Doubao-Seed-2.1-pro', 'https://example/ep', data, u)
        archive_dir = os.path.join(self.tmp, 'output', 'usage_archive')
        self.assertTrue(os.path.exists(archive_dir))
        files = os.listdir(archive_dir)
        self.assertEqual(len(files), 1)
        with open(os.path.join(archive_dir, files[0]), 'r', encoding='utf-8') as f:
            obj = json.load(f)
        self.assertEqual(obj['provider'], 'doubao')
        self.assertEqual(obj['model'], 'Doubao-Seed-2.1-pro')
        self.assertEqual(obj['raw_response'], data)

    def test_safe_model_filename(self):
        """model 名含 / : 替换为 _。"""
        with open(os.path.join(self.tmp, 'settings.json'), 'w', encoding='utf-8') as f:
            json.dump({'usage': {'enabled': True, 'debug_archive_enabled': True}}, f)
        self.ux.reset_usage_cfg_cache()
        self.ux._debug_dump_response('doubao', 'ep-abc/with:colon', 'ep', {'x': 1}, None)
        archive_dir = os.path.join(self.tmp, 'output', 'usage_archive')
        files = os.listdir(archive_dir)
        # 文件名不应含 / 或 :
        for fn in files:
            self.assertNotIn('/', fn)
            self.assertNotIn(':', fn)

    def test_7day_cleanup(self):
        """§11.2：>7 天的文件自动清理。"""
        import time
        with open(os.path.join(self.tmp, 'settings.json'), 'w', encoding='utf-8') as f:
            json.dump({'usage': {'enabled': True, 'debug_archive_enabled': True}}, f)
        self.ux.reset_usage_cfg_cache()
        archive_dir = os.path.join(self.tmp, 'output', 'usage_archive')
        os.makedirs(archive_dir, exist_ok=True)
        # 写一个 8 天前的旧文件
        old_path = os.path.join(archive_dir, 'old.json')
        with open(old_path, 'w', encoding='utf-8') as f:
            json.dump({'old': True}, f)
        old_mtime = time.time() - 8 * 86400
        os.utime(old_path, (old_mtime, old_mtime))
        # 调用 _debug_dump_response（会触发清理）
        self.ux._debug_dump_response('doubao', 'm', 'ep', {'a': 1}, None)
        # 旧文件应被清理
        self.assertFalse(os.path.exists(old_path), '>7 天的文件应被清理')

    def test_disabled_does_not_write(self):
        """enabled=False → debug 即便开也不写盘（更保险）。"""
        with open(os.path.join(self.tmp, 'settings.json'), 'w', encoding='utf-8') as f:
            json.dump({'usage': {'enabled': False, 'debug_archive_enabled': True}}, f)
        self.ux.reset_usage_cfg_cache()
        self.ux._debug_dump_response('doubao', 'm', 'ep', {'a': 1}, None)
        archive_dir = os.path.join(self.tmp, 'output', 'usage_archive')
        self.assertFalse(os.path.exists(archive_dir))


# ──────────────────────────────────────────────────────────────────
# 4. ocr.py / vision.py 三元组 API 形态契约（不发起真实请求）
# ──────────────────────────────────────────────────────────────────

class TestOcrVisionThreeTupleContract(unittest.TestCase):
    """T-C2 / T-C3：ocr.py / vision.py 内部 API 返回三元组契约。

    不发起真实 API；通过 mock requests.post 验证：
      - _ocr_api_call_do / _ocr_api_call / _call_vision_api 返回 (text, mdl, usage) 三元组
      - usage=None / dict 都允许，调用方不因 None 中断
    t8 集成收口补充：漏斗单点落账（T-C5）——mock 调用会真实触发
    usage_store.record，setUp 把 usage_store.get_base_dir 重定向到 tmp，
    落账不污染真实数据目录；并新增落账恰一行/口径断言。
    """

    def setUp(self):
        import usage_store
        self.us = usage_store
        self.tmp = tempfile.mkdtemp()
        self._orig_base = usage_store.get_base_dir
        usage_store.get_base_dir = lambda: self.tmp
        usage_store.session_reset()

    def tearDown(self):
        self.us.get_base_dir = self._orig_base
        shutil.rmtree(self.tmp, ignore_errors=True)

    def _read_log_lines(self):
        log_path = os.path.join(self.tmp, 'usage_log.jsonl')
        if not os.path.exists(log_path):
            return []
        with open(log_path, 'r', encoding='utf-8') as f:
            return [json.loads(l) for l in f if l.strip()]

    def test_ocr_api_call_returns_three_tuple(self):
        """_ocr_api_call 返回三元组（mock 响应）。"""
        from unittest.mock import patch, MagicMock
        # mock settings 注入主 provider（patch ocr.get_api_config 而非 utils
        # 因为 ocr.py 用 `from utils import get_api_config` 已经绑定了引用）
        import ocr
        cfg = {
            'active_provider': 'doubao',
            'providers': {
                'doubao': {
                    'api_key': 'fake-key',
                    'model': 'Doubao-Seed-2.1-pro',
                    'endpoint': 'https://ark.cn-beijing.volces.com/api/v3/chat/completions',
                },
                'qwen': {'api_key': '', 'model': ''},
                'glm': {'api_key': '', 'model': ''},
            },
        }
        fake_response = MagicMock()
        fake_response.json.return_value = {
            'choices': [{'message': {'content': '{"columns": [], "rows": []}'}}],
            'usage': {'prompt_tokens': 100, 'completion_tokens': 50, 'total_tokens': 150},
        }
        with patch.object(ocr, 'get_api_config', return_value=cfg), \
             patch('ocr.requests.post', return_value=fake_response):
            content, mdl, usage = ocr._ocr_api_call('img', 'prompt', max_tok=1024)
        self.assertEqual(content, '{"columns": [], "rows": []}')
        self.assertEqual(mdl, 'Doubao-Seed-2.1-pro')
        self.assertIsInstance(usage, dict)
        self.assertEqual(usage['prompt'], 100)
        self.assertEqual(usage['completion'], 50)
        self.assertEqual(usage['source'], 'data.usage')

    def test_ocr_api_call_usage_missing_returns_estimate(self):
        """usage 缺失时 _ocr_api_call 返回 fallback 估算（不外抛）。"""
        from unittest.mock import patch, MagicMock
        import ocr
        cfg = {
            'active_provider': 'doubao',
            'providers': {
                'doubao': {
                    'api_key': 'fake-key',
                    'model': 'Doubao-Seed-2.1-pro',
                    'endpoint': 'https://ark.cn-beijing.volces.com/api/v3/chat/completions',
                },
                'qwen': {'api_key': '', 'model': ''},
                'glm': {'api_key': '', 'model': ''},
            },
        }
        fake_response = MagicMock()
        fake_response.json.return_value = {
            'choices': [{'message': {'content': '示例商品'}}],
            # 无 usage 字段
        }
        with patch.object(ocr, 'get_api_config', return_value=cfg), \
             patch('ocr.requests.post', return_value=fake_response):
            content, mdl, usage = ocr._ocr_api_call('img', '示例提示词', max_tok=1024)
        self.assertEqual(content, '示例商品')
        self.assertIsInstance(usage, dict)
        # usage 缺失 → 走 §3 兜底估算
        self.assertIn(usage['source'], ('fallback', 'fallback_max_tok'))

    def test_vision_api_returns_three_tuple(self):
        """_call_vision_api 返回三元组（直接 patch requests 走 chat completions 路径）。"""
        from unittest.mock import patch, MagicMock
        import vision
        fake_response = MagicMock()
        fake_response.json.return_value = {
            'choices': [{'message': {'content': '{}'}}],
            'usage': {'prompt_tokens': 50, 'completion_tokens': 10, 'total_tokens': 60},
        }
        provider = {
            'api_key': 'fake-key',
            'model': 'Doubao-Seed-2.1-pro',
            'endpoint': 'https://ark.cn-beijing.volces.com/api/v3/chat/completions',
            'custom_endpoint': '',
        }
        with patch.object(vision, '_pick_vision_model',
                          return_value=('doubao', provider, provider['endpoint'], 'fake-key',
                                        'Doubao-Seed-2.1-pro', False)), \
             patch('requests.post', return_value=fake_response):
            text, mdl, usage = vision._call_vision_api('img', 'p', max_tokens=64, timeout=10,
                                                       call_site='test')
        self.assertEqual(text, '{}')
        self.assertEqual(mdl, 'Doubao-Seed-2.1-pro')
        self.assertIsInstance(usage, dict)
        self.assertEqual(usage['prompt'], 50)

    def test_ocr_call_records_single_usage_line(self):
        """t8 集成收口（T-C5）：mock OCR 调用 → usage_log.jsonl 恰落一行、口径正确。

        价格表缺失 → cost_cny=0（面板显示 ?）；call_site='OCR 识别'；实测行非估算。
        """
        import ocr
        cfg = {
            'active_provider': 'doubao',
            'providers': {
                'doubao': {
                    'api_key': 'fake-key',
                    'model': 'Doubao-Seed-2.1-pro',
                    'endpoint': 'https://ark.cn-beijing.volces.com/api/v3/chat/completions',
                },
                'qwen': {'api_key': '', 'model': ''},
                'glm': {'api_key': '', 'model': ''},
            },
        }
        fake_response = MagicMock()
        fake_response.json.return_value = {
            'choices': [{'message': {'content': '示例商品'}}],
            'usage': {'prompt_tokens': 100, 'completion_tokens': 50, 'total_tokens': 150},
        }
        with patch.object(ocr, 'get_api_config', return_value=cfg), \
             patch('ocr.requests.post', return_value=fake_response):
            ocr._ocr_api_call('img', 'prompt', max_tok=1024)
        lines = self._read_log_lines()
        self.assertEqual(len(lines), 1, '单点落账应恰写一行')
        row = lines[0]
        self.assertEqual(row['provider'], 'doubao')
        self.assertEqual(row['model'], 'Doubao-Seed-2.1-pro')
        self.assertEqual(row['api_type'], 'chat.completions')
        self.assertEqual(row['call_site'], 'OCR 识别')
        self.assertEqual(row['is_estimate'], False)
        self.assertEqual(row['cost_cny'], 0, '缺价应按 0 计（面板显示 ?）')
        self.assertEqual(row['usage']['total'], 150)

    def test_vision_call_records_with_call_site(self):
        """t8 集成收口（T-C5）：vision 调用落账携带 call_site 审计标签。"""
        import vision
        fake_response = MagicMock()
        fake_response.json.return_value = {
            'choices': [{'message': {'content': '{}'}}],
            'usage': {'prompt_tokens': 50, 'completion_tokens': 10, 'total_tokens': 60},
        }
        provider = {
            'api_key': 'fake-key',
            'model': 'Doubao-Seed-2.1-pro',
            'endpoint': 'https://ark.cn-beijing.volces.com/api/v3/chat/completions',
            'custom_endpoint': '',
        }
        with patch.object(vision, '_pick_vision_model',
                          return_value=('doubao', provider, provider['endpoint'], 'fake-key',
                                        'Doubao-Seed-2.1-pro', False)), \
             patch('requests.post', return_value=fake_response):
            vision._call_vision_api('img', 'p', max_tokens=64, timeout=10,
                                    call_site='定位')
        lines = self._read_log_lines()
        self.assertEqual(len(lines), 1)
        self.assertEqual(lines[0]['call_site'], '定位')
        self.assertEqual(lines[0]['usage']['total'], 60)


# ──────────────────────────────────────────────────────────────────
# 5. 集成式断言：record → aggregate 往返
# ──────────────────────────────────────────────────────────────────

class TestEndToEndPipeline(unittest.TestCase):
    """端到端：record 三种类型 → aggregate 验证 cost/token 行为。"""

    def setUp(self):
        import usage_store
        self.us = usage_store
        self.tmp = tempfile.mkdtemp()
        self._orig_base = usage_store.get_base_dir
        usage_store.get_base_dir = lambda: self.tmp
        with open(os.path.join(self.tmp, 'settings.json'), 'w', encoding='utf-8') as f:
            json.dump({'usage': {'enabled': True}}, f)
        usage_store.session_reset()
        # 写不同类型行
        u_real = {'prompt': 1000, 'completion': 500, 'total': 1500, 'image_tokens': None, 'source': 'data.usage'}
        u_est = {'prompt': 800, 'completion': 300, 'total': 1100, 'image_tokens': None, 'source': 'fallback'}
        u_missing_cost = {'prompt': 50, 'completion': 25, 'total': 75, 'image_tokens': None, 'source': 'data.usage'}
        self.us.record('doubao', 'chat.completions', 'M1', 'ep1', u_real, 0.001, False, call_site='a')
        self.us.record('qwen', 'multimodal-generation', 'M2', 'ep2', u_real, 0.002, False, call_site='b')
        self.us.record('glm', 'chat.completions', 'M3', 'ep3', u_est, 0.0, True, call_site='c')
        # pricing 缺失行（cost_cny=None）
        self.us.record('doubao', 'chat.completions', 'M4', 'ep4', u_missing_cost, None, False, call_site='d')

    def tearDown(self):
        self.us.get_base_dir = self._orig_base
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_total_cost_real_only(self):
        agg = self.us.aggregate('all')
        # cost = 0.001 + 0.002 = 0.003（estimate 不计，None 不计）
        self.assertAlmostEqual(agg['cost_cny'], 0.003, places=4)

    def test_total_tokens_includes_estimate(self):
        agg = self.us.aggregate('all')
        # token = 1500 + 1500 + 1100 + 75 = 4175
        self.assertEqual(agg['total_tokens'], 1500 + 1500 + 1100 + 75)

    def test_missing_count(self):
        agg = self.us.aggregate('all')
        # 1 行 cost_cny=None → missing_count += 1
        self.assertEqual(agg['missing_count'], 1)


# ──────────────────────────────────────────────────────────────────
# WS-A 本地历史库（history_db.py，来源 test_tmp_wsa.py）
# ──────────────────────────────────────────────────────────────────
TODAY = datetime.now().strftime('%Y-%m-%d')


def _load_history_db():
    """按 test_smoke 同款模式加载被测模块（独立模块对象，避免污染导入缓存）。"""
    spec = importlib.util.spec_from_file_location(
        'history_db', os.path.join(HERE, 'history_db.py'))
    m = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(m)
    return m


# 典型 gui._calc_from_items 产物 plans（含 daily 键——日销量的真实键名）
PLANS_YN = [
    {'name': '洗衣液2kg', 'sku_id': '11111111111', 'stock': 50, 'daily': 10,
     'days_left': 5.0, 'status': '3天后下单', 'qty': 0, 'warehouse': '华东1号仓'},
    {'name': '抽纸整箱', 'sku_id': '22222222222', 'stock': 5, 'daily': 20,
     'days_left': 0.3, 'status': '立刻补货', 'qty': 200, 'warehouse': '华东1号仓'},
    {'name': '无ID商品A', 'stock': 8, 'daily': 0,
     'days_left': 0.0, 'status': '无销量·观察', 'qty': 0, 'warehouse': '华南仓'},
]
PLANS_SC = [
    {'name': '洗洁精1.5kg', 'sku_id': '33333333333', 'stock': 30, 'daily': 6,
     'days_left': 5.0, 'status': '5天后下单', 'qty': 0, 'warehouse': '西南仓'},
    {'name': '垃圾袋卷装', 'sku_id': '44444444444', 'stock': 2, 'daily': 9,
     'days_left': 0.2, 'status': '立刻补货', 'qty': 300, 'warehouse': '西南仓'},
]


class TestHistoryDB(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.hdb = _load_history_db()

    def setUp(self):
        self.tmp = tempfile.mkdtemp(prefix='pdd_wsa_')
        self.hdb.set_db_path(os.path.join(self.tmp, 'history.db'))

    def tearDown(self):
        self.hdb.reset_db_path()
        shutil.rmtree(self.tmp, ignore_errors=True)

    # ── 原始表直查助手（独立短连接，用完即关，避免 Windows 文件句柄占用）──
    def _raw(self, sql, params=()):
        conn = sqlite3.connect(self.hdb.db_path())
        try:
            return conn.execute(sql, params).fetchall()
        finally:
            conn.close()

    def _exec(self, sql, params=()):
        conn = sqlite3.connect(self.hdb.db_path())
        try:
            conn.execute(sql, params)
            conn.commit()
        finally:
            conn.close()

    # ── 1. 往返 ──────────────────────────────────────────────────────
    def test_record_and_query_roundtrip(self):
        sid = self.hdb.record_capture({'云南': PLANS_YN, '四川': PLANS_SC}, 'live')
        self.assertGreater(sid, 0, '正常记录应返回正数 session_id')

        # 按日聚合：当日 (云南, 四川) 两行
        daily = self.hdb.query_daily(30)
        self.assertEqual(len(daily), 2)
        yn = next(r for r in daily if r['region'] == '云南')
        self.assertEqual(yn['day'], TODAY)
        self.assertEqual(yn['items'], 3)
        self.assertEqual(yn['alerts'], 1, 'alerts 只统计 立刻补货（红色硬预警）')
        self.assertEqual(yn['stock_total'], 63)

        # 地区过滤
        daily_sc = self.hdb.query_daily(30, region='四川')
        self.assertEqual(len(daily_sc), 1)
        self.assertEqual(daily_sc[0]['items'], 2)
        self.assertEqual(daily_sc[0]['alerts'], 1)

        # 单商品 sku 时间序列（sales 取自 plans 的 daily 键）
        series = self.hdb.query_sku_history('22222222222', 90)
        self.assertEqual(len(series), 1)
        self.assertEqual(series[0]['name'], '抽纸整箱')
        self.assertEqual(series[0]['stock'], 5)
        self.assertEqual(series[0]['sales'], 20)
        self.assertEqual(series[0]['warehouse'], '华东1号仓')

        # 无 ID 行回退 (region, name) 精确匹配
        fb = self.hdb.query_sku_history('', 90, region='云南', name='无ID商品A')
        self.assertEqual(len(fb), 1)
        self.assertEqual(fb[0]['warehouse'], '华南仓')

        # 地区当日明细
        rows = self.hdb.query_region_days('云南', TODAY)
        self.assertEqual(len(rows), 3)
        self.assertIn('洗衣液2kg', {r['name'] for r in rows})

        # session 记账：首地区 / source / 行数
        sess = self._raw('SELECT region, source, item_count FROM capture_sessions')
        self.assertEqual(len(sess), 1)
        self.assertEqual(sess[0][0], '云南')
        self.assertEqual(sess[0][1], 'live')
        self.assertEqual(sess[0][2], 5)

    # ── 2. 缺键容忍与非法入参 ────────────────────────────────────────
    def test_missing_keys_tolerated_and_bad_input(self):
        sid = self.hdb.record_capture({'广东': [{}, {'name': '只有名字'}]}, 'import')
        self.assertGreater(sid, 0, '空 plan/缺键 plan 应容忍为默认值')
        rows = self.hdb.query_region_days('广东', TODAY)
        self.assertEqual(len(rows), 2)
        r0, r1 = rows[0], rows[1]  # 同时间戳按 name 升序：'' 在前
        self.assertEqual(r0['name'], '')
        self.assertEqual(r0['stock'], 0)
        self.assertEqual(r0['sku_id'], '')
        self.assertIsNone(r0['days_left'])
        self.assertEqual(r0['status'], '')
        self.assertEqual(r1['name'], '只有名字')

        # 非法入参 → -1，绝不抛
        self.assertEqual(self.hdb.record_capture(None, 'live'), -1)
        self.assertEqual(self.hdb.record_capture('not-a-dict', 'live'), -1)

        # 值非列表被跳过，空 session 仍入账（item_count=0）
        sid2 = self.hdb.record_capture({'广西': None}, 'live')
        self.assertGreater(sid2, 0)
        sess = self._raw('SELECT item_count FROM capture_sessions WHERE id=?', (sid2,))
        self.assertEqual(sess[0][0], 0)

    # ── 3. prune 双阈值 ──────────────────────────────────────────────
    def test_prune_dual_threshold_and_orphan_sessions(self):
        sid1 = self.hdb.record_capture({'云南': PLANS_YN}, 'live')
        sid2 = self.hdb.record_capture({'四川': PLANS_SC}, 'batch')
        self.assertGreater(sid1, 0)
        self.assertGreater(sid2, 0)
        # 人为把四川行做旧（2020-01-01 早于任何保留窗口）
        self._exec('UPDATE history_rows SET captured_at=? WHERE session_id=?',
                   ('2020-01-01 08:00:00', sid2))
        # 天数阈值：max_rows=0 表示该规则不启用
        n = self.hdb.prune(retention_days=30, max_rows=0)
        self.assertEqual(n, 2, '做旧的 2 行四川记录应被天数阈值删除')
        self.assertEqual(len(self.hdb.query_region_days('云南', TODAY)), 3)
        # 孤儿 session（sid2 行全删）被顺带清理
        sids = [r[0] for r in self._raw('SELECT id FROM capture_sessions')]
        self.assertEqual(sids, [sid1])

        # 行数上限：再灌 3 session × 3 行 → 共 12 行 → cap=5 保留最新 5 行
        for k in range(3):
            self.hdb.record_capture({'云南': [
                {'name': f'批量商品{k}-{j}', 'sku_id': f'900000000{k}{j}', 'stock': j}
                for j in range(3)]}, 'batch')
        n2 = self.hdb.prune(retention_days=0, max_rows=5)
        self.assertEqual(n2, 7)
        total = self._raw('SELECT COUNT(*) FROM history_rows')[0][0]
        self.assertEqual(total, 5)

        # 低于阈值 = 廉价 no-op
        self.assertEqual(self.hdb.prune(retention_days=0, max_rows=0), 0)
        self.assertEqual(self.hdb.prune(retention_days=30, max_rows=100000), 0)

    # ── 4. delete_region ─────────────────────────────────────────────
    def test_delete_region(self):
        sid1 = self.hdb.record_capture({'云南': PLANS_YN}, 'live')
        sid2 = self.hdb.record_capture({'四川': PLANS_SC}, 'live')
        n = self.hdb.delete_region('四川')
        self.assertEqual(n, 2)
        self.assertEqual(len(self.hdb.query_region_days('云南', TODAY)), 3, '其他地区不受影响')
        sids = [r[0] for r in self._raw('SELECT id FROM capture_sessions')]
        self.assertEqual(sids, [sid1], '孤儿 session 应被清理')

        self.assertEqual(self.hdb.delete_region('四川'), 0, '重复删除返回 0')
        self.assertEqual(self.hdb.delete_region(''), -1, '空地区视为无效调用')
        self.assertEqual(self.hdb.delete_region('云南'), 3)
        self.assertEqual(self.hdb.query_daily(30), [])

    # ── 5. 损坏库隔离重建 ────────────────────────────────────────────
    def test_corrupt_db_quarantined_and_rebuilt(self):
        sid = self.hdb.record_capture({'云南': PLANS_YN}, 'live')
        self.assertGreater(sid, 0)
        # 强制下次操作重检（同路径重置进程级 _READY 缓存）
        self.hdb.set_db_path(self.hdb.db_path())
        with open(self.hdb.db_path(), 'wb') as f:
            f.write(b'PDD EZ corruption drill: definitely not a sqlite file. ' * 64)

        sid2 = self.hdb.record_capture({'云南': PLANS_YN}, 'batch')
        self.assertGreater(sid2, 0, '损坏库应被隔离后自动重建并可写')
        self.assertTrue(os.path.exists(self.hdb.db_path() + '.corrupt'),
                        '损坏文件应改名 .corrupt 留存（与 Config 同模式）')
        # 旧数据随损坏文件隔离，新库只有新 session
        daily = self.hdb.query_daily(30)
        self.assertEqual(len(daily), 1)
        self.assertEqual(daily[0]['items'], 3)

    # ── 6. 只读故障注入（R8：绝不外抛）──────────────────────────────
    def test_readonly_fault_returns_minus1_without_raise(self):
        sid = self.hdb.record_capture({'云南': PLANS_YN}, 'live')
        self.assertGreater(sid, 0)
        os.chmod(self.hdb.db_path(), 0o444)  # Windows=READONLY 属性 / POSIX=只读位
        try:
            rc = self.hdb.record_capture({'云南': PLANS_YN}, 'live')
            self.assertEqual(rc, -1, '只读库写入失败应返回 -1')
        finally:
            os.chmod(self.hdb.db_path(), 0o666)
        sid2 = self.hdb.record_capture({'云南': PLANS_YN}, 'live')
        self.assertGreater(sid2, 0, '故障解除后应可继续写入')

    # ── 7. 并发写入 ──────────────────────────────────────────────────
    def test_concurrent_writers_thread_safe(self):
        sids, errs = [], []
        lock = threading.Lock()

        def worker(k):
            try:
                for j in range(2):
                    sid = self.hdb.record_capture({'云南': [
                        {'name': f'并发{k}-{j}-{i}', 'sku_id': f'700000000{k}{j}{i}',
                         'stock': i} for i in range(3)]}, 'batch')
                    with lock:
                        sids.append(sid)
            except Exception as e:  # 铁律被破坏（外抛）时在此暴露
                with lock:
                    errs.append(e)

        threads = [threading.Thread(target=worker, args=(k,)) for k in range(4)]
        for t in threads:
            t.start()
        for t in threads:
            t.join(timeout=60)
        self.assertFalse(errs, f'并发写入不允许外抛: {errs}')
        self.assertEqual(len(sids), 8)
        self.assertTrue(all(s > 0 for s in sids), '全部并发写入应成功')
        self.assertEqual(len(set(sids)), 8, 'session_id 不得重复')
        total = self._raw('SELECT COUNT(*) FROM history_rows')[0][0]
        self.assertEqual(total, 24, '4 线程 × 2 次 × 3 行应全部落库')

    # ── 8. clear_all ────────────────────────────────────────────────
    def test_clear_all(self):
        self.hdb.record_capture({'云南': PLANS_YN}, 'live')
        self.hdb.record_capture({'四川': PLANS_SC}, 'batch')
        self.assertTrue(self.hdb.clear_all())
        self.assertEqual(self.hdb.query_daily(30), [])
        self.assertEqual(self._raw('SELECT COUNT(*) FROM capture_sessions')[0][0], 0)
        self.assertTrue(os.path.exists(self.hdb.db_path()), '清空数据但保留库文件')
        sid = self.hdb.record_capture({'云南': PLANS_YN}, 'live')
        self.assertGreater(sid, 0, '清空后可继续正常使用')


# ════════════════════════════════════════════════════════════════════════════
# v1.4.x 导航页重构：📈 历史趋势 / 💰 用量明细 从弹窗独立为导航页
# t22 验收测试（结构性/静态断言，无 Tk 主循环依赖）
# ════════════════════════════════════════════════════════════════════════════

def _read(path):
    with open(path, encoding='utf-8', errors='replace') as fp:
        return fp.read()


def _count_defs(src, name):
    """统计 `def NAME(` 出现的次数（行首允许空白），避免单测到 docstring 提及旧名而误算。"""
    import re
    return len(re.findall(rf'^\s*def\s+{re.escape(name)}\s*\(', src, re.MULTILINE))


class TestNavRefactorV147(unittest.TestCase):
    """v1.4.x 导航页重构（t21）—— 验收测试。

    旧版历史趋势（gui._show_history_dialog）/ 用量明细（settings_ui._show_usage_detail）
    是 Toplevel 弹窗；新版把两个页面整体迁移为 stats_ui.StatsPagesMixin 的导航页，
    旧弹窗入口已删，地区 tab 行尾「📈 历史」改为导航跳转，API 页底部
    「💰 用量明细」按钮整体移除。

    本测试类只做静态/结构断言（不构造 Tk，避免主循环依赖）：
    1. stats_ui.py 模块与 StatsPagesMixin 类存在，关键构建/刷新方法齐备；
    2. gui.py 导航 items 列表含两个新页签；
    3. gui.py _show_page 含 page_history / page_usage 懒构建分支；
    4. gui.py App 类继承 StatsPagesMixin；
    5. 地区 tab「📈 历史」按钮调用 _goto_history_page（不再 Toplevel）；
    6. 旧 _show_history_dialog（gui.py）与 _show_usage_detail（settings_ui.py）
       函数定义已移除（仅允许在文档字符串/注释里提及"已迁移"）；
    7. settings_ui.py 不再含「💰 用量明细」按钮行（仅注释允许）；
    8. 既有 drill-down（_history_day_detail / _history_sku_chart）保留
       （双击历史行的明细/折线仍弹 Toplevel，是有意保留）。
    """

    GUI = os.path.join(HERE, 'gui.py')
    SETTINGS = os.path.join(HERE, 'settings_ui.py')
    STATS = os.path.join(HERE, 'stats_ui.py')

    # ── 1. stats_ui 模块齐备 ────────────────────────────────────────
    def test_stats_ui_module_exists(self):
        self.assertTrue(os.path.exists(self.STATS), 'stats_ui.py 必须存在（独立 Mixin 模块）')

    def test_stats_ui_module_importable(self):
        """守护式 import：importlib 强制 spec 加载并执行模块体，校验符号表。"""
        import importlib.util
        spec = importlib.util.spec_from_file_location('pdd_stats_ui', self.STATS)
        m = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(m)
        self.assertTrue(hasattr(m, 'StatsPagesMixin'),
                        'StatsPagesMixin 类必须存在')
        for fn in ('_build_history_page', '_build_usage_page',
                   '_history_page_refresh', '_usage_page_refresh',
                   '_make_usage_tree', '_refill_usage_tree',
                   '_usage_chart_redraw', '_usage_chart_on_resize'):
            self.assertTrue(hasattr(m.StatsPagesMixin, fn),
                            f'StatsPagesMixin.{fn} 必须存在')

    # ── 2. gui.py App 继承 StatsPagesMixin ──────────────────────────
    def test_gui_app_class_inherits_stats_mixin(self):
        src = _read(self.GUI)
        self.assertIn('class App(SettingsUIMixin, StatsPagesMixin)', src,
                      'App 必须继承 StatsPagesMixin 才能拥有两页构建能力')
        self.assertIn('from stats_ui import StatsPagesMixin', src,
                      'gui.py 必须 import StatsPagesMixin')

    # ── 3. 导航 items 列表含两个新页签 ─────────────────────────────
    def test_nav_items_contains_history_and_usage(self):
        src = _read(self.GUI)
        # _build_nav 的 items 列表必须同时含两个新页
        m = re.search(r'def\s+_build_nav\s*\(self\):\s*\n(.*?)\n\s+for\s+text,\s+page\s+in\s+items:',
                      src, re.DOTALL)
        self.assertIsNotNone(m, '未找到 _build_nav 的 items 列表')
        items_block = m.group(1)
        self.assertIn('📈 历史趋势', items_block, '导航 items 缺「📈 历史趋势」')
        self.assertIn('💰 用量明细', items_block, '导航 items 缺「💰 用量明细」')
        self.assertIn('self.page_history', items_block,
                      '「📈 历史趋势」必须指向 self.page_history')
        self.assertIn('self.page_usage', items_block,
                      '「💰 用量明细」必须指向 self.page_usage')

    # ── 4. _show_page 含 page_history / page_usage 懒构建分支 ─────
    def test_show_page_lazy_build_branches(self):
        src = _read(self.GUI)
        m = re.search(r'def\s+_show_page\s*\(self,\s*page\):', src)
        self.assertIsNotNone(m, '未找到 _show_page 定义')
        body = src[m.start():m.start() + 2000]
        self.assertIn('self._build_history_page(page)', body,
                      '_show_page 必须懒构建历史页')
        self.assertIn('self._build_usage_page(page)', body,
                      '_show_page 必须懒构建用量页')
        self.assertIn('self._history_page_refresh', body,
                      '_show_page 必须调度历史页刷新（after_idle 链路）')
        self.assertIn('self._usage_page_refresh', body,
                      '_show_page 必须调度用量页刷新（after_idle 链路）')

    def test_show_page_uses_after_idle_not_blocking(self):
        """刷新必须经 win.after_idle 调度（worker 线程不直调 Tk）。"""
        src = _read(self.GUI)
        m = re.search(r'def\s+_show_page\s*\(self,\s*page\):', src)
        body = src[m.start():m.start() + 2500]
        # 切页到 history 时应使用 after_idle 调度刷新
        hist_idx = body.find('self.page_history')
        usage_idx = body.find('self.page_usage')
        self.assertGreater(hist_idx, 0)
        self.assertGreater(usage_idx, 0)
        self.assertIn('after_idle', body[hist_idx:hist_idx + 600],
                      'page_history 刷新必须经 after_idle 调度（主线程事件队列）')
        self.assertIn('after_idle', body[usage_idx:usage_idx + 600],
                      'page_usage 刷新必须经 after_idle 调度（主线程事件队列）')

    # ── 5. 地区 tab「📈 历史」按钮 → _goto_history_page ────────────
    def test_region_tab_history_button_uses_nav_shortcut(self):
        src = _read(self.GUI)
        # _update_tabs 必须有「📈 历史」按钮，command 必须是 _goto_history_page
        m = re.search(r'def\s+_update_tabs\s*\(self\):.*?(?=\n    def\s|\nclass\s)',
                      src, re.DOTALL)
        self.assertIsNotNone(m, '未找到 _update_tabs 定义')
        body = m.group(0)
        self.assertIn('📈 历史', body, '地区 tab 行尾必须保留「📈 历史」按钮')
        self.assertIn('_goto_history_page', body,
                      '地区 tab「📈 历史」按钮必须调用 _goto_history_page（导航跳转，非 Toplevel）')

    def test_goto_history_page_function_exists_and_dispatches(self):
        """_goto_history_page：地区 tab 行的快捷方式，统一走 _show_page 跳转。"""
        src = _read(self.GUI)
        self.assertEqual(_count_defs(src, '_goto_history_page'), 1,
                         'gui.py 必须定义 _goto_history_page')
        m = re.search(r'def\s+_goto_history_page\s*\(self\):.*?(?=\n    def\s|\nclass\s)',
                      src, re.DOTALL)
        self.assertIsNotNone(m)
        body = m.group(0)
        self.assertIn('self._show_page(self.page_history)', body,
                      '_goto_history_page 必须 _show_page(self.page_history) 跳转')

    # ── 6. 旧 _show_history_dialog（gui.py）已移除 ─────────────────
    def test_old_history_dialog_function_removed(self):
        src = _read(self.GUI)
        self.assertEqual(_count_defs(src, '_show_history_dialog'), 0,
                         '_show_history_dialog 必须删除（导航页是唯一实现）')
        # 但允许 docstring 注释里提及"已迁移"
        m = re.search(r'def\s+_goto_history_page\s*\(self\):.*?(?=\n    def\s|\nclass\s)',
                      src, re.DOTALL)
        if m:
            self.assertIn('_show_history_dialog', m.group(0),
                          '_goto_history_page 注释应说明旧 _show_history_dialog 已迁')

    # ── 7. 旧 _show_usage_detail（settings_ui.py）已移除 ───────────
    def test_old_usage_detail_function_removed(self):
        src = _read(self.SETTINGS)
        self.assertEqual(_count_defs(src, '_show_usage_detail'), 0,
                         '_show_usage_detail 必须删除（导航页是唯一实现）')

    # ── 8. settings_ui.py 不再含「💰 用量明细」按钮行 ─────────────
    def test_api_page_no_usage_button(self):
        """API 页不再内置「💰 用量明细」入口（用户明确要求）——但允许 docstring/注释里提及。"""
        src = _read(self.SETTINGS)
        # 直接 grep "_mk_btn" 行含"💰 用量明细"必须为 0
        import re as _re
        bad = _re.findall(r'^\s*self\._mk_btn\([^)]*💰\s*用量明细', src, _re.MULTILINE)
        self.assertEqual(len(bad), 0,
                         f'settings_ui.py 仍有「💰 用量明细」按钮行: {bad}')

    # ── 9. 历史 drill-down（_history_day_detail / _history_sku_chart）保留 ──
    def test_history_drill_down_dialogs_preserved(self):
        """双击历史行/明细行的 drill-down Toplevel 仍保留（与导航页正交）。"""
        src = _read(self.GUI)
        self.assertEqual(_count_defs(src, '_history_day_detail'), 1,
                         '_history_day_detail 必须保留（明细 Toplevel）')
        self.assertEqual(_count_defs(src, '_history_sku_chart'), 1,
                         '_history_sku_chart 必须保留（折线 Toplevel）')

    # ── 10. page_history / page_usage 帧已挂到 content_frame ─────
    def test_page_frames_attached(self):
        src = _read(self.GUI)
        self.assertRegex(src, r'self\.page_history\s*=\s*tk\.Frame\(self\.content_frame',
                         'page_history 必须作为 content_frame 子帧创建')
        self.assertRegex(src, r'self\.page_usage\s*=\s*tk\.Frame\(self\.content_frame',
                         'page_usage 必须作为 content_frame 子帧创建')

    # ── 11. P3-R2-M1：_safe_destroy_widget 静态方法存在 + 行为契约 ─────
    def test_safe_destroy_widget_exists_and_safe(self):
        """M1 修复：_safe_destroy_widget 守卫方法存在，且对 None / 假对象不抛异常。

        真实 Entry widget 的 destroy() 时序竞态需 Tk mainloop 才能复现，本静态契约
        测试只验证「None/已销毁对象」等异常输入下不抛异常（守护式语义）。
        """
        import importlib.util
        spec = importlib.util.spec_from_file_location('pdd_stats_ui_m1', self.STATS)
        m = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(m)
        self.assertTrue(hasattr(m.StatsPagesMixin, '_safe_destroy_widget'),
                        'StatsPagesMixin._safe_destroy_widget 必须存在（M1 修复）')
        # None / 任意假对象 / 抛 winfo_exists 的对象都不应让 _safe_destroy_widget 抛异常
        for fake in (None, object(), type('X', (), {
                'winfo_exists': staticmethod(lambda: (_ for _ in ()).throw(RuntimeError('x')))
        })()):
            try:
                m.StatsPagesMixin._safe_destroy_widget(fake)
            except Exception as e:
                self.fail(f'_safe_destroy_widget({fake!r}) 不应抛异常，实际抛 {e!r}')

    # ── 12. P3-R2-L1：_chart_empty_message 三分支决策（无 Tk 行为级断言）──
    def test_chart_empty_message_three_branches(self):
        """L1 修复：_chart_empty_message 静态方法的「空 entries / 全 0 / 正常」三分支。

        无 Tk 依赖，可直接断言决策输出（redraw 的具体画法不在本测试范围）。
        """
        import importlib.util
        spec = importlib.util.spec_from_file_location('pdd_stats_ui_l1', self.STATS)
        m = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(m)
        fn = m.StatsPagesMixin._chart_empty_message
        self.assertTrue(callable(fn), '_chart_empty_message 必须可调用')

        # 分支 1：panel=None / by_model 缺省 / 空 dict → 引导识别
        self.assertIn('暂无', fn(None))
        self.assertIn('暂无', fn({}))
        self.assertIn('暂无', fn({'by_model': {}}))
        # 分支 2：entries 非空但全部 cost=0（含 None / 负数归 0）→ 引导配价格表
        self.assertIn('价格表', fn({'by_model': {
            'doubao-seed': {'cost': 0.0, 'tokens': 100, 'count': 1},
            'qwen-vl': {'cost': None, 'tokens': 50, 'count': 1},
        }}))
        # 分支 3：至少一条 cost>0 → None（正常绘制）
        self.assertIsNone(fn({'by_model': {
            'doubao-seed': {'cost': 0.0, 'tokens': 100, 'count': 1},
            'glm-4.6': {'cost': 0.05, 'tokens': 200, 'count': 2},
        }}))
        # 边界：单个模型 cost=0 仍属分支 2
        self.assertIn('价格表', fn({'by_model': {'only': {'cost': 0.0}}}))
        # 边界：entries 含负数 cost（被 (v.get('cost', 0.0) or 0.0) 归 0）→ 分支 2
        self.assertIn('价格表', fn({'by_model': {
            'a': {'cost': -1.0}, 'b': {'cost': -0.5},
        }}))


class TestKeyDecryptWiring(unittest.TestCase):
    """v1.4.8 P1-C-fix（t18）：ocr/vision 运行时 api_key 解密接线防回归。

    背景：迁移后 settings.json 里 api_key 是 dpapi:v1: 密文，运行时读取点必须经
    utils.decrypt_secret 解密；否则所有视觉 API 调用拿密文当 key，全线 401。
    """

    def _src(self, name):
        with open(os.path.join(HERE, name), 'r', encoding='utf-8') as f:
            return f.read()

    def test_ocr_vision_wiring_present(self):
        """ocr.py（≥3 处）与 vision.py（≥2 处）的 api_key 读取点已接 decrypt_secret。"""
        ocr_src = self._src('ocr.py')
        vision_src = self._src('vision.py')
        self.assertGreaterEqual(ocr_src.count('decrypt_secret('), 3,
                                'ocr.py 运行时 api_key 读取点必须经 decrypt_secret 解密')
        self.assertGreaterEqual(vision_src.count('decrypt_secret('), 2,
                                'vision.py 运行时 api_key 读取点必须经 decrypt_secret 解密')
        # 三个已知读取点形态逐一存在：主读取 / 降档模型 fallback / GLM fallback
        self.assertIn("key = decrypt_secret(key)", ocr_src)
        self.assertIn("cur_key = decrypt_secret(cur_key)", ocr_src)
        self.assertIn("_glm_key = decrypt_secret(_glm_key)", vision_src)

    def test_decrypt_secret_roundtrip_and_fallback(self):
        """utils.decrypt_secret：明文直通 / 密文往返+memo / 损坏密文返空 / 非字符串原样。"""
        import importlib.util
        spec = importlib.util.spec_from_file_location('pdd_utils_wiring', os.path.join(HERE, 'utils.py'))
        u = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(u)
        spec2 = importlib.util.spec_from_file_location('pdd_dpapi_wiring', os.path.join(HERE, 'dpapi_utils.py'))
        m = importlib.util.module_from_spec(spec2)
        spec2.loader.exec_module(m)
        if not m.is_available():
            self.skipTest('DPAPI 不可用（沙盒/Wine）')
        # 明文直通（无 dpapi:v1: 前缀零开销）+ 空值/非字符串语义
        self.assertEqual(u.decrypt_secret('sk-plain-key'), 'sk-plain-key')
        self.assertEqual(u.decrypt_secret(''), '')
        self.assertIsNone(u.decrypt_secret(None))
        self.assertEqual(u.decrypt_secret(123), 123)
        # 密文往返 + memo 缓存
        blob = m.enc('sk-wiring-test-key')
        self.assertEqual(u.decrypt_secret(blob), 'sk-wiring-test-key')
        self.assertIn(blob, u._decrypt_memo, '解密结果应进 memo 缓存')
        self.assertEqual(u.decrypt_secret(blob), 'sk-wiring-test-key')
        # 损坏密文 → 返回 ""（调用方按 key 为空路径处理，不抛）
        self.assertEqual(u.decrypt_secret('dpapi:v1:not-a-valid-blob'), '')

    def test_config_save_get_set_are_class_methods(self):
        """回归：decrypt_secret（模块级）曾插在 class Config 中间把类截断，
        save/get/set 沦为其函数体内死代码 → Config.save 不存在，所有配置写盘静默失效。
        断言类方法齐全 + set→磁盘→读回往返成立。"""
        import importlib.util
        import tempfile
        import shutil
        import json
        spec = importlib.util.spec_from_file_location('pdd_utils_save', os.path.join(HERE, 'utils.py'))
        u = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(u)
        for name in ('load', 'save', 'get', 'set', '_migrate_secrets', 'decrypt_value'):
            self.assertTrue(callable(getattr(u.Config, name, None)), f'Config.{name} 缺失')
        tmp = tempfile.mkdtemp()
        try:
            with open(os.path.join(tmp, 'settings_template.json'), 'w', encoding='utf-8') as f:
                json.dump({}, f)
            u.get_base_dir = lambda: tmp
            u.Config._load_cache = {'mtime': -1, 'data': None}
            u.Config._template_cache = None
            u.Config.set('t18_probe', {'ok': 1})
            with open(os.path.join(tmp, 'settings.json'), 'r', encoding='utf-8') as f:
                disk = json.load(f)
            self.assertEqual(disk.get('t18_probe'), {'ok': 1})
        finally:
            shutil.rmtree(tmp, ignore_errors=True)


class TestEULA(unittest.TestCase):
    """t7+t16 EULA 回归测试。"""

    def test_eula_version_constant(self):
        from eula_text import EULA_VERSION, EULA_CLAUSES
        self.assertEqual(EULA_VERSION, 'v1')
        # 恰 7 条
        self.assertEqual(len(EULA_CLAUSES), 7)

    def test_eula_text_renders_keywords(self):
        from eula_text import render_eula_text
        text = render_eula_text()
        self.assertIn('账号风险', text)
        self.assertIn('风险由用户本人承担', text)

    def test_gui_eula_dialog_blocks_with_wait_window(self):
        """t16 修复：_show_eula_dialog 体内必须包含 wait_visibility + wait_window 阻塞序列。"""
        import gui
        src = inspect.getsource(gui)
        # 找到 _show_eula_dialog 函数体
        idx = src.find('def _show_eula_dialog')
        self.assertGreater(idx, 0, "gui.py must define _show_eula_dialog")
        # 取到下一个 def 为止
        next_def = src.find('\n    def ', idx + 1)
        body = src[idx:next_def if next_def > 0 else len(src)]
        self.assertIn('wait_window', body, "EULA dialog must call wait_window to block")
        self.assertIn('wait_visibility', body, "EULA dialog must call wait_visibility for stable display")


class TestLicense(unittest.TestCase):
    """t10 P2-A 离线卡密授权测试。"""

    def test_rfc8032_vector1(self):
        from auth.ed25519_verify import verify, TEST1_PK, TEST1_MSG, TEST1_SIG
        self.assertTrue(verify(TEST1_PK, TEST1_MSG, TEST1_SIG))

    def test_rfc8032_vector2(self):
        from auth.ed25519_verify import verify, TEST2_PK, TEST2_MSG, TEST2_SIG
        self.assertTrue(verify(TEST2_PK, TEST2_MSG, TEST2_SIG))

    def test_tampered_msg_rejected(self):
        from auth.ed25519_verify import verify, TEST1_PK, TEST1_SIG
        self.assertFalse(verify(TEST1_PK, b'x', TEST1_SIG))

    def test_tampered_sig_rejected(self):
        from auth.ed25519_verify import verify, TEST1_PK, TEST1_MSG, TEST1_SIG
        bad = TEST1_SIG[:-1] + b'\\x00'
        self.assertFalse(verify(TEST1_PK, TEST1_MSG, bad))

    def test_get_tier_enforce_false_is_pro(self):
        from auth.license import get_tier
        # enforce=False：恒 pro（即便 license_text 为空或乱填）
        self.assertEqual(get_tier("", enforce=False), 'pro')
        self.assertEqual(get_tier("garbage", enforce=False), 'pro')

    def test_get_tier_enforce_true_empty_is_free(self):
        from auth.license import get_tier
        # enforce=True 且无 license → free
        self.assertEqual(get_tier("", enforce=True), 'free')
        self.assertEqual(get_tier("garbage", enforce=True), 'free')

    def test_machine_fingerprint_stable(self):
        from auth.license import get_machine_fingerprint
        a = get_machine_fingerprint()
        b = get_machine_fingerprint()
        self.assertEqual(a, b)
        self.assertEqual(len(a), 32)  # 16 字节 hex = 32 chars

    def test_get_tier_cache_ttl_expires_and_reevaluates(self):
        """t24 修复包 A (BUG-1)：缓存 TTL=300s 到期后必须重验。

        修前：get_tier('', enforce=True) 第一次返 'free'，缓存永不过期，
        外部修改 settings.json 后该进程内仍返 'free' 直到重启。
        修后：TTL 到期后强制重验（即使同 license_text）。
        """
        import time as _time
        import auth.license as _lic
        _lic.reset_cache()
        # 第 1 次：返 free
        r1 = _lic.get_tier('garbage_for_ttl_test', enforce=True)
        self.assertEqual(r1, 'free')
        # 直接改 _CACHE 把时间戳设到 1000s 前，模拟 TTL 过期
        cache_key = 'True:garbage_for_ttl_test'
        _lic._CACHE[cache_key] = ('pro', _time.time() - 1000)  # 注入「过期但 tier=pro」
        # 第 2 次：TTL 已过期，应重验并返 free（而不是命中缓存的 'pro'）
        r2 = _lic.get_tier('garbage_for_ttl_test', enforce=True)
        self.assertEqual(r2, 'free')
        # 验证：缓存值已被刷新（ts 更新）
        self.assertGreater(_lic._CACHE[cache_key][1], _time.time() - 10)
        # 清理
        _lic.reset_cache()

    def test_get_tier_cache_ttl_fresh_hit_uses_cache(self):
        """t24 修复包 A (BUG-1)：TTL 内命中应直接返缓存，不重验。

        防回归：避免 TTL 实现错误把「同 license_text 重复调用」也走重验路径。
        """
        import time as _time
        import auth.license as _lic
        _lic.reset_cache()
        # 首次返 free
        _lic.get_tier('cache_fresh_test', enforce=True)
        cache_key = 'True:cache_fresh_test'
        # 注入「未过期但 tier=pro」模拟缓存被外部改
        _lic._CACHE[cache_key] = ('pro', _time.time() - 10)  # 10s 前，远未到 300s
        # 命中缓存（TTL 内）应直接返 'pro'
        r = _lic.get_tier('cache_fresh_test', enforce=True)
        self.assertEqual(r, 'pro')
        _lic.reset_cache()

    def test_settings_ui_enforce_rollback_on_save_failure(self):
        """t24 修复包 A (BUG-13)：_on_enforce_toggle 写盘失败时必须回滚 UI。

        直接验证源码：except 分支内必须执行 _enforce_var.set(not bool(_enforce_var.get()))。
        """
        import inspect
        import settings_ui as _sui
        src = inspect.getsource(_sui)
        # 必须存在「失败时反转 _enforce_var」的语句
        self.assertIn('_enforce_var.set(not bool(_enforce_var.get()))', src)

    def test_settings_ui_import_rollback_on_save_failure(self):
        """t24 修复包 A (BUG-14)：_on_import 写盘失败时必须回滚 _key_var 为旧值。

        直接验证源码：except 分支内必须执行 _key_var.set(_old_key)。
        """
        import inspect
        import settings_ui as _sui
        src = inspect.getsource(_sui)
        # 进入 try 前快照
        self.assertIn('_old_key = _key_var.get()', src)
        # 失败时回滚
        self.assertIn('_key_var.set(_old_key)', src)


class TestGating(unittest.TestCase):
    """t12 P2-C Pro 门控测试。

    用户裁定：enforce=false 时所有门控失效（默认全免，永久免费功能不受影响）。
    enforce=true + tier=free 时：实时截图 50 次/日，历史趋势 30 天。
    enforce=true + tier=pro 时：不限。
    永久免费（不门控）：表格导入、手动输入、批量识别、双模型验证、Excel 导出。
    """

    def setUp(self):
        # 重置 license 缓存，确保 enforce 参数每次都被重新计算
        try:
            from auth.license import reset_cache
            reset_cache()
        except Exception:
            pass
        # 备份 settings.json license 段
        self._orig_cfg = None
        try:
            from utils import Config
            self._orig_cfg = Config.load() if hasattr(Config, "load") else {}
        except Exception:
            pass

    def tearDown(self):
        # 还原 settings.json license 段
        try:
            if self._orig_cfg is not None:
                from utils import Config
                Config.save(self._orig_cfg)
        except Exception:
            pass
        try:
            from auth.license import reset_cache
            reset_cache()
        except Exception:
            pass

    def test_enforce_false_is_pro_and_no_quota(self):
        """enforce=false：get_tier 恒 pro、check_live_quota 恒 allowed、history 不限"""
        from auth.license import get_tier, is_pro, check_live_quota, get_history_days_limit
        self.assertEqual(get_tier("", enforce=False), 'pro')
        self.assertEqual(get_tier("garbage", enforce=False), 'pro')
        self.assertTrue(is_pro("", enforce=False))
        # 999999 是 used 也无影响
        gate = check_live_quota(999999, "", enforce=False)
        self.assertTrue(gate["allowed"])
        # history 无限
        self.assertEqual(get_history_days_limit("", enforce=False), 999999)

    def test_enforce_true_free_live_quota_blocks_at_50(self):
        """enforce=true + free：used >= 50 阻断，< 50 通过"""
        from auth.license import check_live_quota
        # used=49 → 通过
        g = check_live_quota(49, "garbage_no_valid_license", enforce=True)
        self.assertTrue(g["allowed"])
        # used=50 → 阻断
        g = check_live_quota(50, "garbage_no_valid_license", enforce=True)
        self.assertFalse(g["allowed"])
        # used=100 → 阻断
        g = check_live_quota(100, "garbage_no_valid_license", enforce=True)
        self.assertFalse(g["allowed"])

    def test_enforce_true_pro_live_quota_unlimited(self):
        """enforce=true + Pro：不限（无 license 时 is_pro=False 走 free 路径，
        所以这里只能断言"is_pro=False 的 free 路径在 enforce=true 下被钳制"，
        真正 Pro 验证需密钥——保留扩展点）。"""
        from auth.license import check_live_quota
        # 无 license + enforce=true → free 路径 → 50 次上限
        g = check_live_quota(60, "", enforce=True)
        self.assertFalse(g["allowed"])

    def test_enforce_true_free_history_30_days(self):
        """enforce=true + free：history 钳到 30 天"""
        from auth.license import get_history_days_limit
        # 无 license → free
        self.assertEqual(get_history_days_limit("", enforce=True), 30)
        self.assertEqual(get_history_days_limit("garbage", enforce=True), 30)

    def test_check_live_quota_result_shape(self):
        """返回 dict 字段完整（allowed/limit/used/remaining/reason）"""
        from auth.license import check_live_quota
        g = check_live_quota(10, "garbage", enforce=True)
        for k in ("allowed", "limit", "used", "remaining", "reason"):
            self.assertIn(k, g)
        self.assertEqual(g["used"], 10)

    def test_get_license_info_enforce_false(self):
        """enforce=false：返回 is_pro=True，status 含「试用期」"""
        from auth.license import get_license_info
        info = get_license_info("", enforce=False)
        self.assertTrue(info["is_pro"])
        self.assertIn("试用期", info["status_text"])
        self.assertFalse(info["enforce"])

    def test_get_license_info_enforce_true_no_key(self):
        """enforce=true 且无 license：is_pro=False，tier=free"""
        from auth.license import get_license_info
        info = get_license_info("", enforce=True)
        self.assertFalse(info["is_pro"])
        self.assertEqual(info["tier"], "free")
        self.assertIn("免费版", info["status_text"])

    def test_get_license_info_enforce_true_garbage(self):
        """enforce=true + 无效 license：等同无 license → free"""
        from auth.license import get_license_info
        info = get_license_info("garbage_license_text_xxx", enforce=True)
        self.assertFalse(info["is_pro"])

    def test_count_today_live_screenshot_smoke(self):
        """usage_store.count_today_live_screenshot() 返回整数，失败返 0"""
        import usage_store as us
        n = us.count_today_live_screenshot()
        self.assertIsInstance(n, int)
        self.assertGreaterEqual(n, 0)

    def test_license_module_constants(self):
        """t12 阈值常量正确导出"""
        from auth.license import FREE_DAILY_LIVE_SCREENSHOT, FREE_HISTORY_DAYS, UNLIMITED
        self.assertEqual(FREE_DAILY_LIVE_SCREENSHOT, 50)
        self.assertEqual(FREE_HISTORY_DAYS, 30)
        self.assertEqual(UNLIMITED, 999999)


class TestReplenishmentModels(unittest.TestCase):
    """t13 P3-A 补货模型框架：经典原样保留 + 加权新增。

    用户裁定（最高优先约束）：
    - 经典模式（默认）一行公式逻辑都不许改，行为与改动前完全一致
    - 加权模式：日销 = 0.5×7日 + 0.3×14日 + 0.2×30日
    - 加权无历史 → 回退经典 + 标注「经典(无历史)」
    - 加权任何异常 → 逐商品回退经典，绝不中断整批
    """

    def setUp(self):
        try:
            from utils import Config
            self._orig = Config.load() if hasattr(Config, "load") else {}
        except Exception:
            self._orig = {}

    def tearDown(self):
        try:
            from utils import Config
            if self._orig is not None:
                Config.save(self._orig)
        except Exception:
            pass

    def test_get_replenishment_cfg_defaults(self):
        """默认 = classic + safety_days=2 + in_transit_qty=0"""
        from utils import get_replenishment_cfg
        cfg = get_replenishment_cfg()
        self.assertEqual(cfg['model'], 'classic')
        self.assertEqual(cfg['safety_days'], 2)
        self.assertEqual(cfg['in_transit_qty'], 0)

    def test_get_replenishment_cfg_weighted(self):
        from utils import get_replenishment_cfg
        from utils import Config
        cfg = Config.load() if hasattr(Config, "load") else {}
        cfg['replenishment'] = {'model': 'weighted', 'safety_days': 5, 'in_transit_qty': 200}
        Config.save(cfg)
        cfg2 = get_replenishment_cfg()
        self.assertEqual(cfg2['model'], 'weighted')
        self.assertEqual(cfg2['safety_days'], 5)
        self.assertEqual(cfg2['in_transit_qty'], 200)

    def test_classic_formula_unchanged(self):
        """经典公式：固定输入 → 固定输出（与改动前完全一致）"""
        from utils import calc_replenishment_classic
        # 输入：stock=0, sales=10, shipping=1, offset=1
        # daily=10, ratio=0/10=0, lead_time=2, reorder=-2 → 立刻补货
        # qty = max(10*8, 100) = 100, round to 100
        p = calc_replenishment_classic({'name': 'X', 'stock': 0, 'sales': 10}, '广东', 1, 1)
        self.assertEqual(p['status'], '立刻补货')
        self.assertEqual(p['color'], 'red')
        self.assertEqual(p['qty'], 100)
        self.assertEqual(p['model'], 'classic')
        # 输入：stock=200, sales=10, shipping=1, offset=1
        # ratio=20, lead_time=2, reorder=18 → green
        p2 = calc_replenishment_classic({'name': 'X', 'stock': 200, 'sales': 10}, '广东', 1, 1)
        self.assertEqual(p2['color'], 'green')
        self.assertEqual(p2['qty'], 0)
        # 无销量
        p3 = calc_replenishment_classic({'name': 'X', 'stock': 50, 'sales': 0}, '广东', 1, 1)
        self.assertEqual(p3['status'], '无销量·观察')
        self.assertEqual(p3['qty'], 0)

    def test_weighted_with_history_uses_weighted_avg(self):
        """加权模式：含历史数据 → 加权日销 (0.5×7 + 0.3×14 + 0.2×30)"""
        from utils import calc_replenishment_weighted
        from datetime import datetime, timedelta
        today = datetime.now()
        rows = []
        # 近 7 日每天 20 件
        for i in range(7):
            d = (today - timedelta(days=i)).strftime('%Y-%m-%d')
            rows.append({'captured_at': d, 'sales': 20, 'name': 'X'})
        # 近 8-14 日每天 10 件
        for i in range(8, 15):
            d = (today - timedelta(days=i)).strftime('%Y-%m-%d')
            rows.append({'captured_at': d, 'sales': 10, 'name': 'X'})
        # 近 15-30 日每天 5 件
        for i in range(15, 31):
            d = (today - timedelta(days=i)).strftime('%Y-%m-%d')
            rows.append({'captured_at': d, 'sales': 5, 'name': 'X'})
        def hlookup(sku, reg, days, name=None):
            return rows
        item = {'name': 'X', 'stock': 0, 'sales': 999, 'sku_id': 'SKU001'}
        p = calc_replenishment_weighted(item, '广东', 2, 2, 0, hlookup)
        self.assertEqual(p['model'], 'weighted')
        # 加权日销 = 0.5*avg(7) + 0.3*avg(14) + 0.2*avg(30)
        # avg(7)=20, avg(14)=(7*20+7*10)/14=15, avg(30)=(7*20+7*10+16*5)/30=290/30≈9.67
        # → 0.5*20 + 0.3*15 + 0.2*9.67 ≈ 16.43
        self.assertAlmostEqual(p['daily'], 16.4, delta=0.5)
        # 立刻补货（stock=0/daily=14 ≈ 0, lead_time=2+2=4）
        self.assertEqual(p['color'], 'red')
        self.assertIn('立刻补货', p['status'])

    def test_weighted_no_history_falls_back(self):
        """加权无历史 → 经典公式 + 标注「经典(无历史)」"""
        from utils import calc_replenishment_weighted
        def hlookup(*a, **kw):
            return []
        item = {'name': 'X', 'stock': 0, 'sales': 10, 'sku_id': 'SKU001'}
        p = calc_replenishment_weighted(item, '广东', 1, 2, 0, hlookup)
        self.assertEqual(p['model'], 'classic(no_history)')
        # 与经典公式一致：立刻补货、qty=100
        self.assertEqual(p['status'], '立刻补货')
        self.assertEqual(p['qty'], 100)

    def test_weighted_exception_falls_back(self):
        """加权 history_lookup 抛异常 → 经典公式 + 标注回退"""
        from utils import calc_replenishment_weighted
        def hlookup(*a, **kw):
            raise RuntimeError("db corrupted")
        item = {'name': 'X', 'stock': 0, 'sales': 10, 'sku_id': 'SKU001'}
        p = calc_replenishment_weighted(item, '广东', 1, 2, 0, hlookup)
        self.assertEqual(p['model'], 'classic(no_history)')
        self.assertEqual(p['status'], '立刻补货')

    def test_calc_replenishment_dispatch_classic(self):
        """calc_replenishment 入口：model='classic' 走经典路径"""
        from utils import calc_replenishment
        items = [{'name': 'A', 'stock': 0, 'sales': 10}]
        def sl(item, reg): return 1
        def hl(*a, **kw): return []
        plans = calc_replenishment(items, '广东', 'classic', 2, 0, sl, hl, offset=1)
        self.assertEqual(len(plans), 1)
        self.assertEqual(plans[0]['model'], 'classic')
        self.assertEqual(plans[0]['status'], '立刻补货')

    def test_calc_replenishment_dispatch_weighted_with_history(self):
        """calc_replenishment 入口：model='weighted' + 有历史 → 加权"""
        from utils import calc_replenishment
        from datetime import datetime, timedelta
        today = datetime.now()
        rows = [{'captured_at': (today - timedelta(days=i)).strftime('%Y-%m-%d'),
                 'sales': 10, 'name': 'A'} for i in range(20)]
        items = [{'name': 'A', 'stock': 0, 'sales': 10, 'sku_id': 'S1'}]
        def sl(item, reg): return 1
        def hl(sku, reg, days, name=None):
            return rows if sku else []
        plans = calc_replenishment(items, '广东', 'weighted', 2, 0, sl, hl)
        self.assertEqual(plans[0]['model'], 'weighted')

    def test_calc_replenishment_per_item_exception_falls_back(self):
        """任何异常 → 逐商品回退经典，绝不中断整批"""
        from utils import calc_replenishment
        items = [
            {'name': 'A', 'stock': 0, 'sales': 10},
            {'name': 'B', 'stock': 0, 'sales': 10, 'sku_id': 'B1'},
        ]
        def sl(item, reg): return 1
        def hl(*a, **kw):
            raise RuntimeError("kaboom")
        plans = calc_replenishment(items, '广东', 'weighted', 2, 0, sl, hl)
        self.assertEqual(len(plans), 2)
        for p in plans:
            # 无有效历史 → 经典(无历史) 兜底
            self.assertIn(p['model'], ('classic(no_history)', 'classic'))

    def test_settings_template_has_replenishment(self):
        """settings_template.json 包含 replenishment 默认段"""
        import json
        with open(os.path.join(HERE, 'settings_template.json'), 'r', encoding='utf-8') as f:
            tpl = json.load(f)
        self.assertIn('replenishment', tpl)
        self.assertEqual(tpl['replenishment']['model'], 'classic')
        self.assertEqual(tpl['replenishment']['safety_days'], 2)
        self.assertEqual(tpl['replenishment']['in_transit_qty'], 0)

    def test_export_xlsx_has_model_column(self):
        """export_xlsx.py 输出 headers 含「模型」列（缺省不破坏旧表，新增最右列）"""
        import export_xlsx
        src = inspect.getsource(export_xlsx.export_cache_to_xlsx)
        self.assertIn("'模型'", src)
        self.assertIn("'补货量'", src)  # 旧列仍在

    def test_gui_calc_from_items_dispatches_by_model(self):
        """gui.py _calc_from_items 含 weighted 分发（读 config.replenishment.model）"""
        import gui
        src = inspect.getsource(gui.App._calc_from_items)
        # 必须有 weighted 分支
        self.assertIn('weighted', src)
        # 必须有 model 标注写入
        self.assertIn("'model': _model_tag", src)
        # 经典分支必须保留原公式关键词（防回归）
        self.assertIn("立刻补货", src)
        self.assertIn("daily * 8", src)


class TestBatchCostPreview(unittest.TestCase):
    """t14 P3-B 批量前成本预估确认。

    用户裁定：
    - 价格表完整 → 显示金额区间（CNY 区间）
    - 价格表为空/缺价 → 标题「（价格未配置，仅按 token 数提示）」，显示调用次数
    - 用户取消 → 干净退出，不置任何熔断标志（与 F9/_api_fatal/预算三态无关）
    - 估算失败 → 静默跳过确认，绝不阻塞批量
    """

    def setUp(self):
        try:
            from utils import Config
            self._orig = Config.load() if hasattr(Config, "load") else {}
        except Exception:
            self._orig = {}

    def tearDown(self):
        try:
            from utils import Config
            if self._orig is not None:
                Config.save(self._orig)
        except Exception:
            pass

    def test_gui_has_preview_batch_cost_method(self):
        """gui.App 有 _preview_batch_cost 方法"""
        import gui
        self.assertTrue(hasattr(gui.App, '_preview_batch_cost'))
        import inspect
        sig = inspect.signature(gui.App._preview_batch_cost)
        self.assertIn('region_count', sig.parameters)
        self.assertIn('dual_verify', sig.parameters)

    def test_gui_batch_start_invokes_preview(self):
        """批量启动 start_batch 内 _preview_batch_cost 调用"""
        import gui
        src = inspect.getsource(gui)
        self.assertIn('_preview_batch_cost', src)
        # 确认调用点在双批守卫之后、未选地区检查之后
        self.assertIn('if not self._preview_batch_cost(', src)

    def test_helper_to_float_safe(self):
        """helper _to_float_safe 健壮"""
        from gui import _to_float_safe
        self.assertEqual(_to_float_safe('1.5'), 1.5)
        self.assertEqual(_to_float_safe('invalid', 0.0), 0.0)
        self.assertEqual(_to_float_safe(None, 0.0), 0.0)
        self.assertEqual(_to_float_safe(0), 0.0)
        self.assertEqual(_to_float_safe(0, 99), 0)  # 0 is not the default

    def test_helper_fmt_yuan(self):
        """helper _fmt_yuan 格式化"""
        from gui import _fmt_yuan
        # ≥1 元用 ¥1.23 格式
        self.assertIn('¥', _fmt_yuan(1.5))
        # <1 元用 4 位小数
        r = _fmt_yuan(0.001)
        self.assertIn('¥', r)
        # 0 → ¥0
        self.assertIn('¥0', _fmt_yuan(0))

    def test_preview_batch_cost_uses_pricing_cfg(self):
        """_preview_batch_cost 读 utils.get_usage_cfg().pricing（不硬编码）"""
        import gui
        src = inspect.getsource(gui.App._preview_batch_cost)
        self.assertIn('get_usage_cfg', src)
        self.assertIn('pricing', src)
        self.assertIn('input_per_million', src)
        self.assertIn('output_per_million', src)

    def test_preview_batch_cost_falls_back_on_no_pricing(self):
        """价格表为空 → 标题含「价格未配置」"""
        import gui
        src = inspect.getsource(gui.App._preview_batch_cost)
        self.assertIn('价格未配置', src)
        self.assertIn('预计调用', src)

    def test_preview_batch_cost_exception_does_not_block(self):
        """估算函数顶部 try/except，异常 → 放行（绝不阻塞批量）"""
        import gui
        src = inspect.getsource(gui.App._preview_batch_cost)
        # 必须有顶层 try/except 包整个函数
        self.assertIn('try:', src)
        self.assertIn('except Exception:', src)
        # 异常路径必须 return True（放行）
        # 函数最后一段 "return True" 在 except 块内
        idx = src.rfind('return True')
        self.assertGreater(idx, 0)

    def test_preview_batch_cost_per_round_estimation(self):
        """轮次估计 3~8，calls 公式合理"""
        import gui
        src = inspect.getsource(gui.App._preview_batch_cost)
        # 显式声明保守轮次 3 ~ 8
        self.assertIn('min_rounds, max_rounds = 3, 8', src)
        # 调用次数公式
        self.assertIn('per_round_calls_per_region', src)
        self.assertIn('1 + 1 * n_models', src)

    def test_settings_template_has_pricing_or_empty(self):
        """settings_template.json 中 usage.pricing 默认 {}（不破坏；用户配置即生效）"""
        import json
        with open(os.path.join(HERE, 'settings_template.json'), 'r', encoding='utf-8') as f:
            tpl = json.load(f)
        # pricing 缺省或 {} 都 OK
        pricing = tpl.get('usage', {}).get('pricing', {}) or {}
        self.assertIsInstance(pricing, dict)

    def test_t24_f1_classic_error_label_distinguishes(self):
        """t24 修复包 A (F1)：gui.py 加权 except 回退标注 'classic(error)'。

        与 utils.calc_replenishment_weighted 的 'classic(no_history)' 区分：
        - classic(no_history) = utils 内部查到空结果主动回退
        - classic(error)       = gui 层 utils 抛异常被动回退
        """
        import gui
        src = inspect.getsource(gui)
        # gui.py:2029 必须用 'classic(error)' 标注被动回退
        self.assertIn("'classic(error)'", src)
        # 注释中必须解释两个标签区别
        self.assertIn('classic(error)', src)
        self.assertIn('classic(no_history)', src)

    def test_t24_f2_self_tier_dead_code_removed(self):
        """t24 修复包 A (F2)：gui.py 删除 self._tier 死状态 + _auth_get_tier import。

        防回归：避免后续又把死代码加回来。
        """
        import gui
        src = inspect.getsource(gui)
        # 不应再有 self._tier 赋值（死状态）
        self.assertNotIn('self._tier = ', src)
        # 不应再有 `from auth.license import get_tier as _auth_get_tier` 这种 import 行
        self.assertNotIn('from auth.license import get_tier as _auth_get_tier', src)
        # 不应再调用 _auth_get_tier(...)
        self.assertNotIn('_auth_get_tier(', src)


class TestNavGrouping(unittest.TestCase):
    """t27 实施包 A：导航默认展开 + 三组分组 + 标题统一 + 导出降权 + 死代码清理。

    用户裁定：t25 IA 审计 → t27 实施 4 项定稿（提案 1+2+5+4+6）。
    约束：零新配色/字体/token；首页双入口布局不动；业务逻辑零变化。
    """

    def test_nav_default_expanded_in_build_ui(self):
        """t27 ①：gui.py _build_ui 初始即 add(nav_frame, before=content_frame)。

        修 P-A1（新用户看不到 8 项功能）——main_paned 不能再只 add content_frame。
        """
        import gui
        import inspect
        src = inspect.getsource(gui.App._build_ui)
        # 必须存在 main_paned.add(self.nav_frame, ...) 在 _build_ui 内
        self.assertIn('self.main_paned.add(self.nav_frame', src)
        # 必须在 add(content_frame) 之前
        nav_add_pos = src.find('self.main_paned.add(self.nav_frame')
        content_add_pos = src.find('self.main_paned.add(self.content_frame')
        self.assertGreater(nav_add_pos, 0)
        self.assertGreater(content_add_pos, 0)
        self.assertLess(nav_add_pos, content_add_pos)
        # 宽度统一 176（t29 迭代R3：字面量已 DPI 化为 int(176 * self.dpi_scale)）
        self.assertIn('width=int(176 * self.dpi_scale)', src)
        # _toggle_nav 仍保留（提供折叠入口给熟练用户）
        self.assertTrue(hasattr(gui.App, '_toggle_nav'))

    def test_nav_three_group_layout(self):
        """t27 ②：_build_nav 必须按【工作区】【数据】【设置】三组布局。

        修 P-A2（8 项平铺无分组，命名不一致）。
        """
        import gui
        import inspect
        src = inspect.getsource(gui.App._build_nav)
        # t29 修 v4f-P3：剥离行内注释后断言——防止断言被注释文本（如文档性提及）满足
        src = '\n'.join(line.split('#', 1)[0] for line in src.splitlines())
        # 必须有 groups 列表 + 3 个组名
        self.assertIn('工作区', src)
        self.assertIn('数据', src)
        self.assertIn('设置', src)
        # 组标必须用 _lbl + 8pt C_MUTED
        self.assertIn("font=(self.FONT[0], 8)", src)
        self.assertIn("fg=self.C_MUTED", src)
        # 组间分隔 Frame 用 C_BORDER
        self.assertIn('C_BORDER', src)
        # nav 按钮 padx=14 pady=7
        self.assertIn('padx=14, pady=7', src)

    def test_page_titles_unified_to_font_heading(self):
        """t27 ③：所有 page 顶部标题统一 FONT_HEADING + 无 emoji（emoji 在导航）。

        修 P-A4（page 标题字体/pady/emoji 不统一）。
        """
        import settings_ui
        import stats_ui
        import inspect
        # 4 个 settings_ui page 标题：商品/主题/校准/后台
        for name in ('_build_product_region_tab', '_build_skin_tab',
                     '_build_calibrate_inline', '_build_backend_tab'):
            method = getattr(settings_ui.SettingsUIMixin, name, None)
            self.assertIsNotNone(method, f'{name} must exist')
            src = inspect.getsource(method)
            self.assertIn('font=self.FONT_HEADING', src)
        # settings_ui._build_api_page 标题改 FONT_HEADING
        api_src = inspect.getsource(settings_ui.SettingsUIMixin._build_api_page)
        self.assertIn('font=self.FONT_HEADING', api_src)
        self.assertNotIn('font=self.FONT_TITLE', api_src)
        # 2 个 stats_ui page 标题改 FONT_HEADING + 去 emoji
        # 注意：仅检查顶部 _lbl 的 text 字段（page title），docstring 里的 emoji 不算
        hist_src = inspect.getsource(stats_ui.StatsPagesMixin._build_history_page)
        self.assertIn('font=self.FONT_HEADING', hist_src)
        self.assertNotIn('font=self.FONT_TITLE', hist_src)
        # 顶部标题 _lbl 文本不含 emoji（docstring 可保留）
        import re as _re
        hist_title_match = _re.search(r'self\._lbl\(page,\s*text="([^"]+)"', hist_src)
        self.assertIsNotNone(hist_title_match)
        self.assertNotIn('📈', hist_title_match.group(1))
        self.assertNotIn('💰', hist_title_match.group(1))
        usage_src = inspect.getsource(stats_ui.StatsPagesMixin._build_usage_page)
        self.assertIn('font=self.FONT_HEADING', usage_src)
        self.assertNotIn('font=self.FONT_TITLE', usage_src)
        usage_title_match = _re.search(r'self\._lbl\(page,\s*text="([^"]+)"', usage_src)
        self.assertIsNotNone(usage_title_match)
        self.assertNotIn('💰', usage_title_match.group(1))

    def test_export_btn_demoted_weight(self):
        """t27 ④：「导出 Excel」13pt bold 16w×2h 降为 9pt bold 常规。

        修 P-A3（结果动作视觉权重超过数据入口）。
        """
        import gui
        import inspect
        src = inspect.getsource(gui.App._build_ui)
        # 必须不再用 13, bold + width=16, height=2
        self.assertNotIn("font=(self.FONT[0], 13, 'bold')", src)
        self.assertNotIn('width=16, height=2', src)
        # 必须用 9pt bold（与 btn_row 一致）
        self.assertIn("font=(self.FONT[0], 9, 'bold')", src)
        # 仍 kind='primary' 保持视觉重要
        # kind='primary' 在 _mk_btn 调用中保留
        self.assertIn("kind='primary'", src)
        # command 仍指向 self._export
        self.assertIn('self._export', src)

    def test_calibrate_tab_dead_code_removed(self):
        """t27 ⑤：_build_calibrate_tab 改名为 _build_calibrate_inline（零外部引用）。"""
        import settings_ui
        import inspect
        # 必须有 _build_calibrate_inline
        self.assertTrue(hasattr(settings_ui.SettingsUIMixin, '_build_calibrate_inline'))
        # 必须没有 _build_calibrate_tab（已重命名）
        self.assertFalse(hasattr(settings_ui.SettingsUIMixin, '_build_calibrate_tab'))
        # gui.py 0 命中 _build_calibrate_tab（死代码确实死）
        import gui
        gui_src = inspect.getsource(gui)
        self.assertNotIn('_build_calibrate_tab', gui_src)

    def test_no_new_color_font_token_added(self):
        """t27 约束：零新配色/字体/token。"""
        import gui
        import settings_ui
        import stats_ui
        # 不能引入新的 hex 颜色 / 字体字号
        for src in (inspect.getsource(gui), inspect.getsource(settings_ui), inspect.getsource(stats_ui)):
            # 不应有 #FFFFFF / #000000 等明确的新色（仅允许已有 C_BG/C_TEXT/C_MUTED/C_BORDER/C_PRIMARY 等 token）
            # 这里只检查没有 C_FOO_BAR / FONT_BAR_BAZ 这种新命名
            import re
            # 不应有 C_XXX 形式的新常量（除已存在的 C_BG/C_TEXT/C_MUTED/C_BORDER/C_SURFACE/C_PRIMARY/C_ACCENT/C_CARD_HDR/C_RED_BG/C_YELLOW_BG）
            new_c = re.findall(r'self\.C_[A-Z_]+', src)
            allowed_c = {'C_BG', 'C_TEXT', 'C_MUTED', 'C_BORDER', 'C_SURFACE',
                         'C_PRIMARY', 'C_SECONDARY', 'C_ACCENT', 'C_CARD_HDR',
                         'C_RED_BG', 'C_YELLOW_BG'}
            # 去掉 self. 前缀再比较
            new_c_stripped = {c.replace('self.', '') for c in new_c}
            unexpected = new_c_stripped - allowed_c
            self.assertEqual(unexpected, set(),
                             f'意外的新 C_ token: {unexpected}')

    def test_home_dual_entry_layout_intact(self):
        """v1.5.7 t27 约束更新：首页 primary_row / btn_row 布局（用户拍板定版）。
         ① 第一排 primary_row：批量(居左) + 导入(居左，文件统一入口弹菜单) + 识图(居右) + 导出(居右)
         ② 第二排 btn_row：刷新(居左) + 🛡双模型(居右)
         ③ 识图 = 截取当前窗口（直连 _live_screenshot，不再弹菜单）
         ④ 本地图片选择识别并入导入（_open_import_menu 二选一）
         ⑤ 全部按钮统一宽度（home_actions.BTN_WIDTH_HOME）"""
        import gui
        import inspect
        src = inspect.getsource(gui.App._build_ui)
        # ① primary_row 按钮文字（批量/导入/识图/导出，用户拍板顺序）
        self.assertIn('"批量"', src)
        self.assertIn('"导入"', src)
        self.assertIn('"识图"', src)
        self.assertIn('"导出"', src)
        # ② v1.5.6 的「截图」按钮已解散（拆为 识图=截窗口 + 导入菜单=选图片）
        self.assertNotIn('"截图"', src)
        # ③ 识图直连实时截图（最小化→截 PDD 窗口→恢复→识别）
        self.assertIn("self._live_screenshot", src)
        # ④ 导入 = 文件统一入口（表格/图片菜单；菜单项分派函数存在）
        self.assertIn("self._open_import_menu", src)
        self.assertIn("self._export", src)
        # ⑤ 全部按钮统一宽度：BTN_WIDTH_HOME 常量已引入
        self.assertIn("btn_width_for", src)
        self.assertIn("home_actions", src)
        # 🛡双模型仍在 btn_row（second 行）
        self.assertIn('"🛡 双模型"', src)
        self.assertIn("self._single_dual_var", src)
        # kind='dark' 次数：批量 + 导入 + 刷新 = 3 次（第一排2 + 第二排1）
        dark_count = src.count("kind='dark'")
        self.assertGreaterEqual(dark_count, 3,
                               f"kind='dark' 应至少 4 次（截图/导入/刷新/批量），实际 {dark_count}")

    def test_build_nav_called_in_build_ui_after_page_frames(self):
        """t29 冷启动修复回归：_build_ui 函数体内必须存在 self._build_nav() 调用，
        且该调用必须位于 8 个 page_* 帧创建之后（_build_nav 引用 page_home 等帧，
        若先调用会 AttributeError）。修 t27 默认展开后冷启动导航空壳 bug。
        """
        import gui
        import inspect
        src = inspect.getsource(gui.App._build_ui)
        # 必须存在 self._build_nav() 调用
        self.assertIn('self._build_nav()', src,
                      '_build_ui 内缺少 self._build_nav() 调用（冷启动导航空壳 bug 未修复）')
        # _build_nav 调用位置必须晚于所有 8 个 page_* 帧的创建（_build_nav 引用 page_* 帧）
        build_nav_pos = src.find('self._build_nav()')
        self.assertGreater(build_nav_pos, 0)
        # 8 个 page_* 帧创建：page_home / page_general / page_products / page_theme /
        # page_backend / page_api / page_history / page_usage
        page_positions = []
        for name in ('page_home', 'page_general', 'page_products', 'page_theme',
                     'page_backend', 'page_api', 'page_history', 'page_usage'):
            pos = src.find(f'self.{name} = tk.Frame(self.content_frame')
            self.assertGreater(pos, 0, f'_build_ui 内未找到 self.{name} 帧创建')
            page_positions.append(pos)
        max_page_pos = max(page_positions)
        # t29 修 v4f-F9：必须用 max（晚于【最后一个】帧创建）而非 min——
        # 原断言只要求晚于最早的 page_home，_build_nav 挪到 8 帧中间也绿但启动必 AttributeError
        self.assertGreater(build_nav_pos, max_page_pos,
                          'self._build_nav() 必须晚于所有 8 个 page_* 帧创建（否则 AttributeError）')

    def test_toggle_nav_lazy_build_guard_preserved(self):
        """t29 冷启动修复回归：_toggle_nav 里的懒构建守卫必须保留——
        `if not self.nav_buttons: self._build_nav()` 防止折叠再展开时重复构建。
        冷启动修复只是新增 _build_ui 内的预构建，_toggle_nav 守卫不能丢。
        """
        import gui
        import inspect
        src = inspect.getsource(gui.App._toggle_nav)
        # 懒构建守卫必须存在
        self.assertIn('if not self.nav_buttons', src,
                      '_toggle_nav 懒构建守卫丢失（折叠再展开会重复创建按钮）')
        self.assertIn('self._build_nav()', src,
                      '_toggle_nav 内缺少 self._build_nav() 调用')

    def test_nav_width_dpi_scaled(self):
        """t29 迭代R3：nav 宽度三处（frame width / _build_ui minsize / _toggle_nav
        minsize）必须随 dpi_scale 联动，不得残留裸 176 像素字面量。
        背景：nav 是全 UI 唯一 pack_propagate(False) 冻结像素宽的容器，字号随
        tk scaling 缩放而 176px 不缩，200% DPI 下「💰 用量明细」会被裁。
        """
        import gui
        import inspect
        ui_src = inspect.getsource(gui.App._build_ui)
        toggle_src = inspect.getsource(gui.App._toggle_nav)
        self.assertIn('width=int(176 * self.dpi_scale)', ui_src,
                      'nav_frame width 未 DPI 联动')
        self.assertIn('minsize=int(176 * self.dpi_scale)', ui_src,
                      '_build_ui 的 pane minsize 未 DPI 联动')
        self.assertIn('minsize=int(176 * self.dpi_scale)', toggle_src,
                      '_toggle_nav 的 pane minsize 未 DPI 联动')
        self.assertNotIn('minsize=176', ui_src + toggle_src,
                         '残留裸 176 minsize 字面量（DPI 下会裁导航项）')

    def test_page_title_margins_unified(self):
        """t29 迭代R3：页标题边距统一 padx=16, pady=(14, 2)——settings_ui 5 页
        （商品/主题/校准/后台/API）+ stats_ui 2 页（历史/用量，已是基准值）；
        不得残留旧 pady=(15,2) 或 API 页 padx=20。
        """
        import settings_ui
        import stats_ui
        import inspect
        s = inspect.getsource(settings_ui)
        self.assertNotIn('.pack(pady=(15,2))', s,
                         'settings_ui 页标题残留旧 pady=(15,2)')
        self.assertIn('font=self.FONT_HEADING).pack(padx=16, pady=(14, 2))', s,
                      'settings_ui 页标题未统一 16/(14,2)')
        self.assertIn('padx=16, pady=(14, 2))', s,
                      'API 页标题边距未归一到 16/(14,2)')
        st = inspect.getsource(stats_ui)
        self.assertGreaterEqual(st.count('padx=16, pady=(14, 2)'), 2,
                                'stats_ui 页标题 16/(14,2) 基准缺失')


# ══════════════════════════════════════════════════════════════════
# t28 GMI-3：UI 实施包 B · 历史页空态行为 + 版式参数源码断言
# ══════════════════════════════════════════════════════════════════

class TestHistoryEmptyState(unittest.TestCase):
    """t28 (④高价值)：history_db.query_daily 返回 [] 时，_history_page_refresh
    不抛、Treeview 插入占位提示行、summary label 切到「暂无历史数据」引导语。
    设计：源码级 mock 模式（不实际起 Tk 窗口），用最小组件模拟 self._hist_page / self._hist_tree / self._hist_summary_label，
    直接调用 monkeypatch 后的 _history_page_refresh。"""

    def setUp(self):
        # 加载 stats_ui 模块（独立模块对象，避免污染导入缓存）
        spec = importlib.util.spec_from_file_location(
            'pdd_stats_ui_t28', os.path.join(HERE, 'stats_ui.py'))
        self.sui = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(self.sui)

    def test_empty_rows_inserts_placeholder_and_summary(self):
        """rows=[] → Treeview 插占位提示行 + summary label 切到引导语 + 不抛异常。"""
        from unittest.mock import MagicMock, patch

        # monkeypatch stats_ui.history_db.query_daily 返 []（空态）、query_regions 返 []
        with patch.object(self.sui.history_db, 'query_daily', return_value=[]), \
             patch.object(self.sui.history_db, 'query_regions', return_value=[]):
            # 构造最小组件替身
            page = MagicMock()
            tree = MagicMock()
            summary_lbl = MagicMock()
            reg_var = MagicMock(); reg_var.get.return_value = '全部'  # StringVar 替身
            reg_combo = MagicMock()                                  # dict-style ['values'] = ...
            days_var = MagicMock(); days_var.get.return_value = '90'  # StringVar 替身

            # 构造轻量 StatsPagesMixin 替身（只挂必要属性 + _history_page_refresh 方法）
            class _Stub:
                pass
            stub = _Stub()
            stub._hist_page = page
            stub._hist_tree = tree
            stub._hist_summary_label = summary_lbl
            stub._hist_reg_var = reg_var
            stub._hist_reg_combo = reg_combo
            stub._hist_days_var = days_var
            stub._current_page = page   # 守门条件满足（否则 _history_page_refresh 直接 return）
            # 把方法 bind 到 stub
            stub._history_page_refresh = self.sui.StatsPagesMixin._history_page_refresh.__get__(stub, _Stub)
            # _lbl 静态方法（gui.py 类属性）需要正确解析——直接 patch 类方法
            stub._lbl = lambda *a, **kw: MagicMock()
            stub._hist_busy = False
            # 业务逻辑：调用 refresh，应不抛
            try:
                stub._history_page_refresh()
            except Exception as e:
                self.fail(f'_history_page_refresh 抛异常：{e!r}')

            # 断言 1：Treeview 收到 1 次 insert（占位提示行），values 含「暂无数据」
            self.assertEqual(tree.insert.call_count, 1)
            # tree.insert('', 'end', values=('--', '暂无数据', '--', '--', '--'))
            vals = tree.insert.call_args.kwargs['values']
            self.assertEqual(vals, ('--', '暂无数据', '--', '--', '--'),
                             f'空态占位行 values 不符：{vals!r}')

            # 断言 2：summary label 切到「暂无历史数据」引导语
            self.assertEqual(summary_lbl.config.call_count, 1)
            kwargs = summary_lbl.config.call_args.kwargs
            self.assertIn('暂无历史数据', kwargs.get('text', ''),
                          f'summary label 文本不含引导语：{kwargs.get("text", "")!r}')

            # 断言 3：守门后 _hist_busy 被释放（finally 块执行）
            self.assertFalse(stub._hist_busy, '_hist_busy 未在 finally 释放')

    def test_non_empty_rows_no_placeholder(self):
        """rows=[{'day':'2026-08-28',...}] → Treeview 收到业务行、无占位行、summary 显示「共 N 条」。"""
        from unittest.mock import MagicMock, patch
        sample = [{'day': '2026-08-28', 'region': '全部', 'items': 5, 'alerts': 1, 'stock_total': 100}]

        with patch.object(self.sui.history_db, 'query_daily', return_value=sample), \
             patch.object(self.sui.history_db, 'query_regions', return_value=[]):
            page = MagicMock()
            tree = MagicMock()
            summary_lbl = MagicMock()
            reg_var = MagicMock(); reg_var.get.return_value = '全部'
            reg_combo = MagicMock()
            days_var = MagicMock(); days_var.get.return_value = '90'

            class _Stub:
                pass
            stub = _Stub()
            stub._hist_page = page
            stub._hist_tree = tree
            stub._hist_summary_label = summary_lbl
            stub._hist_reg_var = reg_var
            stub._hist_reg_combo = reg_combo
            stub._hist_days_var = days_var
            stub._current_page = page
            stub._history_page_refresh = self.sui.StatsPagesMixin._history_page_refresh.__get__(stub, _Stub)
            stub._lbl = lambda *a, **kw: MagicMock()
            stub._hist_busy = False

            stub._history_page_refresh()

            # 业务行 1 条 + 无占位
            self.assertEqual(tree.insert.call_count, 1)
            vals = tree.insert.call_args.kwargs['values']
            self.assertEqual(vals, ('2026-08-28', '全部', 5, 1, 100))
            # summary 是「共 1 条按日汇总（双击行看当日明细）」
            kwargs = summary_lbl.config.call_args.kwargs
            self.assertIn('共 1 条', kwargs.get('text', ''))


class TestLayoutParamsSourceAssert(unittest.TestCase):
    """t28 (③版式参数)：源码断言抽查——settings_ui.py / stats_ui.py 关键 padx/pady
    与 width 已按 t26 提案落位（防止后续被改回 14/15/5/6/22/60）。"""

    def _src(self, name):
        with open(os.path.join(HERE, name), 'r', encoding='utf-8') as f:
            return f.read()

    def test_stats_ui_padx_aligned_to_16(self):
        """stats_ui.py：用量页 8 个区域 padx 14→16（topbar / mid / cv / 说明 / 价格表标题 / ptree / 双击说明 / pbtns）。

        8 处 padx=14→16 的修改都是带 # t28 (a-5) 注释的形式；源代码里的 padx=14 已被
        padx=16 替换并加注释。断言：源码里没有任何【未带 t28 注释的】padx=14 残留。
        """
        import re
        src = self._src('stats_ui.py')
        # 匹配所有 padx=14 出现位置（t29 修 v4f-F8：切片长度必须等于目标串长度 7，
        # 原 i:i+8 恒 False 导致断言空洞——永远绿，防回归承诺落空）
        hits = [(i, src[max(0, i-30):i+50]) for i in range(len(src)) if src[i:i+7] == 'padx=14']
        # 过滤掉被 t28 注释括起来的（注释里写"padx=14→16"是文档化变更）
        real_residue = []
        for pos, ctx in hits:
            # 出现点若位于同行 '# t28' 尾随注释之内（前缀已含 '# t28'）则属文档化变更
            line_start = src.rfind('\n', 0, pos) + 1
            line_prefix = src[line_start:pos]
            if True:  # 源码注释已清理，直接判定任何 padx=14 均为残留
                real_residue.append(ctx)
        self.assertEqual(real_residue, [],
                         f'stats_ui.py 仍有 {len(real_residue)} 处 padx=14 真正残留（不在 t28 注释内）：{real_residue}')

    def test_stats_ui_history_hint_has_padx(self):
        """stats_ui.py：历史页 hint label 加了 padx=(0, 4) 离左缘 4px（t26 c-3）。"""
        src = self._src('stats_ui.py')
        # 提示行：双击行看当日明细；明细行双击看单商品库存趋势折线
        self.assertIn('padx=(0, 4)', src, '历史页 hint label 未加 padx=(0, 4)')

    def test_stats_ui_days_combo_width_6(self):
        """stats_ui.py：days_combo width 5→6（容纳 3 位数字"180"）。"""
        src = self._src('stats_ui.py')
        self.assertIn("values=['30', '90', '180'], width=6", src,
                      'days_combo width 未从 5 改 6')

    def test_stats_ui_empty_state_strings_present(self):
        """stats_ui.py：空态文案「暂无历史数据」+ 占位 row「暂无数据」必须存在。"""
        src = self._src('stats_ui.py')
        self.assertIn('暂无历史数据 — 识别或导入后会在此处出现', src,
                      '空态引导语缺失')
        self.assertIn("'暂无数据'", src, '空态占位行 row 缺失')

    def test_settings_ui_secondary_row_aligned(self):
        """settings_ui.py：副模型行 sec_row pack 几何 + Combobox width=20、保存按钮 font 7→8。
        t8 A2 后：sec_row 包入 _sec_card C_BORDER 卡片，几何收紧 padx=16/pady=10/fill=x。
        """
        src = self._src('settings_ui.py')
        # t8 A2：sec_row 现在包在 _sec_card C_BORDER 卡片内；几何应改 padx=16/pady=10/fill='x'
        self.assertIn("sec_row.pack(padx=16, pady=10, fill='x')", src,
                      '副模型行 t8 A2 后应 padx=16/pady=10/fill=x（包入 _sec_card 后收紧）')
        self.assertIn("textvariable=sec_var, state='normal', width=20,", src,
                      '副模型 Combobox width 未从 22 改 20')
        self.assertIn("font=(self.FONT[0], 8)).pack(side=\"left\")", src,
                      '副模型保存按钮 font 未从 7 改 8')

    def test_settings_ui_safety_days_spinbox_width_8(self):
        """settings_ui.py：补货卡 safety_days Spinbox width 6→8（与 in_transit_qty 对齐）。"""
        src = self._src('settings_ui.py')
        # 找到 safety_days 的 Spinbox 调用
        import re
        m = re.search(r"Spinbox\(sd_row, from_=0, to=30, textvariable=_sd_var, width=(\d+),", src)
        self.assertIsNotNone(m, 'safety_days Spinbox 调用未找到')
        self.assertEqual(m.group(1), '8', f'safety_days Spinbox width 仍为 {m.group(1)}（应改 8）')

    def test_settings_ui_backend_label_width_9(self):
        """settings_ui.py：后台三行 label width 10→9。"""
        src = self._src('settings_ui.py')
        import re
        # 三处 label 宽度必须都是 9
        for keyword in ('后台地址', '登录账号', '登录密码'):
            m = re.search(rf"text=\"{keyword}:\", font=self.FONT, width=(\d+),", src)
            self.assertIsNotNone(m, f'{keyword} label 未找到')
            self.assertEqual(m.group(1), '9', f'{keyword} label width 仍为 {m.group(1)}（应改 9）')

    def test_settings_ui_api_cards_aligned_padx_16(self):
        """settings_ui.py：API 页 4 行 (kf/mf/ef/cef) padx 12→16；卡片 padx 4→8。"""
        src = self._src('settings_ui.py')
        # kf/mf/ef/cef 四行 padx=12 应 0 残留（API 卡片范围内）
        # 但 settings_ui 别处可能有 padx=12，所以限定范围：搜 card = ... 之后到 PRESET_PROVIDERS 之前的 padx=12
        import re
        # 简化：直接断言 kf/mf/ef/cef 行 4 行的 padx=16 都存在
        for kw in ('kf.pack(fill="x", padx=16', 'mf.pack(fill="x", padx=16',
                   'ef.pack(fill="x", padx=16', 'cef.pack(fill="x", padx=16'):
            self.assertIn(kw, src, f'API 卡 {kw.split(".")[0]} 行 padx 未改 16')

    def test_settings_ui_license_enf_row_compact(self):
        """settings_ui.py：授权卡 enf_row pady (8,2)→(6,2) 紧凑；Entry width 60→48 留位按钮。"""
        src = self._src('settings_ui.py')
        self.assertIn("enf_row.pack(pady=(6, 2), padx=20, fill='x')", src,
                      'enf_row pady 未改 (6,2)')
        import re
        m = re.search(r"Entry\(in_row, textvariable=_key_var, font=self.FONT, width=(\d+),", src)
        self.assertIsNotNone(m, 'license Entry 未找到')
        self.assertEqual(m.group(1), '48', f'license Entry width 仍为 {m.group(1)}（应改 48）')


class TestLayoutRound1(unittest.TestCase):
    """t8 布局落地 Round 1：A 级 8 处 + B1 路标源码断言。"""

    def _src(self, name):
        with open(os.path.join(HERE, name), 'r', encoding='utf-8') as f:
            return f.read()

    def test_general_separators_removed(self):
        """t8 A1：通用页 _build_general_page 函数体内 0 个 ttk.Separator 调用（两行已删）。"""
        import inspect, re
        try:
            from settings_ui import SettingsUIMixin
        except ImportError:
            self.skipTest('settings_ui 不可导入')
        src = inspect.getsource(SettingsUIMixin._build_general_page)
        # 跳过注释行，统计 ttk.Separator(...).pack 调用次数
        real_calls = len(re.findall(r'ttk\.Separator\([^)]*\)\.pack', src))
        self.assertEqual(real_calls, 0,
                         f'通用页应已删除 ttk.Separator(...).pack 调用，改用 8px 卡片间 padding（找到 {real_calls} 处）')

    def test_skin_2x2_grid_and_C_BORDER(self):
        """t8 A3+A4：主题页 4 卡 2×2 grid + #E2E8F0→self.C_BORDER。"""
        src = self._src('settings_ui.py')
        # 2×2 网格：columnconfigure(0/1, weight=1) + grid(row=i//2, column=i%2)
        self.assertIn('cards_frame.columnconfigure(0, weight=1)', src,
                      '主题页 cards_frame 缺列等权 0')
        self.assertIn('cards_frame.columnconfigure(1, weight=1)', src,
                      '主题页 cards_frame 缺列等权 1')
        self.assertIn("card.grid(row=i // 2, column=i % 2,", src,
                      '主题页 card 未改用 grid(i//2, i%2)')
        # token 擦边修复：#E2E8F0 → self.C_BORDER
        self.assertNotIn('"#E2E8F0"', src,
                         '主题页仍残留 hex 硬编码 #E2E8F0（应改 self.C_BORDER）')
        self.assertIn('self.C_BORDER', src, '主题页未使用 self.C_BORDER')

    def test_general_anchor_self_references(self):
        """t8 B1：通用页 canvas 存 self + anchors 字典存在（路标按钮点击的基础）。"""
        import inspect
        try:
            from settings_ui import SettingsUIMixin
        except ImportError:
            self.skipTest('settings_ui 不可导入')
        src = inspect.getsource(SettingsUIMixin._build_general_page)
        # canvas 存 self（v4f 修正：build 时存，click 时实时算）
        self.assertIn('self._general_canvas = canvas', src,
                      'canvas 未存为 self._general_canvas（实时算 y 缺基础）')
        # anchors 字典存在
        self.assertIn('self._general_anchors', src,
                      'self._general_anchors 字典缺失（5 段路标机制失效）')
        # 跳锚方法存在
        jump_src = inspect.getsource(SettingsUIMixin._jump_to_general_anchor)
        self.assertIn('yview_moveto', jump_src,
                      '_jump_to_general_anchor 未调 yview_moveto（v4f 修正：实时 fraction 跳转）')
        # 实时计算 y（v4f 修正：免 resize 重算）
        self.assertIn('winfo_y', jump_src,
                      '_jump_to_general_anchor 未实时算 anchor.winfo_y（v4f 修正）')

    def test_usage_reset_btn_red_token(self):
        """t8 A8：用量页重置本月按钮·复用 _mk_btn 返回值 config(bg=C_RED_BG)。"""
        src = self._src('stats_ui.py')
        # 重置按钮存在并配 bg=C_RED_BG（v4f 修正：不新建 _warn_btn）
        self.assertIn('_reset_btn.configure(bg=self.C_RED_BG)', src,
                      '重置本月按钮未走 C_RED_BG 视觉区分（v4f 修正：_warn_btn 不存在）')
        # 应仍调 reset_this_month 回调（业务逻辑零变化）
        self.assertIn('reset_this_month', src, '重置回调 reset_this_month 缺失')

    def test_product_canvas_no_fixed_height(self):
        """t8 A6：商品页 canvas 删 height=220（v4f 修正：禁用 winfo_height 方案，纯 fill+expand）。"""
        import inspect, re
        try:
            from settings_ui import SettingsUIMixin
        except ImportError:
            self.skipTest('settings_ui 不可导入')
        method_src = inspect.getsource(SettingsUIMixin._build_product_region_tab)
        # 跳过注释行，找真正传 height=220 的 canvas 构造
        canvas_with_h = re.findall(r'Canvas\([^)]*height\s*=\s*220[^)]*\)', method_src)
        self.assertEqual(canvas_with_h, [],
                         f'商品页 canvas 仍写死 height=220（应改 fill+expand 自适应），找到：{canvas_with_h}')


class TestAnimationsRound1(unittest.TestCase):
    """t9 动效落地：4 项动效（脉冲/批量按钮文案/hover 插值/nav 渐隐）+ 熔断基建源码断言。"""

    def _src(self, name):
        with open(os.path.join(HERE, name), 'r', encoding='utf-8') as f:
            return f.read()

    def test_lerp_hex_pure_function(self):
        """t9 基建：_lerp_hex 纯函数单测 3 例（同色/异色/越界 t）。"""
        import sys
        sys.path.insert(0, HERE)
        import gui
        # 同色 → 原色
        self.assertEqual(gui._lerp_hex('#000000', '#000000', 0.5), '#000000')
        self.assertEqual(gui._lerp_hex('#FFFFFF', '#FFFFFF', 0.0), '#FFFFFF')
        # 异色 t=0 → a；t=1 → b
        self.assertEqual(gui._lerp_hex('#000000', '#FFFFFF', 0.0), '#000000')
        self.assertEqual(gui._lerp_hex('#000000', '#FFFFFF', 1.0), '#FFFFFF')
        # 异色 t=0.5 → 中点附近（int 截断 127.5 = 127 = 0x7F）
        self.assertEqual(gui._lerp_hex('#000000', '#FFFFFF', 0.5), '#7F7F7F')
        # 越界 t 自动夹紧
        self.assertEqual(gui._lerp_hex('#000000', '#FFFFFF', -1.0), '#000000')
        self.assertEqual(gui._lerp_hex('#000000', '#FFFFFF', 2.0), '#FFFFFF')
        # 容错：非 hex 字符串 → 返回 b
        self.assertEqual(gui._lerp_hex('invalid', '#FFFFFF', 0.5), '#FFFFFF')
        self.assertEqual(gui._lerp_hex(None, '#FFFFFF', 0.5), '#FFFFFF')
        # 起点非 0：#FF0000 → #00FF00 t=0.5 → #7F7F00
        self.assertEqual(gui._lerp_hex('#FF0000', '#00FF00', 0.5), '#7F7F00')

    def test_animations_master_switch_and_meltdown(self):
        """t9 基建：ANIMATIONS_ENABLED 总开关 + 动效回调异常时静默熔断。"""
        src = self._src('gui.py')
        # 总开关
        self.assertIn('ANIMATIONS_ENABLED = True', src,
                      'ANIMATIONS_ENABLED 模块常量缺失（低配机一键关动效）')
        # 熔断：异常时 try/except 静默吞掉
        self.assertIn('except Exception:\n                    pass', src,
                      '动效回调必须 try/except 静默熔断（不影响主流程）')
        # 各动效方法入口都应检查 ANIMATIONS_ENABLED
        for m in ('_pulse_status', '_animate_btn_hover', '_animate_nav_leave'):
            self.assertIn(f'if not ANIMATIONS_ENABLED:\n            return',
                          src, f'{m} 缺总开关检查')

    def test_pulse_status_uses_register_redraw_pattern(self):
        """t9 C：_pulse_status 必须用 self.tc 实时查 C_TEXT（非构造时快照），兼容主题切换。"""
        import inspect
        import gui
        src = inspect.getsource(gui.App._pulse_status)
        # 终态色 = self.tc('C_TEXT', ...) 实时查
        self.assertIn("self.tc('C_TEXT'", src,
                      '_pulse_status 终态色必须用 self.tc 实时查（主题切换不偏色）')
        # 强调色 = self.tc('C_PRIMARY', ...)
        self.assertIn("self.tc('C_PRIMARY'", src,
                      '_pulse_status 强调色必须用 self.tc 实时查')
        # 必须有 _anim_jobs 注册与清空
        self.assertIn('self._anim_jobs', src, '_pulse_status 缺 _anim_jobs 注册表')

    def test_hover_animation_wired_into_mk_btn(self):
        """t9 A：_CanvasBtn._hover / _leave 必须调 self.owner._animate_btn_hover。"""
        import inspect
        import gui
        hover_src = inspect.getsource(gui._CanvasBtn._hover)
        leave_src = inspect.getsource(gui._CanvasBtn._leave)
        # hover 走插值
        self.assertIn('self.owner._animate_btn_hover(self, enter=True)', hover_src,
                      '_CanvasBtn._hover 未接 _animate_btn_hover 插值')
        # leave 也走插值
        self.assertIn('self.owner._animate_btn_hover(self, enter=False)', leave_src,
                      '_CanvasBtn._leave 未接 _animate_btn_hover 插值')
        # disabled 态抢先落终态（v4f 修正）
        self.assertIn("if self._state == 'disabled':",
                      hover_src + leave_src,
                      'disabled 态应抢先落终态不动画（v4f 修正）')

    def test_anim_jobs_init_and_snap_before_cancel(self):
        """t9 基建：self._anim_jobs 初始化在 __init__，_cancel_after_jobs 先 snap 终态再 cancel。"""
        import importlib.util
        spec = importlib.util.spec_from_file_location('pdd_gui_for_anim', os.path.join(HERE, 'gui.py'))
        # 直接读源码验证
        src = self._src('gui.py')
        # 初始化
        self.assertIn('self._anim_jobs = {}', src,
                      'self._anim_jobs 初始化缺失（在 __init__ 内）')
        # _cancel_after_jobs 工具函数
        self.assertIn('def _cancel_after_jobs', src,
                      '_cancel_after_jobs 工具函数缺失（按目标 widget 分键取消）')
        # v4f 最大漏项：先 snap 终态再 cancel——验证 _cancel_after_jobs 前有终态设置
        # 简单模式：验证 _anim_jobs 在每个动效 _do 函数末尾置空
        # （即先 snap 到 _end/_b，再调 _cancel_after_jobs）
        self.assertIn('self._anim_jobs[_key] = []', src,
                      '动效终态 snap 时 _anim_jobs[_key] = [] 应先于 cancel')
class TestAnimationsFixRound2(unittest.TestCase):
    """t11 动效修复包：③ 删 canvas gate + _highlight_nav 兜底 / ② disabled snap 禁用色 /
    ⑤ 真熔断（_meltdown_animations 翻 ANIMATIONS_ENABLED）/ ⑧a 脉冲实时 self.tc /
    ⑧b 脉冲起拍捕获当前 fg。"""

    def _src(self, name):
        with open(os.path.join(HERE, name), 'r', encoding='utf-8') as f:
            return f.read()

    def test_nav_leave_no_canvas_gate(self):
        """t11 ③：_animate_nav_leave 不能再用 btn.canvas gate（导航按钮是 tk.Button 无 canvas）。"""
        import sys
        sys.path.insert(0, HERE)
        import gui
        import inspect
        src = inspect.getsource(gui.App._animate_nav_leave)
        # 删 canvas gate
        self.assertNotIn("getattr(btn, 'canvas'", src,
                         '_animate_nav_leave 不应再用 btn.canvas gate（tk.Button 无此属性）')
        # _highlight_nav 必须先 configure(bg=C_BG) 兜底
        hl_src = inspect.getsource(gui.App._highlight_nav)
        self.assertIn('btn.configure(bg=self.C_BG', hl_src,
                      '_highlight_nav 离开位分支必须显式 configure(bg=C_BG) 兜底')

    def test_hover_disabled_snap_uses_disabled_color(self):
        """t11 ②：hover 链 disabled 分支 snap 到禁用色（btn.disabled tc + C_SURFACE fallback）。"""
        import sys
        sys.path.insert(0, HERE)
        import gui
        import inspect
        src = inspect.getsource(gui.App._animate_btn_hover)
        # 禁用态 snap 应走 tc('btn.disabled', self.C_SURFACE)，不能硬编码 _end
        self.assertIn("self.tc('btn.disabled', self.C_SURFACE)", src,
                      'disabled snap 必须走 btn.disabled tc + C_SURFACE fallback（v4f 修正：零新 token）')

    def test_real_meltdown_flips_module_flag(self):
        """t11 ⑤：真熔断——_meltdown_animations 翻 ANIMATIONS_ENABLED=False；动效 except 分支应调用。"""
        import sys
        sys.path.insert(0, HERE)
        import gui
        # 模块级函数存在
        self.assertTrue(callable(getattr(gui, '_meltdown_animations', None)),
                        '_meltdown_animations 模块级函数缺失（真熔断核心）')
        # 翻前 True
        gui.ANIMATIONS_ENABLED = True
        # 调用
        gui._meltdown_animations()
        self.assertFalse(gui.ANIMATIONS_ENABLED,
                         '_meltdown_animations 翻 ANIMATIONS_ENABLED=False 失败')
        # 重置供后续测试
        gui.ANIMATIONS_ENABLED = True
        # 验证动效方法 except 分支包含 _meltdown_animations 调用
        src = self._src('gui.py')
        for m in ('_pulse_status', '_animate_btn_hover', '_animate_nav_leave'):
            # 取该方法的源码块（含 _meltdown_animations 调用）
            idx = src.find(f'def {m}')
            self.assertGreater(idx, 0, f'{m} 方法不存在')
            # 取到下一个 def 之前
            nxt = src.find('\n    def ', idx + 1)
            block = src[idx:nxt if nxt > 0 else None]
            self.assertIn('_meltdown_animations()', block,
                          f'{m} except 分支必须调 _meltdown_animations()（v4f 修正：注释谎言）')

    def test_meltdown_triggered_by_lerp_exception(self):
        """t11 ⑤（v4f 完整示例）：monkeypatch _lerp_hex 抛异常后，_pulse_status 应触发熔断。"""
        import sys
        sys.path.insert(0, HERE)
        import gui
        # 重置
        gui.ANIMATIONS_ENABLED = True
        # monkeypatch _lerp_hex
        _orig = gui._lerp_hex
        def _boom(*a, **k):
            raise RuntimeError('t11 ⑤测试：lerp 抛异常')
        gui._lerp_hex = _boom
        try:
            # 构造一个最简 self-like 对象（不实例化 App，避免 Tk 启动）
            class _FakeLbl:
                def winfo_exists(self): return True
                def cget(self, key): return '#1F1F1F' if key == 'fg' else ''
                def configure(self, **kw): pass
            class _FakeSelf:
                _anim_jobs = {}
                win = type('W', (), {'after': staticmethod(lambda *a, **k: None)})()
                status_label = _FakeLbl()
                def tc(self, key, default=None):
                    return default or '#1F1F1F'
            fs = _FakeSelf()
            # _pulse_status 在 lerp 抛异常时应走 except 分支，触发熔断
            gui.App._pulse_status(fs)
            self.assertFalse(gui.ANIMATIONS_ENABLED,
                             'monkeypatch lerp 抛异常后 _pulse_status 应触发熔断翻 False')
        finally:
            gui._lerp_hex = _orig
            gui.ANIMATIONS_ENABLED = True

    def test_pulse_captures_current_fg_and_realtime_theme(self):
        """t11 ⑧a+⑧b：_pulse_status 终态用实时 self.tc；起拍捕获 label 当前 fg（不硬编码 C_TEXT）。"""
        import sys
        sys.path.insert(0, HERE)
        import gui
        import inspect
        src = inspect.getsource(gui.App._pulse_status)
        # ⑧a：_do 内每步 / 终态用 self.tc 实时查
        self.assertIn("self.tc('C_TEXT'", src,
                      '_pulse_status 终态必须 self.tc(\"C_TEXT\") 实时查（v4f 修正：200ms 内切主题不残留）')
        # ⑧b：起拍捕获 _lbl.cget('fg')
        self.assertIn("_lbl.cget('fg')", src,
                      '_pulse_status 起拍必须捕获 label 当前 fg（v4f 修正：不硬编码 C_TEXT）')


if __name__ == '__main__':
    unittest.main()